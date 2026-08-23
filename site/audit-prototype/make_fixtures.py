#!/home/jon/venv/bin/python
"""Generate .xlsx test fixtures for the Migration Audit prototype.

Run:  /home/jon/venv/bin/python make_fixtures.py

Fixtures:
  basic-2sheet.xlsx   - openpyxl; 2 sheets, 10 formulas (VLOOKUP/SUMIFS/TEXTJOIN...),
                        strings containing "&" and embedded quotes.
  shared-formula.xlsx - HAND-CRAFTED XML zipped manually. openpyxl always writes
                        one plain <f> per cell and cannot emit t="shared" groups,
                        so we build the minimal OOXML package ourselves to get a
                        real Excel-style shared formula (master with si+ref, and
                        members carrying only si).
  array-formula.xlsx  - openpyxl ArrayFormula (writes <f t="array" ref="...">).
"""
import os
import zipfile

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(os.path.abspath(__file__))


def basic_2sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)          # A1:A5 = 1..5
        ws.cell(row=i, column=2, value=i * 10)     # B1:B5 = 10..50
    ws["C1"] = 'Tom & Jerry "quoted"'
    ws["C2"] = "Fish & <Chips>"
    # 5 formulas on Data
    ws["D1"] = "=VLOOKUP(A1,$A$1:$B$5,2,FALSE)"
    ws["D2"] = '=SUMIFS($B$1:$B$5,$A$1:$A$5,">2")'
    ws["D3"] = '=TEXTJOIN(", ",TRUE,C1:C2)'
    ws["D4"] = '=IF(A1>2,"a & b","x ""q"" y")'
    ws["D5"] = "=SUM(A1:A5)+MAX(B1:B5)"
    # 5 formulas on Summary
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "=COUNT(Data!A1:A5)"
    ws2["A2"] = "=AVERAGE(Data!B1:B5)"
    ws2["A3"] = '=CONCATENATE("Total: ",SUM(Data!A1:A5))'
    ws2["A4"] = "=ROUND(SUM(Data!B1:B5)/3,2)"
    ws2["A5"] = '=IFERROR(VLOOKUP(99,Data!A1:B5,2,0),"missing & none")'
    wb.save(os.path.join(HERE, "basic-2sheet.xlsx"))


SHARED_SHEET_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheetData>
<row r="1"><c r="A1"><v>1</v></c><c r="B1"><f t="shared" ref="B1:B5" si="0">A1*2+$A$1</f><v>3</v></c><c r="C1"><f t="shared" ref="C1:C3" si="1">SUM($A$1:A1)</f><v>1</v></c></row>
<row r="2"><c r="A2"><v>2</v></c><c r="B2"><f t="shared" si="0"/><v>5</v></c><c r="C2"><f t="shared" si="1"/><v>3</v></c></row>
<row r="3"><c r="A3"><v>3</v></c><c r="B3"><f t="shared" si="0"/><v>7</v></c><c r="C3"><f t="shared" si="1"/><v>6</v></c></row>
<row r="4"><c r="A4"><v>4</v></c><c r="B4"><f t="shared" si="0"/><v>9</v></c></row>
<row r="5"><c r="A5"><v>5</v></c><c r="B5"><f t="shared" si="0"/><v>11</v></c></row>
</sheetData>
</worksheet>
"""

SHARED_WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="SharedDemo" sheetId="1" r:id="rId1"/></sheets>
</workbook>
"""

SHARED_WORKBOOK_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1"
 Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
 Target="worksheets/sheet1.xml"/>
</Relationships>
"""

SHARED_ROOT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1"
 Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
 Target="xl/workbook.xml"/>
</Relationships>
"""

SHARED_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>
"""


def shared_formula():
    path = os.path.join(HERE, "shared-formula.xlsx")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", SHARED_CONTENT_TYPES)
        z.writestr("_rels/.rels", SHARED_ROOT_RELS)
        z.writestr("xl/workbook.xml", SHARED_WORKBOOK_XML)
        z.writestr("xl/_rels/workbook.xml.rels", SHARED_WORKBOOK_RELS)
        z.writestr("xl/worksheets/sheet1.xml", SHARED_SHEET_XML)


def array_formula():
    wb = Workbook()
    ws = wb.active
    ws.title = "Arrays"
    for i in range(1, 4):
        ws.cell(row=i, column=1, value=i)        # A1:A3
        ws.cell(row=i, column=2, value=i * 100)  # B1:B3
    ws["D1"] = ArrayFormula("D1", "=SUM(A1:A3*B1:B3)")
    ws["E1"] = ArrayFormula("E1:E3", "=A1:A3*2")
    ws["F1"] = "=SUM(A1:A3)"  # one normal formula alongside
    wb.save(os.path.join(HERE, "array-formula.xlsx"))


if __name__ == "__main__":
    basic_2sheet()
    shared_formula()
    array_formula()
    # sanity: all three must round-trip through openpyxl's reader
    from openpyxl import load_workbook
    for name in ("basic-2sheet.xlsx", "shared-formula.xlsx", "array-formula.xlsx"):
        wb = load_workbook(os.path.join(HERE, name))
        print(name, "->", wb.sheetnames)
    print("fixtures written to", HERE)
