#!/home/jon/venv/bin/python
"""Generate the E2E verdict fixture for the Migration Audit page.

Run:  /home/jon/venv/bin/python make_fixtures.py

verdict-mix.xlsx — 2 sheets, 12 formulas mixing safe functions with real,
dataset-verified breakers so test.mjs can assert exact classifications per
migration target. Function choices (checked against docs/data/compat.json,
snapshot 2026-08-23):

  SUM, VLOOKUP, IF   x&g&l documented; lv=quirky for SUM/VLOOKUP, supported for IF
  AGGREGATE          Excel-only vs Sheets (g:false); LibreOffice lv=supported
  GROUPBY            Excel-only (g:false); LibreOffice lv=unsupported (executed #NAME?)
  WEBSERVICE         Excel-only (g:false); LibreOffice documented, lv=null (doc-only)
                     NB: this slot exists purely to exercise the DOCUMENTED-ONLY
                     verdict basis, so it must name a function the corpus has not
                     executed yet. It was BAHTTEXT until 2026-08-31 (executed by
                     the alphabetical corpus push, batch A), then ODDLYIELD until
                     later the same day, when batch F executed the whole ODD*
                     family and its lv stopped being null.
                     WEBSERVICE IS THE LAST CANDIDATE IN THE DATASET. Of the
                     ~480 functions, exactly five were ever both x && l && !g and
                     un-executed -- ODDFPRICE, ODDFYIELD, ODDLPRICE, ODDLYIELD
                     and WEBSERVICE -- and batch F executed the first four.
                     THE PUSH HAS NOW REACHED W AND THIS SLOT IS SETTLED: batch G
                     (RTD..ZTEST) closed the Excel-documented set and DELIBERATELY
                     DID NOT EXECUTE WEBSERVICE, for the same class of reason it
                     skips CALL and REGISTER.ID -- WEBSERVICE's documented
                     behaviour IS a network fetch, so any value this harness sees
                     describes the sandbox rather than the engine. (Concretely:
                     _xlfn.WEBSERVICE parses and evaluates in all four pinned
                     LibreOffice builds and returns #N/A for an unreachable host,
                     which classify_verdict() would publish as "quirky" -- a
                     LibreOffice defect that does not exist.) So lv stays null,
                     WEBSERVICE stays doc-only, and this slot keeps working
                     indefinitely.
                     IF THAT EVER CHANGES -- if a later batch decides to execute
                     WEBSERVICE after all -- there is no real function left to
                     point at, and the doc-only verdict basis must then be
                     exercised with a SYNTHETIC dataset entry in test.mjs rather
                     than with a fixture formula over a real name.
                     Note WEBSERVICE sorts AFTER TEXTSPLIT, where the previous two
                     occupants sorted before it, so the at-risk ordering and
                     free-tier assertions in test.mjs name it last; that is the
                     one thing to re-check when this slot moves again.
  TEXTSPLIT          g:false; LibreOffice lv=supported, lnew=25.8.7.3
  FILTER             x&g documented; LibreOffice lv=quirky
  GOOGLEFINANCE      Sheets-only (x:false); l:false, lv=null
  ARRAYFORMULA       Sheets-only (x:false); l:false, lv=null
  NOTAREALFUNCTION   absent from the dataset -> UNKNOWN

The formulas are never executed by any test — only their stored text is
parsed — so writing Sheets-only functions into an .xlsx is fine.
"""
import os

from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))


def verdict_mix():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)          # A1:A5
        ws.cell(row=i, column=2, value=i * 10)     # B1:B5
    ws["C1"] = "=SUM(A1:A5)"
    ws["C2"] = "=VLOOKUP(A1,$A$1:$B$5,2,FALSE)"
    ws["C3"] = "=AGGREGATE(9,6,A1:A5)"
    ws["C4"] = "=GROUPBY(A1:A5,B1:B5,SUM)"
    ws["C5"] = "=FILTER(A1:A5,B1:B5>10)"
    ws["C6"] = '=TEXTSPLIT("a,b",",")'
    ws["C7"] = '=WEBSERVICE("https://example.com/rates.xml")' 

    ws2 = wb.create_sheet("Mix")
    ws2["A1"] = '=GOOGLEFINANCE("GOOG")'
    ws2["A2"] = "=ARRAYFORMULA(Calc!A1:A5*2)"
    ws2["A3"] = "=NOTAREALFUNCTION(A1)"
    ws2["A4"] = "=IF(SUM(A1)>0,GROUPBY(Calc!A1:A5,Calc!B1:B5,SUM),0)"
    ws2["A5"] = "=A1+A2"  # formula with no functions at all

    wb.save(os.path.join(HERE, "verdict-mix.xlsx"))


if __name__ == "__main__":
    verdict_mix()
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(HERE, "verdict-mix.xlsx"))
    print("verdict-mix.xlsx ->", wb.sheetnames)
    print("written to", HERE)
