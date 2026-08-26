#!/usr/bin/env python3
"""
End-to-end tests for scripts/xlsx_recalc_diff.py.

No pytest dependency -- plain unittest. These tests REALLY RUN LibreOffice
(`soffice`), because the whole point of the tool is that the LibreOffice
side is executed rather than assumed.

We cannot run Excel, so the "Excel-saved" side of every fixture is
simulated the honest way: build the workbook with openpyxl, then INJECT a
cached <v> value into the formula cells at the XML level, exactly where
Excel would have written its own computed result. The injected values are
the ones Microsoft documents (cited per fixture); the LibreOffice values
are whatever LibreOffice actually computes right here, right now.

    /home/jon/venv/bin/python scripts/test_xlsx_recalc_diff.py
    python3 scripts/test_xlsx_recalc_diff.py -v
"""
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import openpyxl  # noqa: E402
import xlsx_recalc_diff as X  # noqa: E402


# --------------------------------------------------------------------------
# Fixture helpers: build with openpyxl, then inject Excel-style cached values
# --------------------------------------------------------------------------

def _cell_re(addr):
    return re.compile((r'<c r="%s"[^>]*>.*?</c>' % re.escape(addr)).encode(), re.DOTALL)


def inject_cached(path, values, calc_mode=None, keep_full_calc_on_load=False):
    """Simulate an Excel save: write a cached <v> into the given formula cells.

    values: {sheet_name: {addr: value}} where value is a number, str, bool,
            or an error string like '#VALUE!'.
    Also strips openpyxl's fullCalcOnLoad="1" (Excel does not set it) and can
    set calcMode="manual".
    """
    with zipfile.ZipFile(path) as zin:
        parts = [(i, zin.read(i.filename)) for i in zin.infolist()]
        sheet_part = dict((n, p) for n, p in X.sheet_parts(zin))

    out = []
    for info, data in parts:
        target_sheet = None
        for sname, pname in sheet_part.items():
            if pname == info.filename and sname in values:
                target_sheet = sname
        if target_sheet:
            for addr, val in values[target_sheet].items():
                if isinstance(val, bool):
                    tattr, text = b' t="b"', b"1" if val else b"0"
                elif isinstance(val, str) and val in X.ERROR_STRINGS:
                    tattr, text = b' t="e"', val.encode()
                elif isinstance(val, str):
                    tattr, text = b' t="str"', val.encode()
                else:
                    tattr, text = b"", repr(val).encode()

                def repl(m, tattr=tattr, text=text):
                    cell = m.group(0)
                    open_tag = cell[: cell.find(b">") + 1]
                    rest = cell[len(open_tag):]
                    open_tag = re.sub(rb'\s+t="[^"]*"', b"", open_tag)
                    open_tag = open_tag[:-1] + tattr + b">"
                    rest = X.V_RE.sub(b"", rest)
                    rest = rest.replace(b"</c>", b"<v>" + text + b"</v></c>")
                    return open_tag + rest

                new, n = _cell_re(addr).subn(repl, data)
                if n != 1:
                    raise AssertionError("could not inject into %s!%s (n=%d)"
                                         % (target_sheet, addr, n))
                data = new
        elif info.filename == "xl/workbook.xml":
            if not keep_full_calc_on_load:
                data = re.sub(rb'\s+fullCalcOnLoad="[^"]*"', b"", data)
            if calc_mode:
                data = re.sub(rb"<calcPr\b", b'<calcPr calcMode="%s"' % calc_mode.encode(), data)
        out.append((info, data))

    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in out:
            zi = zipfile.ZipInfo(info.filename, date_time=info.date_time)
            zi.compress_type = info.compress_type
            zout.writestr(zi, data)
    os.replace(tmp, path)
    return path


def run_tool(path, extra=None):
    """Run the CLI in-process; returns (exit_code, report_dict_or_None)."""
    with tempfile.TemporaryDirectory() as td:
        jout = os.path.join(td, "r.json")
        argv = [path, "--json", jout, "--quiet"] + (extra or [])
        buf = _Capture()
        with buf:
            code = X.main(argv)
        report = None
        if os.path.exists(jout):
            with open(jout) as f:
                report = json.load(f)
        return code, report, buf.text


class _Capture:
    def __enter__(self):
        import io
        self._old = sys.stdout
        self._buf = io.StringIO()
        sys.stdout = self._buf
        return self

    def __exit__(self, *a):
        sys.stdout = self._old
        self.text = self._buf.getvalue()
        return False


def cat_of(report, address):
    for r in report["cells"]:
        if r["address"] == address:
            return r
    raise AssertionError("cell %s not in report" % address)


HAVE_SOFFICE = shutil.which(X.SOFFICE_BIN) is not None or os.path.isfile(X.SOFFICE_BIN)


# --------------------------------------------------------------------------
# Pure-function tests (no LibreOffice needed)
# --------------------------------------------------------------------------

class TestNormalization(unittest.TestCase):
    def test_unescape_ooxml(self):
        self.assertEqual(X.unescape_ooxml("_x0000_"), "\x00")
        self.assertEqual(X.unescape_ooxml("a_x000A_b"), "a\nb")
        self.assertEqual(X.unescape_ooxml("plain"), "plain")
        # An escaped underscore means the file literally contained "_x0000_"
        self.assertEqual(X.unescape_ooxml("_x005F_x0000_"), "_x005F_x0000_")

    def test_classify_numeric_tolerance(self):
        self.assertEqual(X.classify(1.0, 1.0 + 1e-15)[0], "match")
        self.assertEqual(X.classify(1.0, 1.0000001)[0], "numeric_mismatch")

    def test_classify_error_families(self):
        self.assertEqual(X.classify("#NUM!", "#VALUE!")[0], "error_vs_error")
        self.assertEqual(X.classify("#VALUE!", "#VALUE!")[0], "match")
        self.assertEqual(X.classify(5, "#VALUE!")[0], "value_vs_error")
        self.assertEqual(X.classify("#VALUE!", 5)[0], "error_vs_value")
        self.assertEqual(X.classify(5, None)[0], "missing_in_lo")

    def test_classify_types(self):
        self.assertEqual(X.classify(True, True)[0], "match")
        self.assertEqual(X.classify(True, 1)[0], "type_mismatch")
        self.assertEqual(X.classify("a", "b")[0], "text_mismatch")
        self.assertEqual(X.classify("", None)[0], "match")

    def test_classify_dates_as_serials(self):
        import datetime
        self.assertEqual(X.classify(45000, datetime.datetime(2023, 3, 15))[0], "match")

    def test_function_extraction(self):
        self.assertEqual(X.functions_in("=_xlfn.UNICHAR(SUM(A1:A2))"), ["SUM", "UNICHAR"])
        self.assertTrue(X.is_nondeterministic("=NOW()+1"))
        self.assertFalse(X.is_nondeterministic("=SUM(A1:A2)"))


# --------------------------------------------------------------------------
# XML stripping tests (no LibreOffice needed)
# --------------------------------------------------------------------------

class TestStripping(unittest.TestCase):
    def setUp(self):
        self.td = tempfile.mkdtemp(prefix="recalcdiff_strip_")
        self.path = os.path.join(self.td, "strip.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 2
        ws["A2"] = 3
        ws["A3"] = "text literal"
        ws["A4"] = True
        ws["B1"] = "=A1+A2"
        ws["B1"].number_format = "0.00"
        wb.save(self.path)
        inject_cached(self.path, {"S": {"B1": 5}})

    def tearDown(self):
        shutil.rmtree(self.td, ignore_errors=True)

    def test_strip_removes_only_formula_cached_values(self):
        out = os.path.join(self.td, "stripped.xlsx")
        stats = X.make_stripped_copy(self.path, out)
        self.assertEqual(stats["sheets"], 1)
        self.assertGreaterEqual(stats["values_removed"], 1)

        with zipfile.ZipFile(out) as z:
            sheet = z.read("xl/worksheets/sheet1.xml").decode()
        # formula survives, its cached value does not
        self.assertIn("<f>A1+A2</f>", sheet)
        self.assertNotIn("<v>5</v>", sheet)
        # every non-formula literal cell keeps its value
        self.assertIn('<c r="A1" t="n"><v>2</v></c>', sheet)
        self.assertIn('<c r="A2" t="n"><v>3</v></c>', sheet)
        self.assertIn("text literal", sheet)
        self.assertIn('r="A4"', sheet)

        # openpyxl still reads the literals back correctly
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["S"]
        self.assertEqual(ws["A1"].value, 2)
        self.assertEqual(ws["A2"].value, 3)
        self.assertEqual(ws["A3"].value, "text literal")
        self.assertIs(ws["A4"].value, True)
        self.assertIsNone(ws["B1"].value)  # cached value gone

    def test_strip_preserves_other_parts_byte_for_byte(self):
        out = os.path.join(self.td, "stripped2.xlsx")
        X.make_stripped_copy(self.path, out)
        with zipfile.ZipFile(self.path) as a, zipfile.ZipFile(out) as b:
            self.assertEqual(a.namelist(), b.namelist())
            for name in a.namelist():
                if name.startswith("xl/worksheets/") or name == "xl/workbook.xml":
                    continue
                self.assertEqual(a.read(name), b.read(name),
                                 "part %s was modified by stripping" % name)
        # styles must survive (the openpyxl round-trip alternative loses
        # charts/drawings entirely; this is why we work at the zip level).
        # B1 carries number format "0.00" == builtin numFmtId 2.
        with zipfile.ZipFile(out) as b:
            self.assertIn("xl/styles.xml", b.namelist())
            self.assertIn('numFmtId="2"', b.read("xl/styles.xml").decode())


class TestSharedFormulaXmlScan(unittest.TestCase):
    def test_shared_slaves_are_found_and_resolved(self):
        xml = (b'<worksheet><sheetData><row r="1">'
               b'<c r="A1" t="n"><v>1</v></c>'
               b'<c r="B1"><f t="shared" ref="B1:B3" si="0">A1*2</f><v>2</v></c>'
               b'</row><row r="2">'
               b'<c r="A2" t="n"><v>2</v></c>'
               b'<c r="B2"><f t="shared" si="0"/><v>4</v></c>'
               b'</row><row r="3">'
               b'<c r="A3" t="n"><v>3</v></c>'
               b'<c r="B3"><f t="shared" si="0"/><v>6</v></c>'
               b'</row></sheetData></worksheet>')
        cells = X.scan_sheet_xml(xml)["cells"]
        self.assertEqual(sorted(cells), ["B1", "B2", "B3"])  # NOT A1/A2/A3
        self.assertEqual(cells["B2"]["formula"], "=A1*2")
        self.assertEqual(cells["B2"]["shared_master"], "B1")
        self.assertTrue(all(c["has_cached"] for c in cells.values()))
        stripped = X.strip_sheet_xml(xml, [])
        self.assertNotIn(b"<v>4</v>", stripped)
        self.assertIn(b'<c r="A2" t="n"><v>2</v></c>', stripped)


class TestCalcPr(unittest.TestCase):
    def test_manual_and_fullcalc_detection(self):
        td = tempfile.mkdtemp(prefix="recalcdiff_calcpr_")
        try:
            p = os.path.join(td, "m.xlsx")
            wb = openpyxl.Workbook()
            wb.active["A1"] = "=1+1"
            wb.save(p)
            inject_cached(p, {"Sheet": {"A1": 2}}, calc_mode="manual")
            with zipfile.ZipFile(p) as z:
                calc = X.read_calc_pr(z)
            self.assertEqual(calc["calc_mode"], "manual")
            self.assertFalse(calc["full_calc_on_load"])
        finally:
            shutil.rmtree(td, ignore_errors=True)


# --------------------------------------------------------------------------
# End-to-end tests -- these really execute LibreOffice
# --------------------------------------------------------------------------

@unittest.skipUnless(HAVE_SOFFICE, "soffice not available")
class TestEndToEnd(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.td = tempfile.mkdtemp(prefix="recalcdiff_e2e_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.td, ignore_errors=True)

    def _wb(self, name):
        return os.path.join(self.td, name)

    # --- 1. Injected value equals what LibreOffice computes -> match ------
    def test_matching_cached_value(self):
        p = self._wb("match.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"], ws["A2"] = 2, 3
        ws["B1"] = "=A1+A2"
        ws["B2"] = '=CONCATENATE("a","b")'
        wb.save(p)
        inject_cached(p, {"S": {"B1": 5, "B2": "ab"}})

        code, rep, _ = run_tool(p)
        self.assertTrue(rep["trusted"], rep["warnings"])
        self.assertEqual(cat_of(rep, "S!B1")["category"], "match")
        self.assertEqual(cat_of(rep, "S!B2")["category"], "match")
        self.assertEqual(rep["differing_cells_total"], 0)
        self.assertEqual(code, 0, "all-match workbook must exit 0")

    # --- 2. COUNT + boolean: Excel 1, LibreOffice 2 ----------------------
    def test_count_boolean_divergence(self):
        """=COUNT(A1,1) with A1 holding the boolean TRUE.

        Microsoft's COUNT documentation
        (https://support.microsoft.com/en-us/office/count-function-a59cd7fc-b623-4d93-87a4-d23bf411294c):
        logical values inside a *cell reference* are NOT counted, so Excel
        computes 1 (only the literal 1). LibreOffice treats booleans as
        numbers, so it counts both -> 2. This is the documented divergence
        family recorded in results/libreoffice-25.8.json under
        COUNT_boolean_in_range_excluded (expected 0, LO computed 1).
        """
        p = self._wb("count_bool.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = True
        ws["B1"] = "=COUNT(A1,1)"
        wb.save(p)
        inject_cached(p, {"S": {"B1": 1}})   # Excel's documented result

        code, rep, _ = run_tool(p)
        self.assertTrue(rep["trusted"], rep["warnings"])
        row = cat_of(rep, "S!B1")
        self.assertEqual(row["category"], "numeric_mismatch")
        self.assertEqual(row["excel_value"], 1)
        self.assertEqual(row["libreoffice_value"], 2,
                         "LibreOffice must actually compute 2 here (booleans "
                         "count as numbers); got %r" % (row["libreoffice_value"],))
        self.assertEqual(code, 1)
        self.assertIn("COUNT", [f["function"] for f in rep["top_functions_in_differing_cells"]])

    def test_count_boolean_matches_committed_lo_results(self):
        """The live LO behaviour above must agree with the committed corpus."""
        with open(os.path.join(REPO, "results", "libreoffice-25.8.json")) as f:
            res = json.load(f)
        case = res["function_results"]["COUNT"]["COUNT_boolean_in_range_excluded"]
        self.assertEqual(case["formula_display"], "=COUNT(A1:A1)")
        self.assertEqual(case["expected"], 0, "Excel-documented value")
        self.assertEqual(case["value"], 1, "LibreOffice counts the boolean cell")
        self.assertFalse(case["matched_expected"])

    # --- 3. Error vs value: CHAR(0) --------------------------------------
    def test_error_vs_value_char_zero(self):
        """=CHAR(0): Excel documents #VALUE! (valid range is 1-255); LO
        returns an actual NUL character, stored as the _x0000_ escape.
        Cross-checked against results/libreoffice-25.8.json CHAR_out_of_range_zero."""
        p = self._wb("char0.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "=CHAR(0)"
        wb.save(p)
        inject_cached(p, {"S": {"A1": "#VALUE!"}})

        code, rep, _ = run_tool(p)
        self.assertTrue(rep["trusted"], rep["warnings"])
        row = cat_of(rep, "S!A1")
        self.assertEqual(row["category"], "error_vs_value")
        self.assertEqual(row["excel_value"], "#VALUE!")
        self.assertEqual(row["libreoffice_value"], "\x00",
                         "expected a NUL character from LO, got %r"
                         % (row["libreoffice_value"],))
        self.assertEqual(code, 1)

    # --- 4. Shared formulas ----------------------------------------------
    def test_shared_formula_range(self):
        """A shared formula gives openpyxl a formula string ONLY on the master
        cell. The tool must still diff every slave cell (formula-cell set
        comes from the XML)."""
        p = self._wb("shared.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        for i, v in enumerate([1, 2, 3], start=1):
            ws.cell(row=i, column=1, value=v)
            ws.cell(row=i, column=2, value="=A%d*2" % i)
        wb.save(p)

        # Convert B1:B3 into a real shared-formula block, Excel style.
        with zipfile.ZipFile(p) as z:
            parts = [(i, z.read(i.filename)) for i in z.infolist()]
        rewritten = []
        for info, data in parts:
            if info.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(b'<c r="B1"><f>A1*2</f><v></v></c>',
                                    b'<c r="B1"><f t="shared" ref="B1:B3" si="0">A1*2</f><v>2</v></c>')
                data = data.replace(b'<c r="B2"><f>A2*2</f><v></v></c>',
                                    b'<c r="B2"><f t="shared" si="0"/><v>4</v></c>')
                # B3's injected value is deliberately WRONG (should be 6)
                data = data.replace(b'<c r="B3"><f>A3*2</f><v></v></c>',
                                    b'<c r="B3"><f t="shared" si="0"/><v>99</v></c>')
                data = re.sub(rb'\s+fullCalcOnLoad="[^"]*"', b"", data)
            elif info.filename == "xl/workbook.xml":
                data = re.sub(rb'\s+fullCalcOnLoad="[^"]*"', b"", data)
            rewritten.append((info, data))
        with zipfile.ZipFile(p, "w", zipfile.ZIP_DEFLATED) as zout:
            for info, data in rewritten:
                zi = zipfile.ZipInfo(info.filename, date_time=info.date_time)
                zi.compress_type = info.compress_type
                zout.writestr(zi, data)

        code, rep, _ = run_tool(p)
        self.assertTrue(rep["trusted"], rep["warnings"])
        self.assertEqual(rep["totals"]["formula_cells"], 3)
        self.assertEqual(cat_of(rep, "S!B1")["category"], "match")
        self.assertEqual(cat_of(rep, "S!B2")["category"], "match")
        b2 = cat_of(rep, "S!B2")
        self.assertEqual(b2["shared_master"], "B1")
        self.assertEqual(b2["libreoffice_value"], 4)
        b3 = cat_of(rep, "S!B3")
        self.assertEqual(b3["category"], "numeric_mismatch")
        self.assertEqual(b3["libreoffice_value"], 6)
        self.assertEqual(code, 1)

    # --- 5. Manual calculation mode warning ------------------------------
    def test_manual_calc_mode_warning(self):
        p = self._wb("manual.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"], ws["A2"] = 2, 3
        ws["B1"] = "=A1+A2"
        wb.save(p)
        inject_cached(p, {"S": {"B1": 5}}, calc_mode="manual")

        code, rep, _ = run_tool(p)
        self.assertEqual(rep["calc_pr"]["calc_mode"], "manual")
        self.assertTrue(any("manual" in w for w in rep["warnings"]),
                        rep["warnings"])
        self.assertTrue(rep["trusted"])
        self.assertEqual(code, 0)

    # --- 6. No cached values at all --------------------------------------
    def test_no_cached_values_exit_2(self):
        p = self._wb("nocache.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"], ws["A2"] = 2, 3
        ws["B1"] = "=A1+A2"
        wb.save(p)   # openpyxl never writes cached values

        code, rep, out = run_tool(p)
        self.assertEqual(code, 2)
        self.assertEqual(rep["status"], "no_cached_values")
        self.assertFalse(rep["trusted"])
        self.assertTrue(any("NOT ONE cached value" in w for w in rep["warnings"]))

    def test_repo_verdict_mix_has_no_cached_values(self):
        p = os.path.join(REPO, "site", "audit-page", "verdict-mix.xlsx")
        if not os.path.exists(p):
            self.skipTest("verdict-mix.xlsx not present")
        code, rep, _ = run_tool(p)
        self.assertEqual(code, 2)
        self.assertEqual(rep["status"], "no_cached_values")
        self.assertEqual(rep["totals"]["formula_cells"], 12)
        self.assertEqual(rep["totals"]["formula_cells_with_excel_cached_value"], 0)

    # --- 7. Volatile functions -------------------------------------------
    def test_volatile_is_its_own_category(self):
        p = self._wb("volatile.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "=NOW()"
        ws["A2"] = "=1+1"
        wb.save(p)
        inject_cached(p, {"S": {"A1": 45000.5, "A2": 2}})

        code, rep, _ = run_tool(p)
        self.assertEqual(cat_of(rep, "S!A1")["category"], "volatile")
        self.assertEqual(cat_of(rep, "S!A2")["category"], "match")
        self.assertEqual(rep["differing_cells_total"], 0)
        self.assertEqual(code, 0, "a volatile-only difference is not a mismatch")

        code2, rep2, _ = run_tool(p, ["--include-volatile"])
        self.assertEqual(cat_of(rep2, "S!A1")["category"], "numeric_mismatch")
        self.assertEqual(code2, 1)

    # --- 8. The strip is what makes the run real -------------------------
    def test_stripping_is_necessary(self):
        """Evidence for the design: converting the UNSTRIPPED file through
        soffice can pass Excel's cached value straight through, which would
        silently produce a fake 'everything matches' report. The stripped
        copy defeats that."""
        p = self._wb("cachethru.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = True
        ws["B1"] = "=COUNT(A1,1)"
        wb.save(p)
        inject_cached(p, {"S": {"B1": 1}})   # Excel's value; LO would say 2

        with tempfile.TemporaryDirectory() as td:
            raw = X.recalc_with_libreoffice(p, td)
            passthrough = openpyxl.load_workbook(raw, data_only=True)["S"]["B1"].value
            stripped = os.path.join(td, "stripped.xlsx")
            X.make_stripped_copy(p, stripped)
            os.makedirs(os.path.join(td, "b"), exist_ok=True)
            real = X.recalc_with_libreoffice(stripped, os.path.join(td, "b"))
            recalculated = openpyxl.load_workbook(real, data_only=True)["S"]["B1"].value

        self.assertEqual(recalculated, 2,
                         "stripped conversion must force a genuine recalculation")
        print("\n    [evidence] unstripped conversion -> %r ; stripped "
              "conversion -> %r (Excel's cached value was 1)"
              % (passthrough, recalculated))

    # --- 9. Report writers -----------------------------------------------
    def test_json_and_md_reports(self):
        p = self._wb("reports.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = True
        ws["B1"] = "=COUNT(A1,1)"
        wb.save(p)
        inject_cached(p, {"S": {"B1": 1}})
        with tempfile.TemporaryDirectory() as td:
            j = os.path.join(td, "r.json")
            m = os.path.join(td, "r.md")
            buf = _Capture()
            with buf:
                code = X.main([p, "--json", j, "--md", m, "--limit", "5"])
            self.assertEqual(code, 1)
            with open(j) as f:
                data = json.load(f)
            self.assertEqual(data["status"], "ok")
            with open(m) as f:
                md = f.read()
            self.assertIn("xlsx recalc diff", md)
            self.assertIn("S!B1", md)
            self.assertIn("numeric_mismatch", md)
            self.assertIn("canispreadsheet.com/audit.html", md)
            self.assertIn("canispreadsheet.com/quirks.html", md)
            self.assertIn("S!B1", buf.text)

    # --- 10. Bad input ----------------------------------------------------
    def test_missing_file_exits_2(self):
        buf = _Capture()
        with buf:
            self.assertEqual(X.main(["/nonexistent/nope.xlsx"]), 2)

    def test_non_xlsx_exits_2(self):
        p = self._wb("notazip.xlsx")
        with open(p, "w") as f:
            f.write("this is not a zip file")
        buf = _Capture()
        with buf:
            self.assertEqual(X.main([p]), 2)


if __name__ == "__main__":
    if not HAVE_SOFFICE:
        print("WARNING: soffice not found -- end-to-end tests will be skipped.")
    unittest.main(verbosity=2)
