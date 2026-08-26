#!/usr/bin/env python3
"""
xlsx_recalc_diff.py -- offline, per-cell "will this workbook compute
differently in LibreOffice Calc than it did in Excel?" report.

WHY THIS WORKS AT ALL
---------------------
An .xlsx saved by Excel carries, for every formula cell, BOTH the formula
(`<f>`) and the value Excel last computed for it (`<v>`, the "cached
value"). That cached value is real, executed Excel ground truth sitting on
your disk. So:

    Excel's cached <v>  vs  a forced LibreOffice recalculation of the same
    formulas                                                     = per-cell diff

No Excel install is needed, nothing is uploaded, nothing leaves the
machine. (canispreadsheet.com/audit.html does a *function-level* verdict in
the browser; this tool answers the narrower, sharper question about YOUR
actual numbers.)

THE CRITICAL TRAP: LIBREOFFICE MAY NOT RECALCULATE
--------------------------------------------------
LibreOffice's "Recalculation on File Load" setting for Excel files defaults
to "Never recalculate" (or prompts), so simply running
`soffice --convert-to xlsx` on an Excel-saved file can copy Excel's cached
values straight through -- producing a beautiful, totally fake, 100%-match
report. To defeat that we build a STRIPPED COPY of the workbook at the XML
level: every `<v>` under a cell that has an `<f>` is deleted (and the
`t="str"/"e"/"b"` result-type attribute with it), so there is literally
nothing left to fall back to. Everything else in the file -- charts,
styles, defined names, pivot caches, drawings -- is copied byte-for-byte,
because round-tripping through openpyxl would silently drop most of it.

The stripped copy is a throwaway in a temp dir. YOUR FILE IS NEVER
MODIFIED and never has anything injected into it.

TRUST CHECK
-----------
After conversion we check what fraction of the formula cells that HAD an
Excel cached value came back from LibreOffice with a non-None value. If
that fraction is ~0, LibreOffice did not evaluate anything (bad install,
crash, macro-security prompt) and the whole run is marked UNTRUSTED --
reported loudly, exit code 2. Never trust a clean report without the
"recalculation verified" line.

USAGE
-----
    python3 scripts/xlsx_recalc_diff.py BOOK.xlsx [options]

      --json OUT.json    write a machine-readable report
      --md OUT.md        write a Markdown report
      --limit N          how many differing cells to print (default 25)
      --keep-temp        keep the stripped copy + LO output for debugging
      --include-volatile treat NOW/TODAY/RAND/... as real mismatches
      --sheet NAME       restrict to one sheet (repeatable)
      --quiet            summary only

    Exit codes: 0 = every cell matches, 1 = differences found,
                2 = untrusted run / unusable input / error.

Requires: openpyxl, and `soffice` on PATH (or $SOFFICE_BIN).
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from collections import Counter, OrderedDict

try:
    import openpyxl
    from openpyxl.utils.cell import range_boundaries, coordinate_from_string, column_index_from_string
except ImportError:  # pragma: no cover - environment problem, not logic
    sys.stderr.write("ERROR: openpyxl is required (pip install openpyxl)\n")
    sys.exit(2)

SOFFICE_BIN = os.environ.get("SOFFICE_BIN", "soffice")

AUDIT_URL = "https://canispreadsheet.com/audit.html"
QUIRKS_URL = "https://canispreadsheet.com/quirks.html"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

ERROR_STRINGS = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
    "#GETTING_DATA", "#CALC!", "#SPILL!", "#FIELD!", "#UNKNOWN!",
    "#BLOCKED!", "#CONNECT!", "#BUSY!", "#EXTERNAL!",
}

# Functions whose value legitimately changes every recalculation. A diff on
# these is expected and meaningless, so they get their own category instead
# of being reported as a compatibility problem.
VOLATILE_FUNCTIONS = {
    "NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT",
    "CELL", "INFO",
}
# OFFSET/INDIRECT/CELL/INFO are volatile in the recalculation-dependency
# sense but ARE deterministic given the same data, so they are not treated
# as time-varying here -- only the four true nondeterministic ones are.
NONDETERMINISTIC = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY"}

EXCEL_EPOCH = _dt.datetime(1899, 12, 30)  # serial 0 in the 1900 date system
NUMERIC_REL_TOL = 1e-9

CATEGORY_ORDER = [
    "match",
    "volatile",
    "numeric_mismatch",
    "text_mismatch",
    "type_mismatch",
    "value_vs_error",
    "error_vs_value",
    "error_vs_error",
    "missing_in_lo",
    "no_excel_value",
]

CATEGORY_HELP = {
    "match": "Excel's cached value and LibreOffice's recalculation agree.",
    "volatile": "Formula is nondeterministic (NOW/TODAY/RAND/...); a difference here is expected, not a compatibility problem.",
    "numeric_mismatch": "Both computed a number, but different numbers.",
    "text_mismatch": "Both computed text, but different text.",
    "type_mismatch": "Same-ish value, different type (e.g. boolean vs number, number vs text).",
    "value_vs_error": "Excel computed a value; LibreOffice returns an error.",
    "error_vs_value": "Excel cached an error; LibreOffice computes a value.",
    "error_vs_error": "Both error, but with different error codes (breaks ERROR.TYPE / error-sniffing formulas).",
    "missing_in_lo": "Excel had a value; LibreOffice produced nothing (blank/empty result).",
    "no_excel_value": "Formula cell with no Excel cached value -- nothing to compare against.",
}

MISMATCH_CATEGORIES = {
    "numeric_mismatch", "text_mismatch", "type_mismatch",
    "value_vs_error", "error_vs_value", "error_vs_error", "missing_in_lo",
}


# --------------------------------------------------------------------------
# XML-level workbook inspection
# --------------------------------------------------------------------------

# One <c> element. Cells never nest, and the payload (<f>/<v>/<is>) is small.
CELL_RE = re.compile(rb"<c\b[^>]*/>|<c\b[^>]*>.*?</c>", re.DOTALL)
V_RE = re.compile(rb"<v[^>]*/>|<v\b[^>]*>.*?</v>", re.DOTALL)
F_RE = re.compile(rb"<f\b[^>]*/>|<f\b[^>]*>(.*?)</f>", re.DOTALL)
ATTR_RE = re.compile(rb'(\w+(?::\w+)?)\s*=\s*"([^"]*)"')
TYPE_ATTR_RE = re.compile(rb'\s+t="(?:str|e|b|n)"')


def _attrs(tag_bytes):
    return {k.decode(): v.decode() for k, v in ATTR_RE.findall(tag_bytes)}


def _open_tag(cell_bytes):
    end = cell_bytes.find(b">")
    return cell_bytes[:end + 1]


def _unescape_xml(s):
    return (s.replace("&lt;", "<").replace("&gt;", ">")
             .replace("&quot;", '"').replace("&apos;", "'")
             .replace("&amp;", "&"))


def sheet_parts(zf):
    """Return an ordered list of (sheet_name, zip_part_name) from workbook.xml
    + its rels, so cell addresses can be reported as SheetName!A1."""
    try:
        wb_xml = zf.read("xl/workbook.xml")
    except KeyError:
        return []
    rels = {}
    try:
        rel_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        for m in re.finditer(r"<Relationship\b[^>]*>", rel_xml):
            a = {k: v for k, v in re.findall(r'(\w+)\s*=\s*"([^"]*)"', m.group(0))}
            if "Id" in a and "Target" in a:
                tgt = a["Target"]
                if tgt.startswith("/"):
                    tgt = tgt[1:]
                elif not tgt.startswith("xl/"):
                    tgt = "xl/" + tgt.lstrip("./")
                rels[a["Id"]] = tgt
    except KeyError:
        pass

    out = []
    for m in re.finditer(rb"<sheet\b[^>]*/?>", wb_xml):
        a = _attrs(m.group(0))
        name = _unescape_xml(a.get("name", ""))
        rid = a.get("r:id") or a.get("id")
        part = rels.get(rid)
        if part is None:
            part = "xl/worksheets/sheet%d.xml" % (len(out) + 1)
        out.append((name, part))
    return out


def read_calc_pr(zf):
    """Read <calcPr> from workbook.xml: calcMode + fullCalcOnLoad."""
    info = {"calc_mode": "auto", "full_calc_on_load": False, "calc_id": None}
    try:
        wb_xml = zf.read("xl/workbook.xml")
    except KeyError:
        return info
    m = re.search(rb"<calcPr\b[^>]*/?>", wb_xml)
    if not m:
        return info
    a = _attrs(m.group(0))
    info["calc_mode"] = a.get("calcMode", "auto")
    info["full_calc_on_load"] = a.get("fullCalcOnLoad", "0") in ("1", "true")
    info["calc_id"] = a.get("calcId")
    return info


def _addr_in_ranges(addr, ranges):
    col_letters, row = coordinate_from_string(addr)
    col = column_index_from_string(col_letters)
    for (min_c, min_r, max_c, max_r) in ranges:
        if min_c <= col <= max_c and min_r <= row <= max_r:
            return True
    return False


def scan_sheet_xml(data):
    """
    Parse one sheet XML. Returns dict with:
      cells: OrderedDict addr -> {formula, shared_si, shared_master,
                                  is_array_master, array_ref, has_cached,
                                  cached_raw, t}
      array_ranges: list of range_boundaries tuples for t="array" formulas
    Formula-cell membership comes from the XML (NOT from openpyxl), so
    shared-formula slaves -- which openpyxl reports as plain values -- are
    included.
    """
    cells = OrderedDict()
    array_ranges = []
    shared_master_formula = {}

    for m in CELL_RE.finditer(data):
        cell = m.group(0)
        if b"<f" not in cell:
            continue
        ca = _attrs(_open_tag(cell))
        addr = ca.get("r")
        if not addr:
            continue
        fm = F_RE.search(cell)
        ftext = ""
        fa = {}
        if fm:
            fa = _attrs(_open_tag(fm.group(0)))
            body = fm.group(1)
            if body:
                ftext = _unescape_xml(body.decode("utf-8", "replace"))
        vm = V_RE.search(cell)
        cached_raw = None
        has_cached = False
        if vm:
            inner = re.sub(rb"^<v[^>]*>|</v>$", b"", vm.group(0), flags=re.DOTALL)
            cached_raw = _unescape_xml(inner.decode("utf-8", "replace"))
            has_cached = cached_raw.strip() != ""

        si = fa.get("si")
        ftype = fa.get("t")
        if ftype == "shared" and si is not None and ftext:
            shared_master_formula[si] = (addr, ftext)
        if ftype == "array" and fa.get("ref"):
            try:
                array_ranges.append(range_boundaries(fa["ref"]))
            except Exception:
                pass

        cells[addr] = {
            "formula": ("=" + ftext) if ftext else None,
            "formula_type": ftype,
            "shared_si": si,
            "shared_master": None,
            "array_ref": fa.get("ref") if ftype == "array" else None,
            "has_cached": has_cached,
            "cached_raw": cached_raw,
            "t": ca.get("t"),
            "array_member": False,
        }

    # Resolve shared-formula slaves to their master's formula text.
    for addr, info in cells.items():
        if info["formula"] is None and info["shared_si"] in shared_master_formula:
            master_addr, master_f = shared_master_formula[info["shared_si"]]
            info["formula"] = "=" + master_f
            info["shared_master"] = master_addr

    # Cells inside a legacy array formula's ref range hold results but carry
    # no <f> of their own; include them so their cached values get stripped
    # and compared too.
    if array_ranges:
        for m in CELL_RE.finditer(data):
            cell = m.group(0)
            if b"<f" in cell:
                continue
            ca = _attrs(_open_tag(cell))
            addr = ca.get("r")
            if not addr or addr in cells:
                continue
            try:
                if not _addr_in_ranges(addr, array_ranges):
                    continue
            except Exception:
                continue
            vm = V_RE.search(cell)
            cached_raw = None
            has_cached = False
            if vm:
                inner = re.sub(rb"^<v[^>]*>|</v>$", b"", vm.group(0), flags=re.DOTALL)
                cached_raw = _unescape_xml(inner.decode("utf-8", "replace"))
                has_cached = cached_raw.strip() != ""
            cells[addr] = {
                "formula": None, "formula_type": None, "shared_si": None,
                "shared_master": None, "array_ref": None,
                "has_cached": has_cached, "cached_raw": cached_raw,
                "t": ca.get("t"), "array_member": True,
            }

    return {"cells": cells, "array_ranges": array_ranges}


def strip_sheet_xml(data, array_ranges):
    """Remove cached <v> values from every formula cell (and every cell
    inside an array formula's output range), preserving all other bytes."""
    def repl(m):
        cell = m.group(0)
        is_formula = b"<f" in cell
        if not is_formula:
            if not array_ranges:
                return cell
            ca = _attrs(_open_tag(cell))
            addr = ca.get("r")
            if not addr:
                return cell
            try:
                if not _addr_in_ranges(addr, array_ranges):
                    return cell
            except Exception:
                return cell
        open_tag = _open_tag(cell)
        rest = cell[len(open_tag):]
        rest = V_RE.sub(b"", rest)
        # Drop the result-type attribute; without a <v> it is meaningless and
        # a stale t="e"/t="str" can confuse readers.
        new_open = TYPE_ATTR_RE.sub(b"", open_tag)
        return new_open + rest
    return CELL_RE.sub(repl, data)


def make_stripped_copy(src, dst):
    """Rewrite the workbook with all formula-cell cached values removed.
    Copies every other part byte-for-byte (charts, styles, drawings, pivot
    caches survive -- an openpyxl round trip would destroy them)."""
    stats = {"sheets": 0, "values_removed": 0}
    with zipfile.ZipFile(src) as zin:
        parts = dict((n, i) for i, n in enumerate(zin.namelist()))
        sheet_names = {p for _, p in sheet_parts(zin)}
        if not sheet_names:
            sheet_names = {n for n in parts if n.startswith("xl/worksheets/sheet")}
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in sheet_names:
                    scan = scan_sheet_xml(data)
                    before = len(V_RE.findall(data))
                    data = strip_sheet_xml(data, scan["array_ranges"])
                    after = len(V_RE.findall(data))
                    stats["sheets"] += 1
                    stats["values_removed"] += (before - after)
                elif item.filename == "xl/workbook.xml":
                    # Belt and braces on OUR throwaway copy only: ask for a
                    # full recalculation on load and clear manual calc mode.
                    data = _force_full_calc(data)
                zi = zipfile.ZipInfo(item.filename, date_time=item.date_time)
                zi.compress_type = item.compress_type
                zi.external_attr = item.external_attr
                zout.writestr(zi, data)
    return stats


def _force_full_calc(wb_xml):
    if re.search(rb"<calcPr\b", wb_xml):
        def fix(m):
            tag = m.group(0)
            tag = re.sub(rb'\s+calcMode="[^"]*"', b"", tag)
            tag = re.sub(rb'\s+fullCalcOnLoad="[^"]*"', b"", tag)
            closing = b"/>" if tag.rstrip().endswith(b"/>") else b">"
            body = tag[: tag.rfind(closing)]
            return body + b' calcMode="auto" fullCalcOnLoad="1"' + closing
        return re.sub(rb"<calcPr\b[^>]*/?>", fix, wb_xml, count=1)
    return re.sub(rb"</workbook>", b'<calcPr calcMode="auto" fullCalcOnLoad="1"/></workbook>',
                  wb_xml, count=1)


# --------------------------------------------------------------------------
# LibreOffice
# --------------------------------------------------------------------------

def soffice_version():
    try:
        out = subprocess.run([SOFFICE_BIN, "--version"], capture_output=True,
                             text=True, timeout=60).stdout
        return " ".join(out.split()[:3]) if out.strip() else "unknown"
    except Exception:
        return "unknown"


def recalc_with_libreoffice(src_xlsx, workdir, timeout=600):
    """Convert src_xlsx -> xlsx with soffice in an isolated user profile.
    Returns the output path."""
    outdir = os.path.join(workdir, "lo_out")
    profile = os.path.join(workdir, "lo_profile")
    os.makedirs(outdir, exist_ok=True)
    os.makedirs(profile, exist_ok=True)
    cmd = [
        SOFFICE_BIN,
        "-env:UserInstallation=file://" + profile,
        "--headless", "--invisible", "--nologo", "--nofirststartwizard",
        "--norestore", "--nolockcheck",
        "--convert-to", "xlsx", "--outdir", outdir, src_xlsx,
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    out_path = os.path.join(outdir, os.path.splitext(os.path.basename(src_xlsx))[0] + ".xlsx")
    if proc.returncode != 0 and not os.path.exists(out_path):
        raise RuntimeError("soffice exited %d: %s" % (proc.returncode, proc.stderr.strip()[:800]))
    if not os.path.exists(out_path):
        raise RuntimeError("soffice produced no output file. stderr: %s" % proc.stderr.strip()[:800])
    return out_path


# --------------------------------------------------------------------------
# Value normalization + comparison
# --------------------------------------------------------------------------

_ESCAPE_RE = re.compile(r"_x([0-9A-Fa-f]{4})_")


def unescape_ooxml(s):
    """LibreOffice (and Excel) store control characters as _xNNNN_. Turn
    those back into the real characters so both sides compare like for like.
    `_x005F_x0000_` is the escaped-underscore form of a literal `_x0000_`."""
    if not isinstance(s, str) or "_x" not in s:
        return s
    s = s.replace("_x005F_", "\x00ESCAPED_UNDERSCORE\x00")
    s = _ESCAPE_RE.sub(lambda m: chr(int(m.group(1), 16)), s)
    return s.replace("\x00ESCAPED_UNDERSCORE\x00", "_x005F_")


def to_serial(v):
    if isinstance(v, _dt.datetime):
        d = v - EXCEL_EPOCH
        return d.days + d.seconds / 86400.0 + d.microseconds / 86400e6
    if isinstance(v, _dt.date):
        return float((_dt.datetime(v.year, v.month, v.day) - EXCEL_EPOCH).days)
    if isinstance(v, _dt.time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400.0 + v.microsecond / 86400e6
    if isinstance(v, _dt.timedelta):
        return v.days + v.seconds / 86400.0 + v.microseconds / 86400e6
    return v


def normalize(v):
    """Normalize one side of the comparison. Dates -> Excel serials,
    _xNNNN_ escapes -> real characters, error strings recognized."""
    v = to_serial(v)
    if isinstance(v, str):
        v = unescape_ooxml(v)
        if v in ERROR_STRINGS:
            return ("error", v)
        return ("text", v)
    if isinstance(v, bool):
        return ("bool", v)
    if isinstance(v, (int, float)):
        return ("number", float(v))
    if v is None:
        return ("empty", None)
    return ("text", str(v))


def numbers_close(a, b):
    if a == b:
        return True
    scale = max(1.0, abs(a), abs(b))
    return abs(a - b) <= NUMERIC_REL_TOL * scale


def classify(excel_val, lo_val):
    """Return (category, note). 'match' means the two agree."""
    ek, ev = normalize(excel_val)
    lk, lv = normalize(lo_val)

    # A formula returning "" round-trips through .xlsx as a value-less cell,
    # which reads back as None. Blank and empty-string are genuinely
    # indistinguishable at this layer, so treat them as equal.
    if (ek, ev) == ("text", "") and lk == "empty":
        return "match", None
    if (lk, lv) == ("text", "") and ek == "empty":
        return "match", None

    if ek == "empty" and lk == "empty":
        return "match", None
    if ek == "error" and lk == "error":
        return ("match", None) if ev == lv else ("error_vs_error", "%s -> %s" % (ev, lv))
    if ek == "error":
        return "error_vs_value", "Excel cached %s; LibreOffice computes a value" % ev
    if lk == "error":
        return "value_vs_error", "LibreOffice returns %s" % lv
    if lk == "empty":
        return "missing_in_lo", "LibreOffice produced no value"
    if ek == "empty":
        # Excel had no value but LO produced one -- caller filters these out
        # earlier (no_excel_value); defensive fallback.
        return "no_excel_value", None
    if ek == "bool" and lk == "bool":
        return ("match", None) if ev == lv else ("type_mismatch", "boolean differs")
    if ek == "bool" or lk == "bool":
        en = float(ev) if ek in ("bool", "number") else None
        ln = float(lv) if lk in ("bool", "number") else None
        if en is not None and ln is not None and numbers_close(en, ln):
            return "type_mismatch", "same number, different type (boolean vs number)"
        return "type_mismatch", "boolean vs %s" % (lk if ek == "bool" else ek)
    if ek == "number" and lk == "number":
        return ("match", None) if numbers_close(ev, lv) else (
            "numeric_mismatch", "delta %r" % (lv - ev))
    if ek == "text" and lk == "text":
        return ("match", None) if ev == lv else ("text_mismatch", None)
    return "type_mismatch", "%s vs %s" % (ek, lk)


FUNC_RE = re.compile(r"([A-Z][A-Z0-9_.]*)\s*\(")
# Functions added to Excel after 2007 are stored in the XML with an "_xlfn."
# (or "_xlfn._xlws.") prefix; report the human-typed name instead.
XLFN_PREFIX_RE = re.compile(r"_XLFN\.(?:_XLWS\.)?")


def functions_in(formula):
    if not formula:
        return []
    return sorted(set(FUNC_RE.findall(XLFN_PREFIX_RE.sub("", formula.upper()))))


def is_nondeterministic(formula):
    return bool(set(functions_in(formula)) & NONDETERMINISTIC)


def display(v):
    v = to_serial(v)
    if v is None:
        return "(blank)"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return repr(v)
    if isinstance(v, str):
        s = unescape_ooxml(v)
        s = "".join(ch if ch.isprintable() else "\\x%02x" % ord(ch) for ch in s)
        return "'" + (s if len(s) <= 60 else s[:57] + "...") + "'"
    return repr(v)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def build_report(path, args, workdir):
    report = {
        "tool": "xlsx_recalc_diff.py",
        "file": os.path.abspath(path),
        "libreoffice_version": soffice_version(),
        "generated_at": _dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "warnings": [],
        "trusted": None,
        "cells": [],
    }

    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        if "xl/workbook.xml" not in names:
            raise RuntimeError("not an .xlsx workbook (no xl/workbook.xml): %s" % path)
        calc = read_calc_pr(zf)
        report["calc_pr"] = calc
        sheets = sheet_parts(zf)
        xml_cells = OrderedDict()
        for sname, part in sheets:
            if args.sheet and sname not in args.sheet:
                continue
            try:
                data = zf.read(part)
            except KeyError:
                report["warnings"].append("sheet part missing from zip: %s (%s)" % (part, sname))
                continue
            xml_cells[sname] = scan_sheet_xml(data)["cells"]

    if calc["calc_mode"] == "manual":
        report["warnings"].append(
            "workbook.xml says calcMode=\"manual\": Excel was NOT recalculating "
            "automatically, so the cached values in this file may be STALE and "
            "may not reflect what Excel would compute today. Open it in Excel, "
            "press F9, save, and re-run for a trustworthy baseline.")
    if calc["full_calc_on_load"]:
        report["warnings"].append(
            "workbook.xml sets fullCalcOnLoad=\"1\": the writer of this file "
            "asked every reader to recalculate on open, which usually means the "
            "cached values were not written by a real calculation engine.")

    total_formula_cells = sum(len(c) for c in xml_cells.values())
    cached_cells = sum(1 for c in xml_cells.values() for i in c.values() if i["has_cached"])
    report["totals"] = {
        "sheets": len(xml_cells),
        "formula_cells": total_formula_cells,
        "formula_cells_with_excel_cached_value": cached_cells,
    }

    if total_formula_cells == 0:
        report["status"] = "no_formulas"
        report["trusted"] = False
        return report, 2
    if cached_cells == 0:
        report["status"] = "no_cached_values"
        report["trusted"] = False
        report["warnings"].append(
            "This workbook contains %d formula cells but NOT ONE cached value. "
            "Files written by openpyxl/xlsxwriter/pandas (and files saved with "
            "calculation disabled) store only formulas. There is no Excel "
            "result to diff against, so a recalc diff is meaningless here. "
            "Open and save the file in Excel first." % total_formula_cells)
        return report, 2

    # ---- stripped copy + forced LibreOffice recalculation ----
    stripped = os.path.join(workdir, "stripped.xlsx")
    strip_stats = make_stripped_copy(path, stripped)
    report["strip"] = strip_stats
    lo_path = recalc_with_libreoffice(stripped, workdir)
    report["lo_output"] = lo_path if args.keep_temp else None

    wb_excel = openpyxl.load_workbook(path, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_lo = openpyxl.load_workbook(lo_path, data_only=True, read_only=False)

    counts = Counter()
    func_counter = Counter()
    rows = []
    recalc_hits = recalc_seen = 0

    for sname, cells in xml_cells.items():
        ws_x = wb_excel[sname] if sname in wb_excel.sheetnames else None
        ws_f = wb_formulas[sname] if sname in wb_formulas.sheetnames else None
        ws_l = wb_lo[sname] if sname in wb_lo.sheetnames else None
        if ws_l is None:
            report["warnings"].append(
                "sheet %r missing from the LibreOffice output; skipped" % sname)
            continue
        for addr, info in cells.items():
            excel_val = ws_x[addr].value if ws_x is not None else None
            lo_val = ws_l[addr].value if ws_l is not None else None

            formula = info["formula"]
            if formula is None and ws_f is not None:
                fv = ws_f[addr].value
                if hasattr(fv, "text"):        # ArrayFormula
                    formula = fv.text
                elif isinstance(fv, str) and fv.startswith("="):
                    formula = fv

            if info["has_cached"]:
                recalc_seen += 1
                if lo_val is not None:
                    recalc_hits += 1

            if not info["has_cached"]:
                cat, note = "no_excel_value", None
            else:
                cat, note = classify(excel_val, lo_val)
                if cat != "match" and is_nondeterministic(formula) and not args.include_volatile:
                    cat = "volatile"
                    note = "nondeterministic formula; a difference is expected"

            counts[cat] += 1
            if cat in MISMATCH_CATEGORIES:
                for fn in functions_in(formula):
                    func_counter[fn] += 1
            rows.append({
                "sheet": sname,
                "cell": addr,
                "address": "%s!%s" % (sname, addr),
                "formula": formula,
                "shared_master": info.get("shared_master"),
                "array_member": info.get("array_member", False),
                "excel_value": to_serial(excel_val) if not isinstance(excel_val, str) else unescape_ooxml(excel_val),
                "libreoffice_value": to_serial(lo_val) if not isinstance(lo_val, str) else unescape_ooxml(lo_val),
                "excel_display": display(excel_val),
                "libreoffice_display": display(lo_val),
                "category": cat,
                "note": note,
            })

    rate = (recalc_hits / recalc_seen) if recalc_seen else 0.0
    report["recalc_check"] = {
        "formula_cells_with_excel_cached_value": recalc_seen,
        "of_those_with_a_libreoffice_value": recalc_hits,
        "rate": round(rate, 4),
        "method": "stripped every cached <v> at the XML level, then "
                  "soffice --headless --convert-to xlsx in an isolated user profile",
    }
    if rate < 0.05:
        report["trusted"] = False
        report["warnings"].append(
            "UNTRUSTED RUN: only %.1f%% of the formula cells that had an Excel "
            "cached value came back with any value from LibreOffice. That means "
            "LibreOffice did not actually evaluate this workbook -- every "
            "difference below is an artifact, not a compatibility finding."
            % (100 * rate))
    else:
        report["trusted"] = True
        if rate < 0.5:
            report["warnings"].append(
                "Only %.1f%% of cached formula cells produced a value in "
                "LibreOffice; many formulas may be unsupported (check the "
                "value_vs_error / missing_in_lo categories)." % (100 * rate))

    counts_d = OrderedDict()
    for k in CATEGORY_ORDER:
        if counts.get(k):
            counts_d[k] = counts[k]
    for k, v in counts.items():
        counts_d.setdefault(k, v)

    report["categories"] = counts_d
    report["category_help"] = {k: CATEGORY_HELP[k] for k in counts_d if k in CATEGORY_HELP}
    report["differing_cells_total"] = sum(counts[c] for c in MISMATCH_CATEGORIES)
    report["top_functions_in_differing_cells"] = [
        {"function": f, "differing_cells": n} for f, n in func_counter.most_common(15)]
    report["cells"] = rows
    report["status"] = "ok"

    if not report["trusted"]:
        return report, 2
    return report, (1 if report["differing_cells_total"] else 0)


def differing_rows(report):
    return [r for r in report.get("cells", []) if r["category"] in MISMATCH_CATEGORIES]


def print_console(report, limit, quiet=False):
    w = sys.stdout.write
    w("\n" + "=" * 72 + "\n")
    w("xlsx recalc diff -- Excel cached values vs forced LibreOffice recalc\n")
    w("=" * 72 + "\n")
    w("File:        %s\n" % report["file"])
    w("LibreOffice: %s\n" % report["libreoffice_version"])
    calc = report.get("calc_pr") or {}
    w("Excel calc:  calcMode=%s fullCalcOnLoad=%s\n"
      % (calc.get("calc_mode"), calc.get("full_calc_on_load")))
    t = report.get("totals", {})
    w("Formula cells: %s across %s sheet(s); %s carry an Excel cached value\n"
      % (t.get("formula_cells"), t.get("sheets"),
         t.get("formula_cells_with_excel_cached_value")))

    for msg in report.get("warnings", []):
        w("\n!! WARNING: %s\n" % _wrap(msg, 4))

    if report.get("status") == "no_cached_values":
        w("\nRESULT: no diff is possible for this file.\n")
        _footer(w)
        return
    if report.get("status") == "no_formulas":
        w("\nRESULT: this workbook has no formula cells at all -- nothing to diff.\n")
        _footer(w)
        return

    rc = report.get("recalc_check", {})
    w("\nRecalculation check: %s/%s cached formula cells produced a value in "
      "LibreOffice (%.1f%%) -> %s\n"
      % (rc.get("of_those_with_a_libreoffice_value"),
         rc.get("formula_cells_with_excel_cached_value"),
         100 * rc.get("rate", 0),
         "TRUSTED" if report.get("trusted") else "UNTRUSTED"))

    w("\nPer-category counts\n-------------------\n")
    for cat, n in report.get("categories", {}).items():
        w("  %-18s %6d   %s\n" % (cat, n, CATEGORY_HELP.get(cat, "")))

    diffs = differing_rows(report)
    w("\nCells that will compute differently in LibreOffice: %d\n" % len(diffs))

    tops = report.get("top_functions_in_differing_cells") or []
    if tops:
        w("\nTop functions appearing in differing cells\n"
          "-----------------------------------------\n")
        for item in tops[:10]:
            w("  %-20s %d\n" % (item["function"], item["differing_cells"]))

    if diffs and not quiet:
        w("\nFirst %d differing cell(s)\n" % min(limit, len(diffs)))
        w("-" * 72 + "\n")
        for r in diffs[:limit]:
            w("%s  [%s]\n" % (r["address"], r["category"]))
            w("    formula : %s%s\n" % (
                r["formula"] or "(shared/array member)",
                "   (shared from %s)" % r["shared_master"] if r["shared_master"] else ""))
            w("    Excel   : %s\n" % r["excel_display"])
            w("    LibreO. : %s\n" % r["libreoffice_display"])
            if r["note"]:
                w("    note    : %s\n" % r["note"])
        if len(diffs) > limit:
            w("... %d more (use --limit N, or --json/--md for the full list)\n"
              % (len(diffs) - limit))

    if not diffs:
        w("\nAll formula cells with an Excel cached value recompute identically "
          "in LibreOffice Calc.\n")
    _footer(w)


def _wrap(text, indent=0, width=76):
    import textwrap
    return ("\n" + " " * indent).join(textwrap.wrap(text, width - indent))


def _footer(w):
    w("\nWhy does a given cell differ? Function-level Excel / LibreOffice /\n"
      "Google Sheets behaviour, all of it actually executed:\n"
      "  %s   (per-function verdicts for a whole workbook, in your browser)\n"
      "  %s   (catalogue of known engine divergences)\n"
      % (AUDIT_URL, QUIRKS_URL))


def write_markdown(report, out_path, limit):
    L = []
    a = L.append
    a("# xlsx recalc diff\n")
    a("Excel's cached values vs a forced LibreOffice Calc recalculation of the "
      "same workbook.\n")
    a("| | |")
    a("|---|---|")
    a("| File | `%s` |" % os.path.basename(report["file"]))
    a("| LibreOffice | %s |" % report["libreoffice_version"])
    a("| Generated | %s |" % report["generated_at"])
    calc = report.get("calc_pr") or {}
    a("| Excel calc mode | `%s` (fullCalcOnLoad=%s) |"
      % (calc.get("calc_mode"), calc.get("full_calc_on_load")))
    t = report.get("totals", {})
    a("| Formula cells | %s in %s sheet(s) |" % (t.get("formula_cells"), t.get("sheets")))
    a("| With an Excel cached value | %s |" % t.get("formula_cells_with_excel_cached_value"))
    rc = report.get("recalc_check") or {}
    if rc:
        a("| Recalculation verified | %s/%s cells (%.1f%%) -> **%s** |"
          % (rc.get("of_those_with_a_libreoffice_value"),
             rc.get("formula_cells_with_excel_cached_value"),
             100 * rc.get("rate", 0), "TRUSTED" if report.get("trusted") else "UNTRUSTED"))
    a("")
    if report.get("warnings"):
        a("## Warnings\n")
        for msg in report["warnings"]:
            a("- %s" % msg)
        a("")
    if report.get("status") in ("no_cached_values", "no_formulas"):
        a("**No diff is possible for this file.**\n")
        _md_footer(a)
        return _write(out_path, "\n".join(L))

    a("## Per-category counts\n")
    a("| Category | Cells | Meaning |")
    a("|---|---:|---|")
    for cat, n in report.get("categories", {}).items():
        a("| `%s` | %d | %s |" % (cat, n, CATEGORY_HELP.get(cat, "")))
    a("")
    tops = report.get("top_functions_in_differing_cells") or []
    if tops:
        a("## Functions appearing in differing cells\n")
        a("| Function | Differing cells |")
        a("|---|---:|")
        for item in tops:
            a("| `%s` | %d |" % (item["function"], item["differing_cells"]))
        a("")
    diffs = differing_rows(report)
    a("## Differing cells (%d)\n" % len(diffs))
    if diffs:
        a("| Cell | Category | Formula | Excel | LibreOffice |")
        a("|---|---|---|---|---|")
        for r in diffs[:limit]:
            a("| `%s` | %s | `%s` | %s | %s |" % (
                r["address"], r["category"],
                (r["formula"] or "(shared/array member)").replace("|", "\\|"),
                r["excel_display"].replace("|", "\\|"),
                r["libreoffice_display"].replace("|", "\\|")))
        if len(diffs) > limit:
            a("\n_%d more cells omitted (--limit)._" % (len(diffs) - limit))
    else:
        a("None - every formula cell with an Excel cached value recomputes "
          "identically in LibreOffice Calc.")
    a("")
    _md_footer(a)
    return _write(out_path, "\n".join(L))


def _md_footer(a):
    a("---\n")
    a("Function-level Excel/LibreOffice/Google Sheets behaviour, all actually "
      "executed: [%s](%s) and [%s](%s)." % (AUDIT_URL, AUDIT_URL, QUIRKS_URL, QUIRKS_URL))


def _write(path, text):
    with open(path, "w", encoding="utf-8") as f:
        f.write(text + "\n")
    return path


def json_safe(obj):
    if isinstance(obj, dict):
        return {k: json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [json_safe(v) for v in obj]
    if isinstance(obj, (_dt.datetime, _dt.date, _dt.time)):
        return obj.isoformat()
    if isinstance(obj, bytes):
        return obj.decode("utf-8", "replace")
    return obj


def main(argv=None):
    p = argparse.ArgumentParser(
        prog="xlsx_recalc_diff.py",
        description="Offline per-cell diff: Excel's cached values in an .xlsx "
                    "vs a forced LibreOffice Calc recalculation of the same file. "
                    "Nothing is uploaded; your file is never modified.")
    p.add_argument("xlsx", help="path to the .xlsx saved by Excel")
    p.add_argument("--json", dest="json_out", metavar="OUT.json")
    p.add_argument("--md", dest="md_out", metavar="OUT.md")
    p.add_argument("--limit", type=int, default=25,
                   help="how many differing cells to show (default 25)")
    p.add_argument("--keep-temp", action="store_true",
                   help="keep the stripped copy and LibreOffice output for debugging")
    p.add_argument("--include-volatile", action="store_true",
                   help="count NOW/TODAY/RAND/RANDBETWEEN/RANDARRAY differences as real mismatches")
    p.add_argument("--sheet", action="append", default=None,
                   help="restrict to this sheet name (repeatable)")
    p.add_argument("--quiet", action="store_true", help="summary only")
    args = p.parse_args(argv)

    if not os.path.isfile(args.xlsx):
        sys.stderr.write("ERROR: no such file: %s\n" % args.xlsx)
        return 2
    if shutil.which(SOFFICE_BIN) is None and not os.path.isfile(SOFFICE_BIN):
        sys.stderr.write(
            "ERROR: LibreOffice not found (%r). Install LibreOffice or set "
            "$SOFFICE_BIN to the soffice binary.\n" % SOFFICE_BIN)
        return 2

    workdir = tempfile.mkdtemp(prefix="xlsx_recalc_diff_")
    try:
        try:
            report, code = build_report(args.xlsx, args, workdir)
        except zipfile.BadZipFile:
            sys.stderr.write("ERROR: %s is not a valid .xlsx (zip) file. "
                             ".xls and .ods are not supported.\n" % args.xlsx)
            return 2
        except Exception as e:
            sys.stderr.write("ERROR: %s\n" % e)
            return 2

        print_console(report, args.limit, args.quiet)
        if args.json_out:
            with open(args.json_out, "w", encoding="utf-8") as f:
                json.dump(json_safe(report), f, indent=2, ensure_ascii=False)
            print("\nJSON report: %s" % args.json_out)
        if args.md_out:
            write_markdown(report, args.md_out, args.limit)
            print("Markdown report: %s" % args.md_out)
        if args.keep_temp:
            print("Temp files kept in: %s" % workdir)
        return code
    finally:
        if not args.keep_temp:
            shutil.rmtree(workdir, ignore_errors=True)


if __name__ == "__main__":
    sys.exit(main())
