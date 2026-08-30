#!/usr/bin/env python3
"""Verify each recipe's example formula by actually executing it in headless
LibreOffice (same convert-to recalc trick as the compat harness). Writes
results/recipes-verified.json: slug -> {verified, engine_version, actual, expected}.

A recipe may also carry an optional "variants" list; each variant's "verify" is
either one check dict or a list of them (each {label?, formula, expected,
setup_cells?, setup_sheets?, check_range?}, falling back to the variant's own
setup_cells / setup_sheets).
Those are executed too and stored under the slug as
  "variants": [{"heading": ..., "checks": [{key, label, formula, expected, actual, verified}]}]
and a top-level "extra_checks" list (checks belonging to the recipe but to no
variant) is stored the same way under "extra_checks".

A check may carry "engines": ["google_sheets"] (or ["libreoffice"]); absent
means all engines. This script executes ONLY the checks scoped to LibreOffice,
and the recipe's "verified" flag is therefore the AND over the LibreOffice-
scoped checks alone -- a Sheets-only alternative formula neither runs here nor
counts here. Every stored check carries its stable "key" (see
harness/recipe_corpus.py) so consumers merge by key rather than by position.

The corpus-shaped parts of this script -- which checks exist, how a variant
check inherits setup_cells/setup_sheets, the read-back normalization, and the
expected-vs-actual rule -- live in harness/recipe_corpus.py so that the Google
Sheets recipe runner (harness/run_sheets.py build-recipes / ingest-recipes)
enumerates and judges EXACTLY the same checks the same way. This script keeps
only the LibreOffice-specific part: actually making Calc calculate. The move
was behaviour-preserving; results/recipes-verified.json is unchanged by it.

Usage:
  python3 scripts/verify_recipes.py              # all recipes (rewrites the file)
  python3 scripts/verify_recipes.py <slug> ...   # only those slugs, merged into
                                                 # the existing results file
"""
import json, glob, os, subprocess, tempfile, re
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOFF = "soffice"
import sys
sys.path.insert(0, os.path.join(ROOT, "harness"))
from xlfn_map import to_storage_formula_all  # prefix modern funcs (_xlfn.) for OOXML
from recipe_corpus import (                    # shared with the Sheets recipe runner
    LIBREOFFICE, iter_checks, norm, run_check)

ENGINE = LIBREOFFICE   # the engine identity this script executes as; checks
                       # scoped to any other engine are skipped, not failed

def lo_version():
    out = subprocess.run([SOFF,"--version"],capture_output=True,text=True,timeout=30).stdout
    for t in out.split():
        if t[:1].isdigit() and "." in t: return t
    return "unknown"

def run_case(setup, formula, check_range, setup_sheets=None):
    wb=openpyxl.Workbook(); ws=wb.active
    # Optional extra sheets so recipes can reference another tab (Sheet2!A1).
    # Written before the main sheet's data; the main formula lives on ws.
    for sheet_name, cells in (setup_sheets or {}).items():
        extra=wb.create_sheet(title=sheet_name)
        for a,val in (cells or {}).items(): extra[a]=val
    for a,val in (setup or {}).items(): ws[a]=val
    anchor = check_range.split(":")[0] if check_range else "H1"
    formula = to_storage_formula_all(formula)  # add _xlfn. prefixes for OOXML round-trip
    if check_range: ws[anchor]=ArrayFormula(check_range, formula)
    else: ws[anchor]=formula
    ws["Z1"]="=1+1"  # recalc canary
    d=tempfile.mkdtemp(); p=os.path.join(d,"in.xlsx"); wb.save(p)
    outd=os.path.join(d,"out"); os.makedirs(outd,exist_ok=True)
    subprocess.run([SOFF,"--headless","--convert-to","xlsx","--outdir",outd,p],
                   capture_output=True,timeout=120)
    wb2=openpyxl.load_workbook(os.path.join(outd,"in.xlsx"),data_only=True); ws2=wb2.active
    assert ws2["Z1"].value==2, "recalc canary failed"
    if check_range:
        mn=openpyxl.utils.cell.range_boundaries(check_range)
        vals=[norm(ws2.cell(row=r,column=c).value)
              for r in range(mn[1],mn[3]+1) for c in range(mn[0],mn[2]+1)]
        return [v for v in vals if v is not None]
    return norm(ws2[anchor].value)

def check(v, default_setup=None, default_sheets=None):
    """Execute one check dict in LibreOffice; returns (actual, ok).

    The setup-inheritance, comparison and error-capture rules are
    recipe_corpus.run_check()'s (moved there verbatim from this function);
    run_case above is the LibreOffice engine plugged into it.

    ENGINE is passed so a check scoped to another engine (`"engines":
    ["google_sheets"]`) is never executed here: LibreOffice would happily run
    a Sheets-only alternative formula, fail it, and drag the recipe's
    LibreOffice badge down with it. Checks reach this function already
    filtered by iter_checks(..., engine=ENGINE); the argument is the second
    lock on the same door.
    """
    return run_check(v, run_case, default_setup, default_sheets, engine=ENGINE)

def _key_of(c, default):
    """The stable key of a raw JSON check dict (explicit "id" wins)."""
    return str(c.get("id") or default)

def _main_setup(v):
    """Default (setup_cells, setup_sheets) an extra_check inherits: the main
    example's, matching recipe_corpus.iter_checks()."""
    return v.get("setup_cells"), v.get("setup_sheets")

RESULTS=os.path.join(ROOT,"results/recipes-verified.json")
only=set(sys.argv[1:])
ver=lo_version(); out={}
if only and os.path.exists(RESULTS):   # merge into existing results, don't drop the rest
    out=json.load(open(RESULTS)).get("recipes",{})
n_run=0
for f in sorted(glob.glob(os.path.join(ROOT,"data/recipes/*.json"))):
    r=json.load(open(f))
    if only and r["slug"] not in only: continue
    n_run+=1
    # Enumerate through the shared corpus module, filtered to THIS engine, so
    # that a check carrying "engines": ["google_sheets"] is skipped entirely
    # rather than executed and failed here. Keys come from iter_checks and are
    # written onto every payload: the site merges results by key, not by
    # position, so appending a Sheets-only check to a variant cannot shift an
    # older stored value onto the wrong formula.
    scoped = {c["key"]: c for c in iter_checks(r, engine=ENGINE)}
    v=r["verify"]
    main = next((c for c in scoped.values() if c["kind"]=="main"), None)
    if main is None:
        # A recipe whose worked example is scoped away from LibreOffice has
        # nothing for this script to report; skipping it keeps the results
        # file free of empty half-records.
        print(f"  -- {r['slug']:42} main example not scoped to {ENGINE}; skipped")
        n_run-=1
        continue
    actual, ok = check(v)
    rec={"verified":ok,"engine":"LibreOffice Calc","engine_version":ver,
         "formula":v["formula"],"expected":v["expected"],"actual":actual,
         "key":main["key"]}
    print(f"  {'OK ' if ok else 'XX '} {r['slug']:42} got={actual} want={v['expected']}")
    variants=[]
    for vi, var in enumerate(r.get("variants") or []):
        checks = var.get("verify") or []
        if isinstance(checks, dict): checks=[checks]
        done=[]
        for ci, c in enumerate(checks):
            chk = scoped.get(_key_of(c, f"v{vi}c{ci}"))
            if chk is None: continue          # scoped to another engine
            a, o = check(c, var.get("setup_cells"), var.get("setup_sheets"))
            if not o: ok=False
            done.append({"label":c.get("label",""),"formula":c["formula"],
                         "expected":c["expected"],"actual":a,"verified":o,
                         "key":chk["key"]})
            print(f"    {'ok ' if o else 'XX '} {c['formula'][:64]:66} got={a} want={c['expected']}")
        variants.append({"heading":var.get("heading",""),"checks":done})
    if variants:
        rec["variants"]=variants
        rec["verified"]=ok   # recipe counts as verified only if every variant check passes too
    extra=[]
    for xi, c in enumerate(r.get("extra_checks") or []):
        chk = scoped.get(_key_of(c, f"x{xi}"))
        if chk is None: continue              # scoped to another engine
        a, o = check(c, *_main_setup(v))
        if not o: ok=False
        extra.append({"label":c.get("label",""),"formula":c["formula"],
                      "expected":c["expected"],"actual":a,"verified":o,
                      "key":chk["key"]})
        print(f"    {'ok ' if o else 'XX '} {c['formula'][:64]:66} got={a} want={c['expected']}")
    if extra:
        rec["extra_checks"]=extra
        rec["verified"]=ok
    out[r["slug"]]=rec
if only and not n_run:
    print("no recipe matched:", ", ".join(sorted(only))); sys.exit(2)
json.dump({"generated_at_note":"stamped post-run","engine_version":ver,"recipes":out},
          open(RESULTS,"w"),indent=2,default=str)
print("engine:",ver,"| ran:",n_run,"| verified:",sum(1 for x in out.values() if x['verified']),"/",len(out))
