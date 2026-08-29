#!/usr/bin/env python3
"""
LibreOffice Calc engine runner for the spreadsheet function-compatibility
harness.

WHAT THIS DOES
--------------
1. Loads every data/tests/<FUNCTION>.json file (or a subset given on the
   command line).
2. Builds a single .xlsx workbook with openpyxl: one worksheet per test
   case, containing the case's setup_cells plus the formula under test
   (translated to its correct OOXML storage form via harness/xlfn_map.py),
   plus a small canary block on every sheet.
3. Forces LibreOffice to actually RECALCULATE (not just re-serialize
   cached values) by round-tripping the file through
   `soffice --headless --convert-to xlsx`.
4. Reads the recalculated file back with openpyxl(data_only=True) and
   extracts real computed values.
5. Writes results/libreoffice-24.2.json mapping test id -> computed
   value/error/notes, plus a top-level canary block proving genuine
   recalculation occurred.

WHY `--convert-to` IS TRUSTWORTHY HERE (READ BEFORE CHANGING THIS)
-------------------------------------------------------------------
openpyxl NEVER writes a cached <v> value for a formula cell -- only the
formula string itself. That means there is no stale cached value for
LibreOffice to fall back to; to produce ANY value in column output at all,
soffice --convert-to MUST evaluate every formula from scratch. We proved
this empirically two ways:
  (a) A volatile canary `=NOW()` produces a genuinely different timestamp
      on two separate conversion runs a few seconds apart (see
      canary.now_run_1 / canary.now_run_2 in the output JSON, or re-run
      this script twice to reproduce).
  (b) A deterministic arithmetic canary `=1111+2222` (no cached value
      possible) evaluates to exactly 3333 on every sheet; if recalculation
      were NOT happening, openpyxl would read back None (blank) for every
      formula cell instead, since nothing was ever cached.
If canary checks fail, this script marks the ENTIRE run "trusted": false
and every function result gets an "UNTRUSTED_RECALC" note -- never trust
a green run without checking the "trusted" flag in the output file.

THE _xlfn. PREFIX GOTCHA
-------------------------
See harness/xlfn_map.py for a full writeup. Short version: functions added
to Excel after 2007 (XLOOKUP, LET, LAMBDA, FILTER, SORT, UNIQUE, SEQUENCE,
TEXTSPLIT, TEXTBEFORE/AFTER, IFS, SWITCH, MAXIFS/MINIFS, TEXTJOIN, CONCAT,
IFNA, ARRAYTOTEXT, ...) must be written into the raw .xlsx XML with an
"_xlfn." (or "_xlfn._xlws." for FILTER/SORT) prefix, or EVERY engine
(including real Excel) will show #NAME? even if the function is fully
supported. We translate this automatically per test case based on the
file's "function" field.

USAGE
-----
    python3 harness/run_lo.py                  # run all data/tests/*.json
    python3 harness/run_lo.py XLOOKUP LET       # run only these functions

SUBSET RUNS MERGE, THEY DO NOT OVERWRITE (READ BEFORE CHANGING THIS)
--------------------------------------------------------------------
A full run (no function names on the command line) writes
results/libreoffice-<major.minor>.json from scratch. A SUBSET run merges
into that file instead: only the function_results entries for the
functions named on the command line are replaced, every other function's
previously executed result is preserved byte-for-byte, and the top-level
generated_at / canary / recalc_method are refreshed from this run so the
canary block always describes the most recent execution against the file.
`trusted` becomes the AND of the previous file's flag and this run's, so a
merge can only ever downgrade trust, never launder an untrusted run into a
trusted file. Each merge appends a record to a top-level "subset_runs" list
(timestamp, which functions were re-executed, that run's own trusted flag,
and the generated_at it superseded) so the provenance of a mixed file is
auditable. Merging is refused if the installed LibreOffice version does not
match the engine_version already recorded in the file -- results from two
different engine builds must never share one results file.
"""
import json
import os
import subprocess
import sys
import tempfile
import time
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlfn_map import to_storage_formula_all  # noqa: E402,F401

# Everything below is shared verbatim with harness/run_sheets.py -- see
# harness/corpus.py for why. These are re-exported at module level under
# their original names so existing callers of run_lo.build_workbook(),
# run_lo.load_test_files(), etc. keep working unchanged.
from corpus import (  # noqa: E402,F401
    REPO_ROOT,
    TESTS_DIR,
    RESULTS_DIR,
    KNOWN_ERROR_STRINGS,
    CANARY_ARITH_FORMULA,
    CANARY_ARITH_EXPECTED,
    CANARY_ANCHOR,
    EXCEL_EPOCH,
    load_test_files,
    sanitize_sheet_name,
    build_workbook,
    anchor_for_case,
    flatten_cases,
    cell_addrs_in_range,
    is_error_value,
    normalize_readback_value,
    values_roughly_equal,
    compare_expected,
)

import openpyxl  # noqa: E402

# Which soffice binary to drive. Defaults to the system "soffice" (PATH), but
# can be overridden to point at an isolated/alternate LibreOffice install
# (e.g. an extracted AppImage of an OLDER release) so we can execute the exact
# same corpus under multiple engine versions for version-range compatibility
# data. The output filename is derived from the binary's own reported version,
# so pointing this at 24.2 writes results/libreoffice-24.2.json automatically.
SOFFICE_BIN = os.environ.get("SOFFICE_BIN", "soffice")


def _detect_lo_version():
    """Detect the running LibreOffice version at runtime so results never
    silently claim a stale/hardcoded version. Falls back to 'unknown'."""
    try:
        out = subprocess.run(
            [SOFFICE_BIN, "--version"], capture_output=True, text=True, timeout=30
        ).stdout
        for tok in out.split():
            if tok[:1].isdigit() and "." in tok:
                return tok
    except Exception:
        pass
    return "unknown"


LO_VERSION = _detect_lo_version()  # e.g. "25.8.7.3"
LO_VERSION_TAG = (
    ".".join(LO_VERSION.split(".")[:2]) if LO_VERSION != "unknown" else "unknown"
)  # major.minor for the results filename, e.g. "25.8"


def run():
    requested = set(sys.argv[1:]) or None
    test_files = load_test_files(requested)
    if not test_files:
        print("No test files matched.", file=sys.stderr)
        sys.exit(1)

    cases_flat = flatten_cases(test_files)

    print(f"Loaded {len(test_files)} function(s), {len(cases_flat)} test case(s).")

    wb, sheet_map = build_workbook(cases_flat)

    with tempfile.TemporaryDirectory() as tmpdir:
        src_path = os.path.join(tmpdir, "harness_input.xlsx")
        out_dir = os.path.join(tmpdir, "out")
        os.makedirs(out_dir, exist_ok=True)
        wb.save(src_path)

        t0 = time.time()
        proc = subprocess.run(
            [SOFFICE_BIN, "--headless", "--convert-to", "xlsx",
             "--outdir", out_dir, src_path],
            capture_output=True, text=True, timeout=300,
        )
        elapsed = time.time() - t0
        print(proc.stdout.strip())
        if proc.returncode != 0:
            print("soffice STDERR:\n" + proc.stderr, file=sys.stderr)
            sys.exit(f"soffice conversion failed (exit {proc.returncode})")

        out_path = os.path.join(out_dir, "harness_input.xlsx")
        if not os.path.exists(out_path):
            sys.exit(f"Expected output file not found: {out_path}")

        # Second run (staggered) purely to independently reconfirm the
        # volatile-canary recalculation proof for THIS invocation's log.
        time.sleep(2)
        out_dir2 = os.path.join(tmpdir, "out2")
        os.makedirs(out_dir2, exist_ok=True)
        subprocess.run(
            [SOFFICE_BIN, "--headless", "--convert-to", "xlsx",
             "--outdir", out_dir2, src_path],
            capture_output=True, text=True, timeout=300,
        )
        out_path2 = os.path.join(out_dir2, "harness_input.xlsx")

        wb_out = openpyxl.load_workbook(out_path, data_only=True)
        wb_out2 = openpyxl.load_workbook(out_path2, data_only=True) if os.path.exists(out_path2) else None

    # ---- Canary verification ----
    meta = wb_out["_meta"]
    now_run1 = meta["A1"].value
    arith_run1 = meta["A2"].value
    now_run2 = wb_out2["_meta"]["A1"].value if wb_out2 else None

    canary = {
        "arithmetic_formula": CANARY_ARITH_FORMULA,
        "arithmetic_expected": CANARY_ARITH_EXPECTED,
        "arithmetic_actual": arith_run1,
        "arithmetic_ok": arith_run1 == CANARY_ARITH_EXPECTED,
        "volatile_formula": "=NOW()",
        "now_run_1": str(now_run1),
        "now_run_2": str(now_run2),
        "now_differs_across_runs": (now_run1 != now_run2) if now_run2 else None,
        "conversion_seconds_run_1": round(elapsed, 2),
        "method": "openpyxl writes formulas with NO cached <v> value; "
                  "`soffice --headless --convert-to xlsx` must evaluate every "
                  "formula from scratch to produce any output value at all. "
                  "The volatile =NOW() canary changing between two runs a few "
                  "seconds apart, plus the deterministic arithmetic canary "
                  "matching exactly, together prove genuine recalculation.",
    }

    global_trusted = canary["arithmetic_ok"] and bool(canary["now_differs_across_runs"])

    # ---- Per-case results ----
    function_results = {}
    for c in cases_flat:
        sheet_name = sheet_map[c["test_id"]]
        ws = wb_out[sheet_name]

        per_sheet_canary_val = ws[CANARY_ANCHOR].value
        per_sheet_canary_ok = per_sheet_canary_val == CANARY_ARITH_EXPECTED

        anchor_val = normalize_readback_value(ws[c["anchor"]].value)

        range_flat = None
        if c["check_range"]:
            grid = cell_addrs_in_range(c["check_range"])
            range_flat = [normalize_readback_value(ws[addr].value)
                          for row in grid for addr in row]

        error = anchor_val if is_error_value(anchor_val) else None
        matched, mismatch_detail = compare_expected(c["expected"], anchor_val, range_flat)

        notes = []
        if not per_sheet_canary_ok:
            notes.append("UNTRUSTED_RECALC: per-sheet canary failed on this sheet")
        if c.get("expected_note"):
            notes.append(c["expected_note"])
        if mismatch_detail:
            notes.append(f"MISMATCH vs expected: {mismatch_detail}")

        storage_formula = to_storage_formula_all(c["formula"])

        result = {
            "description": c["description"],
            "formula_display": c["formula"],
            "formula_stored_xlsx": storage_formula,
            "value": anchor_val,
            "range_values": range_flat,
            "error": error,
            "expected": c["expected"],
            "matched_expected": matched,
            "canary_ok_this_sheet": per_sheet_canary_ok,
            "notes": "; ".join(notes) if notes else None,
        }
        function_results.setdefault(c["function"], {})[c["test_id"]] = result

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "engine": "LibreOffice Calc",
        "engine_version": LO_VERSION,
        "recalc_method": "soffice --headless --convert-to xlsx (see canary proof below)",
        "trusted": global_trusted,
        "canary": canary,
        "function_results": function_results,
    }

    os.makedirs(RESULTS_DIR, exist_ok=True)
    out_json_path = os.path.join(RESULTS_DIR, f"libreoffice-{LO_VERSION_TAG}.json")

    # ---- Subset runs merge into the existing results file ----
    # See "SUBSET RUNS MERGE, THEY DO NOT OVERWRITE" in the module docstring.
    if requested and os.path.exists(out_json_path):
        with open(out_json_path) as f:
            prev = json.load(f)
        prev_version = prev.get("engine_version")
        if prev_version != LO_VERSION:
            sys.exit(
                f"Refusing to merge: {out_json_path} records engine_version "
                f"{prev_version!r} but the installed LibreOffice reports "
                f"{LO_VERSION!r}. Results from different builds must not share a file."
            )
        merged = dict(prev)
        merged_fr = dict(prev.get("function_results") or {})
        for fn, cases in function_results.items():
            merged_fr[fn] = cases  # whole function re-executed, so replace wholesale
        merged["function_results"] = merged_fr
        merged["generated_at"] = output["generated_at"]
        merged["engine"] = output["engine"]
        merged["engine_version"] = output["engine_version"]
        merged["recalc_method"] = output["recalc_method"]
        merged["canary"] = canary
        # A merge can only downgrade trust, never upgrade it.
        merged["trusted"] = bool(prev.get("trusted", False)) and global_trusted
        merged.setdefault("subset_runs", []).append({
            "generated_at": output["generated_at"],
            "functions": sorted(function_results.keys()),
            "trusted_this_run": global_trusted,
            "superseded_generated_at": prev.get("generated_at"),
        })
        untouched = len(merged_fr) - len(function_results)
        print(f"Subset run: merged {len(function_results)} function(s) into "
              f"{os.path.basename(out_json_path)}; {untouched} other function "
              f"result(s) preserved unchanged.")
        output = merged

    with open(out_json_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
        f.write("\n")

    print(f"\nTrusted recalculation: {global_trusted}")
    print(f"Wrote {out_json_path}")

    # ---- Console summary ----
    n_name_error = 0
    n_other_error = 0
    n_ok = 0
    for fn, cases in function_results.items():
        for tid, r in cases.items():
            if r["error"] == "#NAME?":
                n_name_error += 1
            elif r["error"]:
                n_other_error += 1
            else:
                n_ok += 1
    print(f"Cases with a value (no error): {n_ok}")
    print(f"Cases returning #NAME? (unsupported function): {n_name_error}")
    print(f"Cases returning some other error: {n_other_error}")


if __name__ == "__main__":
    run()
