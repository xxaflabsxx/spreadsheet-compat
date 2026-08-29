"""
Shared corpus + workbook machinery used by every engine runner.

WHY THIS MODULE EXISTS
----------------------
`harness/run_lo.py` (LibreOffice) and `harness/run_sheets.py` (Google
Sheets, via Drive import) must build IDENTICAL workbooks from IDENTICAL
test-case loading logic and must compare read-back values with IDENTICAL
normalization + comparison rules. If the two runners drifted apart even
slightly -- a different sheet-name sanitizer, a different anchor cell, a
looser float tolerance -- then a "difference" reported between LibreOffice
and Google Sheets could be an artifact of the harness rather than a real
engine difference, which would silently poison the entire dataset.

So all of it lives here, once, and both runners import it. Nothing in this
module knows anything about a specific engine: it turns data/tests/*.json
into a workbook and turns read-back cell values into verdicts. Making the
engine actually recalculate, and proving that it did, is each runner's own
job.

This module was factored OUT of run_lo.py without changing any behaviour:
the functions below are the run_lo.py originals verbatim (docstrings
included). run_lo.py imports them and re-exports them under their original
module-level names, so `run_lo.build_workbook(...)` etc. still resolve.
"""
import glob
import json
import os
import re
from datetime import datetime

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlfn_map import to_storage_formula_all  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TESTS_DIR = os.path.join(REPO_ROOT, "data", "tests")
RESULTS_DIR = os.path.join(REPO_ROOT, "results")

# The Excel/OOXML error vocabulary. This set is byte-for-byte the one
# run_lo.py has always used -- do not add engine-specific error tokens here
# or a previously "some other error" LO result could silently reclassify.
KNOWN_ERROR_STRINGS = {
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
    "#GETTING_DATA", "#CALC!", "#SPILL!", "#FIELD!", "#UNKNOWN!",
    "#BLOCKED!", "#CONNECT!", "#BUSY!",
}

# Additional error tokens only Google Sheets ever produces. Sheets emits
# #ERROR! for a formula it cannot even parse (its analogue of a parse
# failure, distinct from Excel's #NAME? for an unknown function name).
# Sheets has no #CALC!: it returns plain #N/A where Excel 365 would say
# #CALC! (e.g. FILTER matching no rows), and #N/A is already in the set
# above, so nothing else needs translating. Passed explicitly by
# run_sheets.py so the LO path's classification is untouched.
SHEETS_EXTRA_ERROR_STRINGS = {"#ERROR!"}

CANARY_ARITH_FORMULA = "=1111+2222"
CANARY_ARITH_EXPECTED = 3333
CANARY_ANCHOR = "Z1"  # far from any setup_cells/check_range used in data/tests

# Sheet-name characters rejected by BOTH Excel/OOXML and Google Sheets.
_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:\*\?/\\]")
# We keep the OOXML 31-char cap (not Sheets' more generous 100) because the
# transport format is .xlsx in both directions: a >31-char sheet name is
# invalid OOXML regardless of which engine eventually opens the file.
SHEET_NAME_MAX = 31


def load_test_files(names=None):
    """Return list of (function_name, filepath, payload)."""
    files = sorted(glob.glob(os.path.join(TESTS_DIR, "*.json")))
    out = []
    for path in files:
        fn = os.path.splitext(os.path.basename(path))[0]
        if names and fn not in names:
            continue
        with open(path) as f:
            payload = json.load(f)
        out.append((fn, path, payload))
    return out


def sanitize_sheet_name(name, used):
    """Excel/LO sheet names: <=31 chars, no []:*?/\\, must be unique.

    The same output is also valid in Google Sheets, whose rules are a
    superset (<=100 chars, same forbidden character class) -- see
    assert_sheets_safe_name().
    """
    clean = _ILLEGAL_SHEET_CHARS.sub("_", name)[:SHEET_NAME_MAX]
    base = clean
    i = 2
    while clean.lower() in used:
        suffix = f"~{i}"
        clean = base[: SHEET_NAME_MAX - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def assert_sheets_safe_name(name):
    """Raise ValueError if `name` would be rejected/mangled by Google Sheets.

    Sheets' documented limits: name is non-empty, at most 100 characters,
    and must not contain any of [ ] * ? / \\ : . A leading or trailing
    apostrophe also breaks A1 references of the form 'Sheet name'!A1, so we
    reject those too. Returns the name unchanged when it is safe, so it can
    be used inline.
    """
    if not name or not name.strip():
        raise ValueError("sheet name is empty or whitespace-only")
    if len(name) > 100:
        raise ValueError(f"sheet name too long for Google Sheets (>100): {name!r}")
    bad = _ILLEGAL_SHEET_CHARS.findall(name)
    if bad:
        raise ValueError(f"sheet name contains characters Google Sheets rejects "
                         f"{sorted(set(bad))}: {name!r}")
    if name.startswith("'") or name.endswith("'"):
        raise ValueError(f"sheet name starts/ends with an apostrophe, which "
                         f"breaks A1 references: {name!r}")
    return name


def build_workbook(cases_flat, plain_names=False):
    """
    cases_flat: list of dicts with keys:
        test_id, function, formula (original), setup_cells, check_range
    plain_names: if True, write each formula EXACTLY as authored in
        data/tests -- no _xlfn./_xlfn._xlws. storage-form translation at
        all. Default False preserves the original (and run_lo.py's only)
        behaviour byte-for-byte; run_lo.py always calls this positionally
        with one argument, so it is unaffected by this parameter existing.
        See run_sheets.py's `--plain-names` build flag for why this exists:
        Google Sheets' xlsx importer maps plain `_xlfn.NAME` but not
        `_xlfn._xlws.FILTER/SORT`, so those (and LAMBDA-family functions)
        need to be tested with their bare, natural-language name instead.
    Returns (workbook, sheet_map) where sheet_map[test_id] -> sheet_name
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()
    sheet_map = {}

    for c in cases_flat:
        sheet_name = sanitize_sheet_name(c["test_id"], used_names)
        sheet_map[c["test_id"]] = sheet_name
        ws = wb.create_sheet(sheet_name)

        for addr, val in (c.get("setup_cells") or {}).items():
            ws[addr] = val

        # Prefix EVERY known future-function call site (not just the function
        # under test): nested modern calls like UNICHAR(UNICODE(...)) need
        # both names prefixed or the whole formula is #NAME? on all engines.
        storage_formula = c["formula"] if plain_names else to_storage_formula_all(c["formula"])
        anchor = c["anchor"]
        if c.get("check_range"):
            # Functions expected to return a multi-cell array (spill/dynamic
            # array results) are written as a legacy Ctrl+Shift+Enter style
            # array formula covering the full check_range. This matters:
            # older engines (and pre-365 Excel) only spill a range result
            # when the formula is explicitly marked as an array formula --
            # writing it as a plain scalar formula string causes even a
            # SUPPORTED function like INDEX(range,0,col) to return #VALUE!
            # instead of spilling. Verified empirically: wrapping
            # INDEX(A1:C3,0,2) in ArrayFormula(ref="A30:A32") makes
            # LibreOffice 24.2 correctly spill [2,5,8]; the identical
            # formula as a plain string returns #VALUE!.
            ws[anchor] = ArrayFormula(c["check_range"], storage_formula)
        else:
            ws[anchor] = storage_formula

        # Canary: deterministic, non-cacheable arithmetic on every sheet.
        ws[CANARY_ANCHOR] = CANARY_ARITH_FORMULA

    # Dedicated meta sheet with a volatile canary for cross-run recalc proof.
    meta = wb.create_sheet("_meta", 0)
    meta["A1"] = "=NOW()"
    meta["A2"] = CANARY_ARITH_FORMULA

    return wb, sheet_map


def anchor_for_case(case):
    if case.get("check_range"):
        # anchor is the top-left cell of the check range
        first = case["check_range"].split(":")[0]
        return first
    return "F1"


def flatten_cases(test_files):
    """Turn load_test_files() output into the flat case dicts build_workbook
    consumes. Both runners must derive their case list the same way, so the
    loop lives here rather than being copy-pasted per runner."""
    cases_flat = []
    for fn, path, payload in test_files:
        for case in payload["cases"]:
            cases_flat.append({
                "test_id": case["id"],
                "function": fn,
                "formula": case["formula"],
                "setup_cells": case.get("setup_cells"),
                "check_range": case.get("check_range"),
                "expected": case.get("expected"),
                "expected_note": case.get("expected_note"),
                "description": case["description"],
                "anchor": anchor_for_case(case),
            })
    return cases_flat


def cell_addrs_in_range(range_str):
    """Expand 'A30:C32' into a row-major list of lists of addresses."""
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    rows = []
    for r in range(min_row, max_row + 1):
        row = []
        for c in range(min_col, max_col + 1):
            row.append(f"{get_column_letter(c)}{r}")
        rows.append(row)
    return rows


def is_error_value(v, extra=None):
    """True if `v` is a cached error string. `extra` adds engine-specific
    error tokens (see SHEETS_EXTRA_ERROR_STRINGS) without widening the
    shared Excel/OOXML vocabulary for every other runner."""
    if not isinstance(v, str):
        return False
    return v in KNOWN_ERROR_STRINGS or (extra is not None and v in extra)


EXCEL_EPOCH = datetime(1899, 12, 30)  # serial 0 in the 1900 date system


def normalize_readback_value(v):
    """
    Normalize a value read back from the recalculated .xlsx into the same
    domain the test corpus's `expected` values live in.

    - datetime/date/time objects -> Excel serial numbers. LibreOffice
      applies a date/time NUMBER FORMAT to the result cells of DATE()/
      TIME()-style formulas; openpyxl then surfaces the cached value as a
      Python datetime/time object instead of the underlying float serial.
      (Google Sheets' .xlsx export does the same thing for the same
      reason.) The engine's actual computed value IS the serial -- the
      datetime-ness is presentation, so converting back to the serial is
      the faithful raw value, not an interpretation. (Excel 1900 system:
      1899-12-30 = 0. This intentionally reproduces Excel's day-59/60
      Feb-29-1900 compatibility offset for all post-1900-03-01 dates, which
      is every date used in this corpus.)
    """
    import datetime as _dt

    if isinstance(v, _dt.datetime):
        delta = v - EXCEL_EPOCH
        return delta.days + delta.seconds / 86400 + delta.microseconds / 86400e6
    if isinstance(v, _dt.date):
        return (
            _dt.datetime(v.year, v.month, v.day) - EXCEL_EPOCH
        ).days
    if isinstance(v, _dt.time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 86400 + v.microsecond / 86400e6
    return v


def values_roughly_equal(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) and not isinstance(a, bool) and not isinstance(b, bool):
        return abs(a - b) < 1e-9
    # .xlsx storage limitation: a formula legitimately returning the empty
    # string "" round-trips through file conversion as a cell with no cached
    # value at all, which openpyxl reads back as None. Blank-vs-empty-string
    # is genuinely indistinguishable at this layer, so an expected "" is
    # satisfied by a read-back None. (The raw None is still recorded in the
    # results file; only the match verdict treats them as equivalent.)
    if a == "" and b is None or b == "" and a is None:
        return True
    return a == b


def compare_expected(expected, actual_anchor, actual_range_flat):
    """Returns (matched: bool or None, detail: str or None)."""
    if expected is None:
        return None, None
    if isinstance(expected, list):
        if actual_range_flat is None:
            return False, "expected a range of values but no check_range was read"
        # flatten expected (may be nested for 2D)
        flat_expected = []
        for item in expected:
            if isinstance(item, list):
                flat_expected.extend(item)
            else:
                flat_expected.append(item)
        flat_actual = actual_range_flat
        if len(flat_expected) != len(flat_actual):
            return False, f"length mismatch: expected {len(flat_expected)} values, got {len(flat_actual)}"
        for e, a in zip(flat_expected, flat_actual):
            if not values_roughly_equal(e, a):
                return False, f"value mismatch: expected {e!r}, got {a!r}"
        return True, None
    else:
        matched = values_roughly_equal(expected, actual_anchor)
        detail = None if matched else f"expected {expected!r}, got {actual_anchor!r}"
        return matched, detail
