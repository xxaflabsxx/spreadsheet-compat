#!/usr/bin/env python3
"""
Google Sheets engine runner for the spreadsheet function-compatibility
harness.

WHAT THIS DOES (AND WHAT IT DELIBERATELY DOES NOT DO)
-----------------------------------------------------
There is no headless Google Sheets binary to drive, and this script has no
Google credentials of its own. What it does have is the one property that
makes Google Sheets executable at all from here:

  Uploading a formula-only .xlsx to Google Drive with the Google-Sheets
  target MIME type auto-converts it into a real Google Sheet, and that
  conversion RECALCULATES every formula with Google's own calculation
  engine. Exporting that Sheet back out as .xlsx produces a workbook whose
  cells carry Google Sheets' computed cached values.

So this runner is split in half around an external upload/download step
performed by whoever/whatever has Drive access (the orchestrator):

    build   -- emit N small .xlsx chunk workbooks + a manifest.json
               (upload these to Drive, let Drive convert them, export each
               back as .xlsx)
    ingest  -- read the exported .xlsx files back, map every cell to its
               test case via the manifest, and write
               results/google-sheets.json in exactly the schema
               results/libreoffice-*.json uses.

The full orchestrator loop is documented in README.md, "Phase 2: Google
Sheets runner".

WHY THE FILES ARE CHUNKED
-------------------------
The .xlsx bytes travel through the orchestrator's context window as base64
(~4/3 size inflation), so one 841-sheet ~430 KB workbook is not
transportable. `build` therefore splits the corpus by FUNCTION (never
splitting a function's cases across two files) into chunks of ~40 functions,
which measure ~60 KB each -- comfortably under the ~150 KB per-workbook
budget.

WHY THE RESULTS ARE STILL TRUSTWORTHY (THE CANARY, SAME AS THE LO RUNNER)
-------------------------------------------------------------------------
openpyxl NEVER writes a cached <v> value for a formula cell -- it writes
only the formula string. The uploaded chunk therefore contains zero cached
values. That gives us the same proof the LibreOffice runner relies on:

  (a) Deterministic canary: `=1111+2222` sits in Z1 of EVERY sheet. If
      Google Sheets had not recalculated, the exported cell would come back
      blank (None), because there was never a cached value to preserve.
      Reading back exactly 3333 proves the cell was computed by Google.
  (b) Volatile canary: `=NOW()` on the `_meta` sheet. Sheets evaluates it
      at import time, so a plausible fresh timestamp corroborates (a).
      NOTE the honest limitation: unlike the LO runner, which converts the
      same file twice a few seconds apart and shows =NOW() DIFFERING
      between runs, a single Drive import gives us one timestamp only. The
      cross-run volatile proof is therefore weaker here; the deterministic
      canary is the load-bearing one, and it is checked on every single
      sheet.

If any deterministic canary fails, the whole run is marked "trusted": false
and each affected case gets an "UNTRUSTED_RECALC" note.

WHAT WE CANNOT HONESTLY CLAIM
-----------------------------
Google Sheets has no user-visible version number and exposes none through
Drive. So `engine_version` here is NOT a version -- it is a dated label
("Google Sheets (Drive import, YYYY-MM-DD)") recording WHEN the corpus was
executed. Sheets is a continuously-updated hosted product; a result from
this file means "this is what Google Sheets did on that date", and nothing
stronger. Anything presenting Sheets results must say so.

THE _xlfn. PREFIX (VERIFIED, DO NOT "FIX")
------------------------------------------
Google Sheets DOES understand the OOXML `_xlfn.` storage prefix on import:
verified empirically -- `_xlfn.XLOOKUP(...)` imported and computed a real
value (1), it did not come back as #NAME?. So this runner uses the exact
same harness/xlfn_map.py translation as the LibreOffice path, with no
Sheets-specific special-casing. Writing the bare unprefixed name instead
would be the thing that breaks.

USAGE
-----
    # 1. build the chunk workbooks (all 278 functions, ~40 per chunk)
    python3 harness/run_sheets.py build
    python3 harness/run_sheets.py build --chunk-size 25
    python3 harness/run_sheets.py build --only XLOOKUP LET FILTER
    python3 harness/run_sheets.py build --outdir harness/sheets_chunks

    # 2. (external) upload each chunk to Drive as a Google Sheet, export it
    #    back as .xlsx into harness/sheets_exports/chunk-NN-export.xlsx

    # 3. ingest the exports (incremental: run it per chunk as they arrive)
    python3 harness/run_sheets.py ingest \
        --export harness/sheets_exports/chunk-01-export.xlsx \
        --engine-label "Google Sheets (Drive import, 2026-08-29)"

    # dry run of the whole pipeline with LibreOffice standing in for Drive,
    # writing to a scratch results file (never results/google-sheets.json):
    python3 harness/run_sheets.py selftest --only COUNT SUM MROUND

    # ---- a SECOND executed engine: EXCEL FOR THE WEB ----
    # The chunk workbooks `build` emits are engine-NEUTRAL -- formula-only
    # .xlsx with zero cached values -- so anything that recalculates on open
    # can execute them. Excel for the web does, via a different round trip:
    #
    #   upload chunk-NN.xlsx to OneDrive -> open it in Excel for the web
    #   (which recalculates on open) -> File > Create a Copy > Download a
    #   Copy -> the downloaded .xlsx carries Excel's computed cached values.
    #
    # VERIFIED: this path needs the DEFAULT xlfn-translated serialization,
    # NOT --plain-names. Excel for the web resolves the `_xlfn.` storage
    # prefix natively (`_xlfn.PERCENTRANK.INC(...)` came back computed, not
    # #NAME?) -- which is unsurprising, since that prefix is Excel's own
    # storage convention.
    #
    # The engine is declared at INGEST time, because that is where the
    # results IDENTITY is chosen:
    python3 harness/run_sheets.py ingest \
        --export ~/Downloads/chunk-01.xlsx \
        --chunkdir harness/excel_probe_chunk \
        --out results/excel-web.json \
        --engine excel_web
    #
    # --engine selects the results identity, the error vocabulary (Google's
    # #ERROR! parse-failure token does NOT apply here), the canary prose, and
    # the default --out/--chunkdir/--engine-label. Merging results from one
    # engine into another engine's file is refused outright --
    # --allow-label-change covers a DATE change on ONE engine and nothing
    # more.
    #
    # WHAT THIS IS NOT: Excel for the web is not desktop Excel. It is a
    # different implementation of the calculation engine, and this corpus's
    # Excel column remains DOCUMENTED behaviour for the DESKTOP product. An
    # executed Excel-web result is its own evidence and must never be
    # rendered as though desktop Excel produced it. results/excel-web.json is
    # deliberately NOT published by the site yet for exactly that reason --
    # see excel-web-site-plan.md.
    #
    # What the export does to values on the way out was MEASURED from the
    # probe download rather than assumed, and is recorded verbatim in the
    # results file under "readback_artifacts" (see
    # EXCEL_WEB_READBACK_ARTIFACTS below). Headline: unlike Sheets' export,
    # which rounds to 10 significant digits and is LOSSY, Excel-web writes
    # full IEEE-754 round-trip decimal, so no normalization is applied.

    # dry run of the excel_web path: builds its own fixture, synthesizes the
    # export, and asserts the identity + merge discipline. Executes NOTHING.
    python3 harness/run_sheets.py selftest-excel-web

    # ---- the how-to RECIPE corpus (data/recipes/*.json) ----
    # Same three-step loop, one worksheet per recipe check, writing
    # results/recipes-verified-sheets.json alongside the LibreOffice-executed
    # results/recipes-verified.json.

    # 1. build (plain function names by default -- Sheets maps unprefixed
    #    names, and recipes are authored the way a user types them)
    python3 harness/run_sheets.py build-recipes
    python3 harness/run_sheets.py build-recipes --chunk-size 60
    python3 harness/run_sheets.py build-recipes --only how-to-use-xlookup add-days-to-a-date
    python3 harness/run_sheets.py build-recipes --outdir harness/recipe_chunks
    python3 harness/run_sheets.py build-recipes --xlfn-names   # opt out of plain names
    #    checks carrying "engines": [...] are filtered to the engine the
    #    chunks are built FOR (google_sheets by default; --engine to change it)

    # 2. (external) upload each chunk to Drive, export back as
    #    harness/recipe_exports/chunk-NN-export.xlsx

    # 3. ingest (incremental, per chunk as exports land)
    python3 harness/run_sheets.py ingest-recipes \
        --export harness/recipe_exports/chunk-01-export.xlsx \
        --out results/recipes-verified-sheets.json

    # dry run of the recipe pipeline with LibreOffice standing in for Drive;
    # every value must match results/recipes-verified.json or it exits 1
    python3 harness/run_sheets.py selftest-recipes

    # ---- the six MULTI-SHEET recipes (one workbook per check) ----
    # `build-recipes` cannot carry the recipes whose checks declare
    # setup_sheets: their formulas name the data tabs literally, two checks
    # want a tab called Data with different contents, one names a tab that
    # must NOT exist, and a 3-D reference depends on tab ORDER. So each CHECK
    # gets its own workbook holding exactly the tabs it asks for -- nothing
    # renamed, nothing rewritten -- at one Drive round-trip per check.

    # 1. build (34 workbooks across the 6 recipes)
    python3 harness/run_sheets.py build-recipes-multisheet
    python3 harness/run_sheets.py build-recipes-multisheet --only vlookup-from-another-sheet

    # 2. (external) upload every harness/recipe_chunks_multisheet/ms-*.xlsx to
    #    Drive with convert-on, then download each converted Sheet back as
    #    .xlsx into harness/recipe_exports_multisheet/ -- keep the names, they
    #    already carry the ms-NNN id the manifest is keyed by

    # 3. ingest the whole folder (the manifest's "mode" selects this reader)
    python3 harness/run_sheets.py ingest-recipes \
        --chunkdir harness/recipe_chunks_multisheet \
        --export-dir harness/recipe_exports_multisheet \
        --engine-label "Google Sheets (Drive import, YYYY-MM-DD)"

    # If Drive's importer RENAMES a tab (observed: `Jon's Data` came back as
    # `Jons Data`, apostrophe stripped, with Sheets rewriting the formula to
    # match), ingest identifies it by reading its setup literals back, records
    # BOTH names verbatim in the check's notes under SETUP_ALTERED, and keeps
    # that check out of the recipe's verdict -- Sheets executed a workbook the
    # LibreOffice reference run never saw, so the two are not comparable.
    # Nothing is renamed back. A tab that is genuinely gone, or that no
    # leftover tab's literals can identify, still refuses the workbook.

    # dry run: build + expected values synthesized into the builder's own
    # workbooks + ingest, asserting the mapping, the keyed merge, the
    # engine-label guard, canary detection and the partial-download refusal.
    # Executes no engine at all -- no LibreOffice, no Drive.
    python3 harness/run_sheets.py selftest-recipes-multisheet

INCREMENTAL INGESTION MERGES, IT DOES NOT OVERWRITE
---------------------------------------------------
Every ingest is inherently a subset (one or more chunks out of N), so if
--out already exists the new results are MERGED in with exactly the
semantics run_lo.py uses for subset runs: only the functions actually
ingested this time are replaced, every other function's previously executed
result is preserved byte-for-byte, generated_at / canary / recalc_method are
refreshed to describe the most recent execution, and `trusted` becomes the
AND of the previous file's flag and this run's -- a merge can only ever
downgrade trust, never launder an untrusted ingest into a trusted file. Each
merge appends to a top-level "subset_runs" list for auditability. Each
function block ingested is also stamped with its own "executed_at" UTC date
(see harness/results_schema.py); functions the merge does not touch keep the
date of the run that actually produced them, so refreshing the file-level
generated_at can never re-date the rest of the corpus. Merging
into a file recorded under a DIFFERENT engine label is refused unless
--allow-label-change is passed (results executed on two different dates
against a silently-updated hosted product should not be blended without a
deliberate decision; when it is allowed, both labels are recorded).
"""
import argparse
import glob
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl  # noqa: E402

from openpyxl.worksheet.formula import ArrayFormula  # noqa: E402

from corpus import (  # noqa: E402
    RESULTS_DIR,
    CANARY_ANCHOR,
    CANARY_ARITH_EXPECTED,
    CANARY_ARITH_FORMULA,
    SHEET_NAME_MAX,
    SHEETS_EXTRA_ERROR_STRINGS,
    assert_sheets_safe_name,
    build_workbook,
    cell_addrs_in_range,
    compare_expected,
    flatten_cases,
    is_error_value,
    load_test_files,
    normalize_readback_value,
    sanitize_sheet_name,
)
# The RECIPE corpus (data/recipes/*.json) shares its enumeration, setup
# inheritance, normalization and comparison rules with
# scripts/verify_recipes.py -- see harness/recipe_corpus.py.
from recipe_corpus import (  # noqa: E402
    GOOGLE_SHEETS,
    LIBREOFFICE,
    compare_check,
    iter_checks,
    load_recipe_files,
    norm,
    result_checks_by_key,
    setup_sheet_names,
    uses_setup_sheets,
)
from xlfn_map import to_storage_formula_all  # noqa: E402
from results_schema import function_cases, stamp_executed_at  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_CHUNK_DIR = os.path.join(REPO_ROOT, "harness", "sheets_chunks")
DEFAULT_EXPORT_DIR = os.path.join(REPO_ROOT, "harness", "sheets_exports")
DEFAULT_RESULTS_PATH = os.path.join(RESULTS_DIR, "google-sheets.json")
# Scratch output for `selftest`. Deliberately OUTSIDE results/: it holds
# LibreOffice values, and results/ is only ever real engine output.
SELFTEST_RESULTS_PATH = os.path.join(
    REPO_ROOT, "harness", "sheets_selftest", "plumbing-check.json")
MANIFEST_NAME = "manifest.json"

ENGINE_ID = "google_sheets"
RECALC_METHOD = "Drive import + xlsx export readback"

# --------------------------------------------------------------------------
# INGEST ENGINE REGISTRY
# --------------------------------------------------------------------------
# `build` emits engine-neutral chunk workbooks: formula-only .xlsx with no
# cached values anywhere. That same artifact is executable by ANY hosted
# spreadsheet that recalculates on import, so `ingest` -- not `build` -- is
# where the engine identity is chosen, via `--engine`.
#
# NOTE the pre-existing `--engine` flag on build-recipes/-multisheet is a
# DIFFERENT axis: it selects which engine's checks to INCLUDE (recipe checks
# may carry an "engines" scope). It never sets a results-file identity. The
# flag added here to `ingest` sets the identity of the results file being
# written, so the two do not overlap and are deliberately not merged.
EXCEL_WEB_ENGINE_ID = "excel_web"
EXCEL_WEB_RESULTS_PATH = os.path.join(RESULTS_DIR, "excel-web.json")
EXCEL_WEB_RECALC_METHOD = (
    "OneDrive upload + Excel for the web recalculation on open + "
    "File > Create a Copy > Download a Copy readback")

# Excel for the web's .xlsx export writes the ordinary OOXML error vocabulary
# in typed (t="e") cells. It has no counterpart to Google's #ERROR! parse
# failure, so the Sheets-specific extra token is NOT widened to this engine:
# if "#ERROR!" ever appeared in an Excel-web export it would be a string
# result, not an error cell, and must not be silently reclassified.
EXCEL_WEB_EXTRA_ERROR_STRINGS = frozenset()

_SHEETS_CANARY_METHOD = (
    "openpyxl writes formulas with NO cached <v> value, so the uploaded "
    "chunk workbooks contain zero cached results. Google Drive's "
    "auto-conversion to a Google Sheet recalculates every formula with "
    "Google's engine; exporting that Sheet back to .xlsx carries those "
    "computed values out. The deterministic canary =1111+2222 in "
    f"{CANARY_ANCHOR} of EVERY sheet reading back exactly "
    f"{CANARY_ARITH_EXPECTED} therefore proves genuine computation -- "
    "without recalculation the cell would read back blank (None), since "
    "nothing was ever cached. The volatile =NOW() canary is recorded per "
    "chunk as corroboration, but UNLIKE the LibreOffice runner (which "
    "converts the same file twice and shows =NOW() differing) a single "
    "Drive import yields one timestamp, so no cross-run volatile "
    "comparison is possible: now_differs_across_runs is null by design, "
    "and the deterministic canary is the load-bearing proof here.")

_EXCEL_WEB_CANARY_METHOD = (
    "openpyxl writes formulas with NO cached <v> value, so the uploaded "
    "chunk workbooks contain zero cached results. Opening the workbook in "
    "Excel for the web from OneDrive recalculates every formula with "
    "Microsoft's own web calculation engine; File > Create a Copy > "
    "Download a Copy writes those computed values back out as .xlsx. The "
    f"deterministic canary =1111+2222 in {CANARY_ANCHOR} of EVERY sheet "
    f"reading back exactly {CANARY_ARITH_EXPECTED} therefore proves genuine "
    "computation -- without recalculation the cell would read back blank "
    "(None), since nothing was ever cached. The volatile =NOW() canary is "
    "recorded per chunk as corroboration; like the Drive path (and unlike "
    "the LibreOffice runner, which converts the same file twice and shows "
    "=NOW() differing) a single open yields one timestamp, so "
    "now_differs_across_runs is null by design and the deterministic canary "
    "is the load-bearing proof. Corroborating this engine specifically: the "
    "downloaded package's docProps/app.xml names the application that wrote "
    "it, and is recorded verbatim per chunk under app_provenance_per_chunk.")

# WHAT THE EXPORT DOES TO VALUES ON THE WAY OUT -- MEASURED, NOT ASSUMED.
# Every statement below was measured from the validated probe download
# (6 functions / 34 cases, harness/excel_probe_chunk/) by reading the raw
# worksheet XML, NOT by trusting openpyxl's parse. Anything the probe did not
# exercise is recorded as unmeasured rather than guessed at.
EXCEL_WEB_READBACK_ARTIFACTS = {
    "measured_from": "harness/excel_probe_chunk/ (6 functions, 34 cases)",
    "measured_on": "2026-09-01",
    "method": ("raw xl/worksheets/*.xml <v> strings inspected directly, then "
               "compared against Python's float() parse of the same strings"),
    "float_precision": (
        "LOSSLESS -- no normalization applied or needed. Excel for the web "
        "writes full IEEE-754 round-trip decimal, up to 17 significant "
        "digits, and uses scientific notation for small magnitudes: "
        "ROUND(MIRR(...),4) came back as the literal string "
        "'0.12609999999999999' and ROUND(MIRR(...),4) = -0.048 as "
        "'-4.8000000000000001E-2'. Every such string was verified to parse to "
        "the IDENTICAL double as its short form, so float() recovers the "
        "exact value and Python's repr renders it back as 0.1261 / -0.048. "
        "This is the OPPOSITE of Google Sheets' export, which is LOSSY: it "
        "rounds to 10 significant digits (PI() comes back as 3.141592654), "
        "genuinely destroying precision. Excel-web values therefore need no "
        "rounding fix-up, and none is applied -- the raw parsed double is "
        "recorded."),
    "errors": (
        "Standard OOXML typed error cells: t=\"e\" with the error token as "
        "the cached value (#DIV/0!, #VALUE!, #NUM!, #N/A all observed "
        "verbatim). Recorded as-is, exactly like the LibreOffice and Sheets "
        "runners. No Google-style #ERROR! parse-failure token exists here."),
    "date_formatting_of_numerics": (
        "UNMEASURED. Sheets applies a date/time NUMBER FORMAT to DATE()/"
        "TIME()-style results, which makes openpyxl surface a datetime "
        "instead of the serial (normalize_readback_value converts it back). "
        "The probe carried NO date or time function -- every one of its 34 "
        "result cells came back with number format 'General' and the "
        "workbook declares no custom numFmts at all -- so this run cannot "
        "say whether Excel-web does the same. normalize_readback_value is "
        "still applied on the same terms as every other engine, so a "
        "datetime WOULD be handled correctly if one appears; it simply was "
        "not exercised here."),
    "empty_string_results": (
        "UNMEASURED. No probe case returns the empty string, so whether a "
        "formula yielding \"\" round-trips as a blank cell (None) is "
        "untested for this engine. values_roughly_equal() already treats an "
        "expected \"\" as satisfied by a read-back None for every engine, so "
        "the existing handling applies unchanged."),
    "booleans": (
        "UNMEASURED AS A RESULT. The probe contains no boolean-returning "
        "formula. A boolean INPUT literal did survive as a real typed "
        "t=\"b\" cell rather than the text 'TRUE' -- weak evidence that "
        "Excel-web would not stringify a boolean the way Sheets' export "
        "sometimes does -- but an input is not a computed result and no "
        "claim is drawn from it. The _boolean_text_artifact note still fires "
        "for this engine if it ever happens."),
    "formula_prefix": (
        "The _xlfn. storage prefix survives the round trip verbatim: "
        "'_xlfn.PERCENTRANK.INC(A2:A11,2)' came back written exactly that "
        "way and COMPUTED (0.333), confirming Excel for the web resolves the "
        "storage form. This is why the probe chunk uses the DEFAULT "
        "xlfn-translated serialization and NOT --plain-names."),
    "volatile_now": (
        "=NOW() comes back as a RAW FLOAT SERIAL under 'General' format "
        "(46266.42749652778), not as a datetime the way Sheets' export "
        "renders it, and it carries the ca=\"1\" volatile flag. It is "
        "evaluated in the VIEWER'S LOCAL TIME ZONE, not UTC: the probe's "
        "serial decodes to 2026-09-01 10:15:35, which is UTC-7 (America/"
        "Phoenix), and the package's own docProps/core.xml records "
        "dcterms:created 17:13:37Z and dcterms:modified 17:16:41Z -- so "
        "10:15:35 local == 17:15:35Z falls exactly inside the recalculation "
        "window. Anything comparing this timestamp to a UTC clock must "
        "account for the offset or it will look ~7 hours stale."),
}

# Per-engine `ingest` behaviour. Everything that legitimately differs between
# a Drive import and an Excel-for-the-web open lives here, so ingest_exports()
# itself stays one code path shared by both.
INGEST_ENGINES = {
    ENGINE_ID: {
        "engine_id": ENGINE_ID,
        "display": "Google Sheets",
        "label_template": "Google Sheets (Drive import, {date})",
        "default_out": DEFAULT_RESULTS_PATH,
        "default_chunkdir": DEFAULT_CHUNK_DIR,
        "recalc_method": RECALC_METHOD,
        "extra_error_strings": SHEETS_EXTRA_ERROR_STRINGS,
        "canary_method": _SHEETS_CANARY_METHOD,
        "readback_artifacts": None,
        "expect_app": None,
    },
    EXCEL_WEB_ENGINE_ID: {
        "engine_id": EXCEL_WEB_ENGINE_ID,
        "display": "Excel for the web",
        "label_template": "Excel for the web (recalc, {date})",
        "default_out": EXCEL_WEB_RESULTS_PATH,
        "default_chunkdir": os.path.join(REPO_ROOT, "harness", "excel_probe_chunk"),
        "recalc_method": EXCEL_WEB_RECALC_METHOD,
        "extra_error_strings": EXCEL_WEB_EXTRA_ERROR_STRINGS,
        "canary_method": _EXCEL_WEB_CANARY_METHOD,
        "readback_artifacts": EXCEL_WEB_READBACK_ARTIFACTS,
        # docProps/app.xml <Application> must look like this or the ingest
        # says so loudly -- see _read_app_provenance().
        "expect_app": "excel online",
    },
}


def _engine_spec(name):
    spec = INGEST_ENGINES.get(name)
    if spec is None:
        sys.exit(f"Unknown ingest engine {name!r}. "
                 f"Known: {', '.join(sorted(INGEST_ENGINES))}.")
    return spec


def _read_app_provenance(path):
    """The application that WROTE this .xlsx, per its docProps/app.xml.

    Excel for the web stamps '<Application>Microsoft Excel Online</Application>'
    plus an AppVersion into every package it produces. That is independent,
    in-band evidence of which engine actually recalculated the workbook --
    something the Google Drive path never offered -- so it is recorded
    verbatim per chunk and checked against the engine being claimed. It is
    NOT proof on its own (a string in a file is forgeable), which is why the
    deterministic canary remains the load-bearing check; this corroborates it
    and, more usefully, catches an honest mistake: ingesting a
    LibreOffice-converted or openpyxl-written file under an Excel-web label.
    """
    import xml.etree.ElementTree as ET
    import zipfile
    try:
        with zipfile.ZipFile(path) as z:
            raw = z.read("docProps/app.xml")
    except (KeyError, zipfile.BadZipFile, OSError):
        return {"application": None, "app_version": None}
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        return {"application": None, "app_version": None}

    def _text(tag):
        for el in root.iter():
            if el.tag.rsplit("}", 1)[-1] == tag:
                return (el.text or "").strip() or None
        return None

    return {"application": _text("Application"),
            "app_version": _text("AppVersion")}

# `selftest` produces LibreOffice values through the Sheets plumbing. The site
# generator discovers engines by globbing results/*.json and matching on the
# "engine" string ("google" anywhere in it means Google Sheets), so selftest
# output MUST NOT carry ENGINE_ID and MUST NOT land in results/ -- either
# alone would be enough for the site to publish LibreOffice numbers in a
# Google Sheets column. Both are enforced in cmd_selftest().
SELFTEST_ENGINE_ID = "SELFTEST_libreoffice_via_sheets_pipeline"
SELFTEST_RECALC_METHOD = ("selftest plumbing check: soffice --convert-to xlsx "
                          "standing in for the Drive import -- NOT Google Sheets")

# Same rule for the Excel-for-the-web selftest: its values are the corpus's
# own expecteds written back into the builder's workbook, so the id must not
# contain a string any consumer maps to a real engine, and the file must not
# land in results/. Both are enforced in cmd_selftest_excel_web().
EXCEL_WEB_SELFTEST_ENGINE_ID = "SELFTEST_synthesized_via_excel_web_pipeline"
EXCEL_WEB_SELFTEST_RECALC_METHOD = (
    "selftest plumbing check: expected values written into the builder's own "
    "workbook standing in for the Excel-for-the-web round trip -- NOTHING was "
    "executed, and this says nothing about Excel for the web")
EXCEL_WEB_SELFTEST_RESULTS_PATH = os.path.join(
    REPO_ROOT, "harness", "excel_web_selftest", "plumbing-check.json")
EXCEL_WEB_SELFTEST_CHUNK_DIR = os.path.join(
    REPO_ROOT, "harness", "excel_web_selftest", "chunks")
# The six functions the real Excel-web probe covered. Reusing them keeps the
# fixture aligned with results/excel-web.json, so the selftest exercises the
# same shapes (a range argument, an empty-range error, a CJK string, an
# _xlfn.-prefixed name) the real run met.
EXCEL_WEB_SELFTEST_FUNCTIONS = (
    "ABS", "FVSCHEDULE", "LENB", "MIRR", "PERCENTRANK.INC", "VAR.P")

# Advisory ceiling per chunk workbook. The bytes travel as base64 through the
# orchestrator's context (~4/3 inflation), so this is about transportability,
# not about any Drive or Sheets limit.
SIZE_WARN_BYTES = 150 * 1024

META_SHEET = "_meta"
META_VOLATILE_CELL = "A1"
META_ARITH_CELL = "A2"


# --------------------------------------------------------------------------
# build
# --------------------------------------------------------------------------

def chunk_functions(cases_flat, chunk_size):
    """Group flat cases by function, then partition the FUNCTIONS (never a
    single function's cases) into chunks of at most `chunk_size` functions.

    Keeping every case of a function in one workbook matters: a function's
    results are merged into the results file wholesale, so a function split
    across two chunks could end up half-ingested and silently truncated."""
    by_fn = {}
    for c in cases_flat:
        by_fn.setdefault(c["function"], []).append(c)
    fns = sorted(by_fn)
    chunks = []
    for i in range(0, len(fns), chunk_size):
        group = fns[i:i + chunk_size]
        chunks.append((group, [c for fn in group for c in by_fn[fn]]))
    return chunks


def _sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(65536), b""):
            h.update(block)
    return h.hexdigest()


def cmd_build(args):
    requested = set(args.only) if args.only else None
    test_files = load_test_files(requested)
    if not test_files:
        sys.exit("No test files matched.")
    if requested:
        missing = requested - {fn for fn, _, _ in test_files}
        if missing:
            sys.exit(f"No data/tests/*.json for: {', '.join(sorted(missing))}")

    plain_names = bool(getattr(args, "plain_names", False))
    manifest_note = getattr(args, "manifest_note", None)

    cases_flat = flatten_cases(test_files)
    chunks = chunk_functions(cases_flat, args.chunk_size)

    outdir = os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    # Stale chunk files from a previous, differently-sized build would be
    # silently ingestible against the new manifest, so clear them out.
    for stale in glob.glob(os.path.join(outdir, "chunk-*.xlsx")):
        os.remove(stale)

    width = max(2, len(str(len(chunks))))
    manifest_chunks = []
    for idx, (fn_group, chunk_cases) in enumerate(chunks, start=1):
        chunk_id = f"chunk-{idx:0{width}d}"
        wb, sheet_map = build_workbook(chunk_cases, plain_names=plain_names)

        # Every sheet name, including the meta sheet, must survive the Drive
        # import unmangled or the manifest's cell mapping silently breaks.
        for name in list(sheet_map.values()) + [META_SHEET]:
            assert_sheets_safe_name(name)

        path = os.path.join(outdir, f"{chunk_id}.xlsx")
        wb.save(path)

        case_entries = []
        for c in chunk_cases:
            # Must mirror build_workbook()'s own per-case choice exactly --
            # this is what actually got written into the sheet, and ingest
            # trusts it verbatim for both display and provenance.
            stored_formula = c["formula"] if plain_names else to_storage_formula_all(c["formula"])
            case_entries.append({
                "test_id": c["test_id"],
                "function": c["function"],
                "sheet": sheet_map[c["test_id"]],
                "anchor": c["anchor"],
                "check_range": c["check_range"],
                "description": c["description"],
                "formula_display": c["formula"],
                "formula_stored_xlsx": stored_formula,
                "serialization": "plain" if plain_names else "xlfn",
                "expected": c["expected"],
                "expected_note": c["expected_note"],
            })

        manifest_chunks.append({
            "chunk": chunk_id,
            "file": os.path.basename(path),
            "bytes": os.path.getsize(path),
            "sha256": _sha256(path),
            "n_functions": len(fn_group),
            "n_cases": len(chunk_cases),
            "functions": fn_group,
            "sheets": sorted(set(sheet_map.values())),
            "cases": case_entries,
        })

    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "chunk_size": args.chunk_size,
        "n_chunks": len(chunks),
        "n_functions": sum(c["n_functions"] for c in manifest_chunks),
        "n_cases": sum(c["n_cases"] for c in manifest_chunks),
        "subset_only": sorted(requested) if requested else None,
        "plain_names": plain_names,
        "plain_names_functions": sorted({fn for c in manifest_chunks for fn in c["functions"]}) if plain_names else None,
        "manifest_note": manifest_note,
        "canary": {
            "arithmetic_formula": CANARY_ARITH_FORMULA,
            "arithmetic_expected": CANARY_ARITH_EXPECTED,
            "arithmetic_cell_every_sheet": CANARY_ANCHOR,
            "meta_sheet": META_SHEET,
            "meta_volatile_cell": META_VOLATILE_CELL,
            "meta_volatile_formula": "=NOW()",
            "meta_arithmetic_cell": META_ARITH_CELL,
        },
        "chunks": manifest_chunks,
    }
    manifest_path = os.path.join(outdir, MANIFEST_NAME)
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2, default=str)
        f.write("\n")

    mode = "PLAIN NAMES (no _xlfn./_xlfn._xlws. translation)" if plain_names else "xlfn-translated (default)"
    print(f"Built {len(chunks)} chunk(s) from {manifest['n_functions']} function(s), "
          f"{manifest['n_cases']} case(s) -> {outdir}  [{mode}]")
    if manifest_note:
        print(f"Manifest note: {manifest_note}")
    total = 0
    for c in manifest_chunks:
        total += c["bytes"]
        flag = "  <-- OVER SIZE BUDGET" if c["bytes"] > SIZE_WARN_BYTES else ""
        print(f"  {c['file']:>16}  {c['bytes']:>8,} bytes  "
              f"{c['n_functions']:>3} fn  {c['n_cases']:>4} cases{flag}")
    print(f"  {'TOTAL':>16}  {total:>8,} bytes "
          f"(base64 ~{int(total * 4 / 3):,} bytes)")
    print(f"Wrote {manifest_path}")
    return manifest


# --------------------------------------------------------------------------
# ingest
# --------------------------------------------------------------------------

def _chunk_id_from_path(path, manifest):
    """Figure out which manifest chunk an exported workbook corresponds to.

    Filename first (chunk-NN-export.xlsx), then verified against the chunk's
    sheet list so an export uploaded/downloaded out of order can never be
    silently ingested against the wrong cell map."""
    base = os.path.basename(path)
    ids = {c["chunk"] for c in manifest["chunks"]}
    m = re.search(r"(chunk-\d+)", base)
    if m and m.group(1) in ids:
        return m.group(1)
    raise SystemExit(
        f"Cannot tell which chunk {base!r} is. Name exports "
        f"'<chunk-id>-export.xlsx' using an id from the manifest "
        f"({', '.join(sorted(ids))})."
    )


def _boolean_text_artifact(expected, actual):
    """Sheets' .xlsx export can serialize a boolean result as the TEXT
    'TRUE'/'FALSE' rather than a boolean cell. That is an export-format
    artifact, not an engine disagreement, so we never silently coerce it --
    the raw value is recorded as-is and this note flags it for a human."""
    return (isinstance(expected, bool)
            and isinstance(actual, str)
            and actual.upper() in ("TRUE", "FALSE"))


def _decode_volatile_serial(v):
    """Excel serial -> ISO timestamp string, for reading the =NOW() canary.

    Excel for the web exports =NOW() as a bare float under 'General' format
    rather than as a date-formatted cell, so openpyxl hands back the serial
    itself. This renders it human-readable. Note it is the engine's LOCAL
    time, not UTC -- see EXCEL_WEB_READBACK_ARTIFACTS['volatile_now'].
    """
    from datetime import timedelta
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).isoformat()
    except (TypeError, ValueError, OverflowError):
        return None


def ingest_exports(export_paths, manifest, engine_label, engine=ENGINE_ID):
    """Read exported workbooks -> (function_results, canary, trusted, stats).

    `engine` selects an INGEST_ENGINES spec. The chunk workbooks themselves
    are engine-neutral (formula-only, zero cached values), so the same
    manifest and the same cell map serve every engine that recalculates on
    open; only the error vocabulary, the canary prose, the provenance check
    and the results identity differ. See the INGEST ENGINE REGISTRY above.

    Provenance: `manifest["plain_names"]` (set by `build --plain-names`)
    records whether these chunk workbooks had formulas written EXACTLY as
    authored in data/tests (no _xlfn./_xlfn._xlws. storage-form
    translation) rather than the default xlfn-translated serialization.
    Every ingested case gets a "serialization": "plain"/"xlfn" field so a
    later merge (or a human reading results/google-sheets.json) can tell
    which wire form produced a given result -- this matters for Google
    Sheets specifically because its xlsx importer maps plain `_xlfn.NAME`
    but NOT `_xlfn._xlws.FILTER/SORT` or LAMBDA-family functions, so a
    plain-names result for those functions supersedes an earlier
    xlfn-serialized one rather than merely re-confirming it.
    """
    spec = _engine_spec(engine)
    extra_errors = spec["extra_error_strings"]
    plain_names = bool(manifest.get("plain_names", False))
    serialization = "plain" if plain_names else "xlfn"
    manifest_note = manifest.get("manifest_note")
    function_results = {}
    per_chunk = []
    sheet_canary_failures = []
    n_sheet_canaries = 0
    volatile_values = {}
    meta_arith = {}
    app_provenance = {}

    cases_by_chunk = {c["chunk"]: c for c in manifest["chunks"]}

    for path in export_paths:
        chunk_id = _chunk_id_from_path(path, manifest)
        chunk = cases_by_chunk[chunk_id]
        app_provenance[chunk_id] = _read_app_provenance(path)
        wb = openpyxl.load_workbook(path, data_only=True)

        expected_sheets = set(chunk["sheets"])
        present = set(wb.sheetnames)
        missing = expected_sheets - present
        if missing:
            raise SystemExit(
                f"{os.path.basename(path)} claims to be {chunk_id} but is missing "
                f"{len(missing)} of its sheets (e.g. {sorted(missing)[:3]}). "
                f"Wrong file for this chunk, or the manifest is stale -- rebuild."
            )

        if META_SHEET in present:
            meta = wb[META_SHEET]
            volatile_values[chunk_id] = meta[META_VOLATILE_CELL].value
            meta_arith[chunk_id] = meta[META_ARITH_CELL].value
        else:
            volatile_values[chunk_id] = None
            meta_arith[chunk_id] = None

        for case in chunk["cases"]:
            ws = wb[case["sheet"]]
            n_sheet_canaries += 1
            canary_val = ws[CANARY_ANCHOR].value
            canary_ok = canary_val == CANARY_ARITH_EXPECTED
            if not canary_ok:
                sheet_canary_failures.append(
                    {"chunk": chunk_id, "sheet": case["sheet"], "value": canary_val})

            raw_anchor = ws[case["anchor"]].value
            anchor_val = normalize_readback_value(raw_anchor)

            range_flat = None
            if case["check_range"]:
                grid = cell_addrs_in_range(case["check_range"])
                range_flat = [normalize_readback_value(ws[addr].value)
                              for row in grid for addr in row]

            # Google's export writes errors as cached error strings, exactly
            # like LibreOffice's does; the raw string is kept verbatim.
            error = anchor_val if is_error_value(
                anchor_val, extra_errors) else None
            matched, mismatch_detail = compare_expected(
                case["expected"], anchor_val, range_flat)

            notes = []
            if not canary_ok:
                notes.append("UNTRUSTED_RECALC: per-sheet canary failed on this sheet")
            if case.get("expected_note"):
                notes.append(case["expected_note"])
            if mismatch_detail:
                notes.append(f"MISMATCH vs expected: {mismatch_detail}")
            if _boolean_text_artifact(case["expected"], anchor_val):
                notes.append("NOTE: read back as the TEXT 'TRUE'/'FALSE' rather "
                             "than a boolean cell -- likely an .xlsx export "
                             "serialization artifact, verify before reporting it "
                             "as an engine difference")
            if error == "#ERROR!" and "#ERROR!" in extra_errors:
                notes.append("NOTE: #ERROR! is Google Sheets' parse-failure error "
                             "(no Excel equivalent); it means Sheets could not "
                             "parse the formula, which is not the same as #NAME?")

            function_results.setdefault(case["function"], {})[case["test_id"]] = {
                "description": case["description"],
                "formula_display": case["formula_display"],
                "formula_stored_xlsx": case["formula_stored_xlsx"],
                "serialization": case.get("serialization", serialization),
                "value": anchor_val,
                "range_values": range_flat,
                "error": error,
                "expected": case["expected"],
                "matched_expected": matched,
                "canary_ok_this_sheet": canary_ok,
                "notes": "; ".join(notes) if notes else None,
            }

        per_chunk.append({
            "chunk": chunk_id,
            "export_file": os.path.basename(path),
            "n_functions": chunk["n_functions"],
            "n_cases": chunk["n_cases"],
        })

    arith_ok = (not sheet_canary_failures
                and n_sheet_canaries > 0
                and all(v == CANARY_ARITH_EXPECTED for v in meta_arith.values()))

    # Does the package say it was written by the engine we are claiming? A
    # mismatch never blocks the ingest (the canary decides trust) but it is
    # recorded and shouted about, because the failure it guards against is
    # ingesting some other tool's output under this engine's label.
    expect_app = spec["expect_app"]
    app_ok = None
    if expect_app:
        app_ok = bool(app_provenance) and all(
            expect_app in (p.get("application") or "").lower()
            for p in app_provenance.values())

    canary = {
        "arithmetic_formula": CANARY_ARITH_FORMULA,
        "arithmetic_expected": CANARY_ARITH_EXPECTED,
        "arithmetic_actual": (CANARY_ARITH_EXPECTED if arith_ok
                              else (sheet_canary_failures[0]["value"]
                                    if sheet_canary_failures else None)),
        "arithmetic_ok": arith_ok,
        "arithmetic_sheets_checked": n_sheet_canaries,
        "arithmetic_sheet_failures": sheet_canary_failures[:20],
        "meta_arithmetic_per_chunk": meta_arith,
        "volatile_formula": "=NOW()",
        # Raw, verbatim, whatever the export gave us. Sheets renders =NOW() as
        # a datetime; Excel for the web returns a bare float serial. Both are
        # stringified here without interpretation.
        "volatile_per_chunk": {k: str(v) for k, v in volatile_values.items()},
        # Decoded ONLY where the raw value is a numeric serial, so a human can
        # read the timestamp without doing Excel-epoch arithmetic. Presentation
        # of the same number, not a second measurement.
        "volatile_decoded_per_chunk": {
            k: _decode_volatile_serial(v) for k, v in volatile_values.items()
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        } or None,
        "now_differs_across_runs": None,
        "app_provenance_per_chunk": app_provenance,
        "app_provenance_expected": expect_app,
        "app_provenance_ok": app_ok,
        "method": spec["canary_method"],
        "engine_label": engine_label,
    }

    stats = {"chunks": per_chunk,
             "n_functions": len(function_results),
             "n_cases": sum(len(function_cases(v)) for v in function_results.values()),
             "plain_names": plain_names,
             "serialization": serialization,
             "manifest_note": manifest_note,
             "engine": spec["engine_id"],
             "app_provenance_ok": app_ok,
             "app_provenance": app_provenance}
    return function_results, canary, arith_ok, stats


def write_results(out_path, function_results, canary, trusted, engine_label,
                  allow_label_change=False, engine=ENGINE_ID,
                  recalc_method=RECALC_METHOD, serialization=None,
                  manifest_note=None, readback_artifacts=None):
    """Write (or incrementally merge into) a results file in the same schema
    results/libreoffice-*.json uses. See the module docstring."""
    generated_at = datetime.now(timezone.utc).isoformat()
    # Per-function execution date: only the functions in THIS ingest get it.
    # The merge below replaces those blocks wholesale and leaves every other
    # function's executed_at exactly as the run that produced it recorded it.
    stamp_executed_at(function_results, generated_at[:10])
    output = {
        "generated_at": generated_at,
        "engine": engine,
        "engine_version": engine_label,
        "recalc_method": recalc_method,
        "trusted": trusted,
        "canary": canary,
        "function_results": function_results,
    }
    if readback_artifacts:
        # What this engine's export does to values on the way out, measured
        # rather than assumed. Sits next to the results it qualifies so it
        # cannot drift away from them.
        output["readback_artifacts"] = readback_artifacts

    if os.path.exists(out_path):
        with open(out_path) as f:
            prev = json.load(f)
        prev_label = prev.get("engine_version")
        prev_engine = prev.get("engine")
        # An engine-IDENTITY change is never mergeable, with or without
        # --allow-label-change: that flag exists to blend two DATES of the
        # same rolling product, not to pour one engine's results into
        # another's file. Silently allowing it is exactly how Excel-for-the-web
        # values would end up in a file the site reads as desktop Excel.
        if prev_engine and prev_engine != engine:
            sys.exit(
                f"Refusing to merge: {out_path} records engine {prev_engine!r} "
                f"but this ingest is engine {engine!r}. Different engines get "
                f"different results files -- --allow-label-change does NOT "
                f"cover this and never will."
            )
        if prev_label != engine_label and not allow_label_change:
            sys.exit(
                f"Refusing to merge: {out_path} records engine_version "
                f"{prev_label!r} but this ingest is labelled {engine_label!r}. "
                f"These are continuously-updated hosted products, so results "
                f"executed on different dates should not be blended without a "
                f"deliberate decision. Re-run with the same --engine-label, or "
                f"pass --allow-label-change to record both."
            )
        merged = dict(prev)
        if readback_artifacts:
            merged["readback_artifacts"] = readback_artifacts
        merged_fr = dict(prev.get("function_results") or {})
        for fn, cases in function_results.items():
            merged_fr[fn] = cases  # whole function re-executed, so replace wholesale
        merged["function_results"] = merged_fr
        merged["generated_at"] = generated_at
        merged["engine"] = engine
        merged["engine_version"] = engine_label
        merged["recalc_method"] = recalc_method
        merged["canary"] = canary
        # A merge can only downgrade trust, never upgrade it.
        merged["trusted"] = bool(prev.get("trusted", False)) and trusted
        if prev_label != engine_label:
            merged.setdefault("engine_version_history", [])
            if prev_label not in merged["engine_version_history"]:
                merged["engine_version_history"].append(prev_label)
            merged["engine_version_history"].append(engine_label)
        merged.setdefault("subset_runs", []).append({
            "generated_at": generated_at,
            "functions": sorted(function_results.keys()),
            "trusted_this_run": trusted,
            "engine_label": engine_label,
            "superseded_generated_at": prev.get("generated_at"),
            # Provenance: which wire serialization produced this subset's
            # results. A "plain" run for a function supersedes an earlier
            # "xlfn" run for that SAME function (this merge already replaced
            # it above); it is not merely a re-confirmation, because Google
            # Sheets' xlsx importer treats the two serializations
            # differently for FILTER/SORT/LAMBDA-family functions. See
            # README.md "Phase 2: Google Sheets runner".
            "serialization": serialization,
            "manifest_note": manifest_note,
        })
        untouched = len(merged_fr) - len(function_results)
        print(f"Incremental ingest: merged {len(function_results)} function(s) into "
              f"{os.path.basename(out_path)}; {untouched} other function result(s) "
              f"preserved unchanged.")
        output = merged

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
        f.write("\n")
    return output


def summarize(function_results):
    n_name_error = n_other_error = n_ok = 0
    for cases in function_results.values():
        for r in function_cases(cases).values():
            if r["error"] == "#NAME?":
                n_name_error += 1
            elif r["error"]:
                n_other_error += 1
            else:
                n_ok += 1
    return n_ok, n_name_error, n_other_error


def _load_manifest(args):
    path = args.manifest or os.path.join(os.path.abspath(args.chunkdir), MANIFEST_NAME)
    if not os.path.exists(path):
        sys.exit(f"Manifest not found: {path}. Run `run_sheets.py build` first.")
    with open(path) as f:
        return json.load(f)


def cmd_ingest(args):
    engine = getattr(args, "engine", None) or ENGINE_ID
    spec = _engine_spec(engine)
    # argparse defaults are fixed before --engine is known, so the
    # engine-dependent ones resolve here instead.
    if getattr(args, "engine_label", None) is None:
        args.engine_label = spec["label_template"].format(
            date=datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    if getattr(args, "out", None) is None:
        args.out = spec["default_out"]
    if getattr(args, "chunkdir", None) is None:
        args.chunkdir = spec["default_chunkdir"]

    manifest = _load_manifest(args)
    for p in args.export:
        if not os.path.exists(p):
            sys.exit(f"Export not found: {p}")

    function_results, canary, trusted, stats = ingest_exports(
        args.export, manifest, args.engine_label, engine=engine)

    write_results(args.out, function_results, canary, trusted,
                  args.engine_label, args.allow_label_change,
                  engine=spec["engine_id"],
                  recalc_method=spec["recalc_method"],
                  serialization=stats["serialization"],
                  manifest_note=stats["manifest_note"],
                  readback_artifacts=spec["readback_artifacts"])

    n_ok, n_name, n_other = summarize(function_results)
    print(f"Engine: {spec['display']}  [{spec['engine_id']}]")
    print(f"Label:  {args.engine_label}")
    print(f"Ingested {len(stats['chunks'])} chunk(s): "
          f"{stats['n_functions']} function(s), {stats['n_cases']} case(s). "
          f"[serialization={stats['serialization']}]")
    if stats["manifest_note"]:
        print(f"Manifest note: {stats['manifest_note']}")
    print(f"Deterministic canary OK on all {canary['arithmetic_sheets_checked']} "
          f"sheet(s): {canary['arithmetic_ok']}")
    if canary["arithmetic_sheet_failures"]:
        print(f"  CANARY FAILURES: {canary['arithmetic_sheet_failures']}")
    print(f"Volatile =NOW() per chunk: {canary['volatile_per_chunk']}")
    if canary.get("volatile_decoded_per_chunk"):
        print(f"  decoded (engine-local time): "
              f"{canary['volatile_decoded_per_chunk']}")
    if spec["expect_app"]:
        apps = {k: v.get("application") for k, v in
                (canary.get("app_provenance_per_chunk") or {}).items()}
        print(f"Package provenance (docProps/app.xml): {apps}")
        if stats["app_provenance_ok"] is False:
            print(f"  *** WARNING: expected an application naming "
                  f"{spec['expect_app']!r}. These bytes were not written by "
                  f"{spec['display']}. The values are being recorded, but do "
                  f"NOT publish them as {spec['display']} results without "
                  f"explaining this. ***")
    print(f"Trusted recalculation: {trusted}")
    print(f"Cases with a value (no error): {n_ok}")
    print(f"Cases returning #NAME? (unsupported function): {n_name}")
    print(f"Cases returning some other error: {n_other}")
    print(f"Wrote {args.out}")
    return function_results


# --------------------------------------------------------------------------
# selftest: prove the build -> export -> ingest plumbing end-to-end
# --------------------------------------------------------------------------

def cmd_selftest(args):
    """Run the whole pipeline with LibreOffice standing in for Google Drive.

    This proves the CELL MAPPING and the ingest plumbing, nothing about
    Google Sheets: the values recovered are LibreOffice's. It therefore
    writes to a scratch results file and never to results/google-sheets.json.
    """
    print("=== selftest: build ===")
    manifest = cmd_build(args)

    chunkdir = os.path.abspath(args.outdir)
    soffice = os.environ.get("SOFFICE_BIN", "soffice")

    with tempfile.TemporaryDirectory() as tmp:
        export_dir = os.path.join(tmp, "exports")
        os.makedirs(export_dir)
        exports = []
        print("\n=== selftest: simulated export (soffice --convert-to xlsx) ===")
        for chunk in manifest["chunks"]:
            src = os.path.join(chunkdir, chunk["file"])
            outd = os.path.join(tmp, chunk["chunk"])
            os.makedirs(outd, exist_ok=True)
            proc = subprocess.run(
                [soffice, "--headless", "--convert-to", "xlsx",
                 "--outdir", outd, src],
                capture_output=True, text=True, timeout=600)
            produced = os.path.join(outd, chunk["file"])
            if proc.returncode != 0 or not os.path.exists(produced):
                sys.exit(f"soffice conversion failed for {chunk['file']}: "
                         f"{proc.stderr.strip()}")
            dest = os.path.join(export_dir, f"{chunk['chunk']}-export.xlsx")
            shutil.copyfile(produced, dest)
            exports.append(dest)
            print(f"  {chunk['file']} -> {os.path.basename(dest)} "
                  f"({os.path.getsize(dest):,} bytes)")

        print("\n=== selftest: ingest ===")
        function_results, canary, trusted, stats = ingest_exports(
            exports, manifest, args.engine_label)

        out = os.path.abspath(args.out)
        # Hard stop: the site generator globs results/*.json and assigns an
        # engine from the "engine" string, so ANY selftest output inside
        # results/ risks LibreOffice values being published as Google Sheets.
        if os.path.commonpath([out, os.path.abspath(RESULTS_DIR)]) == \
                os.path.abspath(RESULTS_DIR):
            sys.exit(f"selftest refuses to write inside results/ ({out}). These "
                     f"are LibreOffice values produced through the Sheets "
                     f"plumbing, not Google Sheets values, and the site "
                     f"generator picks up every results/*.json by engine name.")
        if os.path.exists(out):
            os.remove(out)  # scratch file: always a fresh write, never a merge
        write_results(out, function_results, canary, trusted, args.engine_label,
                      engine=SELFTEST_ENGINE_ID,
                      recalc_method=SELFTEST_RECALC_METHOD,
                      serialization=stats["serialization"],
                      manifest_note=stats["manifest_note"])

    n_ok, n_name, n_other = summarize(function_results)
    print(f"\nChunks ingested       : {len(stats['chunks'])}")
    print(f"Functions / cases     : {stats['n_functions']} / {stats['n_cases']}")
    print(f"Per-sheet canaries    : {canary['arithmetic_sheets_checked']} checked, "
          f"ok={canary['arithmetic_ok']}, failures={canary['arithmetic_sheet_failures']}")
    print(f"Volatile =NOW()       : {canary['volatile_per_chunk']}")
    print(f"trusted               : {trusted}")
    print(f"value / #NAME? / other: {n_ok} / {n_name} / {n_other}")

    print("\n=== 3 sample round-tripped values ===")
    shown = 0
    for fn in sorted(function_results):
        for tid in sorted(function_cases(function_results[fn])):
            r = function_results[fn][tid]
            print(f"  {tid}: {r['formula_display']}  ->  value={r['value']!r}  "
                  f"expected={r['expected']!r}  matched={r['matched_expected']}")
            shown += 1
            if shown >= 3:
                break
        if shown >= 3:
            break
    print(f"\nWrote scratch results -> {args.out} "
          f"(LibreOffice values; plumbing proof only)")


# ==========================================================================
# RECIPES: the second corpus (data/recipes/*.json), executed in Google Sheets
# ==========================================================================
# The how-to recipes have been executed in headless LibreOffice since day one
# (scripts/verify_recipes.py -> results/recipes-verified.json). These three
# subcommands give them the same Drive-import treatment the FUNCTION corpus
# gets above, writing results/recipes-verified-sheets.json.
#
# Everything about WHICH checks exist, how a variant check inherits its
# setup, how a read-back value is normalized and how it is judged against
# `expected` comes from harness/recipe_corpus.py -- the same module
# verify_recipes.py now imports. That is deliberate and load-bearing: the
# site prints "LibreOffice returned X, Google Sheets returned Y" side by
# side, so any difference in enumeration or comparison between the two paths
# would manufacture a fake engine divergence.
#
# SERIALIZATION: unlike the function corpus, recipe chunks are written with
# PLAIN function names by default (`--plain-names`, on unless you pass
# `--xlfn-names`). Google Sheets' xlsx importer maps bare modern names
# (FILTER, SORT, LAMBDA, TEXTJOIN, ...) but does NOT map the
# `_xlfn._xlws.FILTER` storage form, and recipes are authored the way a user
# would type them into Sheets. The trade-off is recorded per chunk in the
# manifest and per run in the results file, because it is the one thing that
# makes a Sheets recipe result not directly byte-comparable to the
# LibreOffice reference run for those functions.

DEFAULT_RECIPE_CHUNK_DIR = os.path.join(REPO_ROOT, "harness", "recipe_chunks")
DEFAULT_RECIPE_EXPORT_DIR = os.path.join(REPO_ROOT, "harness", "recipe_exports")
DEFAULT_RECIPE_RESULTS_PATH = os.path.join(RESULTS_DIR, "recipes-verified-sheets.json")
# Scratch output for `selftest-recipes`. Deliberately OUTSIDE results/, for
# exactly the reason cmd_selftest's is: these are LibreOffice values.
RECIPE_SELFTEST_DIR = os.path.join(REPO_ROOT, "harness", "recipe_selftest")
RECIPE_SELFTEST_RESULTS_PATH = os.path.join(RECIPE_SELFTEST_DIR, "plumbing-check.json")
RECIPE_SELFTEST_CHUNK_DIR = os.path.join(RECIPE_SELFTEST_DIR, "chunks")
LO_RECIPE_RESULTS_PATH = os.path.join(RESULTS_DIR, "recipes-verified.json")

RECIPE_RECALC_METHOD = "Drive import + xlsx export readback (how-to recipe corpus)"
RECIPE_SELFTEST_ENGINE_ID = "SELFTEST_libreoffice_via_sheets_pipeline"

# Six slugs that exercise a scalar result, a date-as-text result, an array
# result written as a real array formula over a check_range, a 1x1
# check_range, and the only single-sheet recipe that carries variants.
RECIPE_SELFTEST_SLUGS = [
    "abbreviate-large-numbers",       # plain scalar text result
    "add-days-to-a-date",             # TEXT()-wrapped date
    "case-sensitive-lookup",          # 1x1 check_range (array formula)
    "list-the-top-n-values",          # 3-cell spill via check_range
    "transpose-rows-to-columns",      # 3-cell horizontal spill
    "sum-with-multiple-criteria",     # 15 variant checks across 5 sections
]


def recipe_sheet_name(recipe_index, check_key, slug, used):
    """Worksheet name for one check: '<id>_<slug>', truncated to 31 chars.

    The id prefix (r007, r007v2c1) is what makes the name unique and is
    derived from the recipe's position in the FULL corpus, so a `--only`
    build lands its checks on exactly the sheets a full build would. The
    slug tail is purely so a human opening the workbook in Drive can tell
    what they are looking at; ingest reads the name from the manifest and
    never re-derives it.

    31 chars, not Sheets' 100: the transport is .xlsx in both directions and
    a >31-char sheet name is out-of-spec OOXML whatever opens the file (the
    same rule corpus.py applies to the function corpus).
    """
    prefix = f"r{recipe_index:03d}" if check_key == "main" else f"r{recipe_index:03d}{check_key}"
    room = SHEET_NAME_MAX - len(prefix) - 1
    name = f"{prefix}_{slug[:room]}" if room > 0 else prefix
    return assert_sheets_safe_name(sanitize_sheet_name(name, used))


def collect_recipe_checks(only=None, engine=GOOGLE_SHEETS):
    """Return (buildable, skipped, missing).

    buildable: [{slug, title, index, checks:[check dicts from iter_checks]}]
    skipped:   multi-sheet recipes, which v1 does not build (see below)
    missing:   requested slugs with no data/recipes/*.json

    ENGINE SCOPING
    --------------
    `engine` is the engine the built workbook will be executed by, and the
    check list is filtered to it: a check carrying `"engines":
    ["google_sheets"]` (a Sheets-only alternative formula) is built ONLY for
    Sheets, and a check carrying `["libreoffice"]` never is. Checks with no
    `engines` key are built for every engine, so this is a no-op for every
    recipe authored before the field existed. `selftest-recipes` passes
    "libreoffice", because there the engine really is LibreOffice standing in
    for Drive and the reference values it compares against are LibreOffice's.

    Sheet naming and manifest keys come from iter_checks' STABLE KEYS, which
    are positional over the UNFILTERED JSON, so a filtered build lands each
    check on exactly the sheet a full build would.

    WHY MULTI-SHEET RECIPES ARE SKIPPED (v1 decision, deliberate)
    -------------------------------------------------------------
    Six recipes need extra worksheets to exist (`setup_sheets`), and the
    natural fix -- give every check its own copy of those tabs, prefixed to
    avoid collisions, and rewrite the formula's sheet references to match --
    is NOT correct for this corpus. Their checks are precisely the ones a
    rename breaks:

      * `=INDIRECT(A1&"!B2")` builds the sheet reference out of a CELL VALUE,
        so renaming the tab silently changes the answer unless the setup data
        is rewritten too, and one of those checks exists specifically to
        assert a #REF! for a tab that does NOT exist.
      * `=SUM(Q1:Q3!A1)` is a 3-D reference over a SPAN of consecutive tabs;
        it depends on sheet ORDER as well as names.
      * `=$'Q1 Data'.B2` and `=Data.B2` are there to assert #NAME? -- the
        LibreOffice-syntax forms that Excel/Sheets reject. A rewriter that
        "fixed" those references would destroy the very thing under test.
      * Sheet names collide WITHIN a single recipe: one variant defines
        Q1/Q2/Q3 as numbers and a check inside it redefines Q1/Q2/Q3 with
        different contents, so even one-recipe-per-workbook does not
        de-collide them.

    Splitting them out is therefore the simpler CORRECT option, and the
    manifest lists them by name with their extra tabs so the omission is
    visible rather than silent. They are no longer unexecutable, though: the
    v2 hinted at here exists as `build-recipes-multisheet`, which runs one
    workbook per CHECK (not per recipe) and so covers them with no rewriting
    at all, at the cost of ~34 more Drive round-trips.
    """
    all_recipes = load_recipe_files()
    index_of = {slug: i + 1 for i, (slug, _, _) in enumerate(all_recipes)}
    wanted = set(only) if only else None
    missing = sorted(wanted - set(index_of)) if wanted else []

    buildable, skipped = [], []
    for slug, _path, recipe in all_recipes:
        if wanted and slug not in wanted:
            continue
        checks = list(iter_checks(recipe, engine=engine))
        if not checks:
            # Every one of this recipe's checks is scoped to another engine.
            continue
        entry = {
            "slug": slug,
            "title": recipe.get("title", ""),
            "index": index_of[slug],
            "checks": checks,
            "n_checks": len(checks),
        }
        if uses_setup_sheets(recipe):
            entry["setup_sheet_names"] = setup_sheet_names(recipe)
            skipped.append(entry)
        else:
            buildable.append(entry)
    return buildable, skipped, missing


def chunk_recipes(recipes, chunk_size):
    """Partition RECIPES (never one recipe's checks) into chunks.

    Same rule, same reason, as chunk_functions(): a recipe is merged into the
    results file wholesale and its `verified` flag is the AND over all of its
    checks, so a recipe split across two chunks could be written out as
    "verified" from a half-ingested export."""
    return [recipes[i:i + chunk_size] for i in range(0, len(recipes), chunk_size)]


def build_recipe_workbook(recipes, plain_names=True):
    """One worksheet per check. Returns (workbook, sheet_map) where
    sheet_map[(slug, check_key)] -> sheet name."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()
    sheet_map = {}

    for rec in recipes:
        for chk in rec["checks"]:
            name = recipe_sheet_name(rec["index"], chk["key"], rec["slug"], used_names)
            sheet_map[(rec["slug"], chk["key"])] = name
            ws = wb.create_sheet(name)

            for addr, val in (chk["setup_cells"] or {}).items():
                ws[addr] = val

            formula = chk["formula"] if plain_names else to_storage_formula_all(chk["formula"])
            if chk["check_range"]:
                # Same reasoning as corpus.build_workbook(): a result that is
                # meant to spill has to be written as a real array formula
                # over its full range, or an engine returns a scalar #VALUE!
                # for a function it fully supports.
                ws[chk["anchor"]] = ArrayFormula(chk["check_range"], formula)
            else:
                ws[chk["anchor"]] = formula

            ws[CANARY_ANCHOR] = CANARY_ARITH_FORMULA

    meta = wb.create_sheet(META_SHEET, 0)
    meta[META_VOLATILE_CELL] = "=NOW()"
    meta[META_ARITH_CELL] = CANARY_ARITH_FORMULA
    assert_sheets_safe_name(META_SHEET)
    return wb, sheet_map


def cmd_build_recipes(args):
    plain_names = bool(getattr(args, "plain_names", True))
    engine = getattr(args, "engine", GOOGLE_SHEETS)
    buildable, skipped, missing = collect_recipe_checks(args.only, engine=engine)
    if missing:
        sys.exit(f"No data/recipes/*.json for: {', '.join(missing)}")
    if not buildable:
        sys.exit("No buildable recipes matched (all matches were multi-sheet "
                 "recipes, which this command skips -- see --help).")

    outdir = os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    # Stale chunks from a previous, differently-sized build would still be
    # ingestible against the new manifest, so clear them first.
    for stale in glob.glob(os.path.join(outdir, "chunk-*.xlsx")):
        os.remove(stale)

    chunks = chunk_recipes(buildable, args.chunk_size)
    width = max(2, len(str(len(chunks))))
    manifest_chunks = []
    for idx, group in enumerate(chunks, start=1):
        chunk_id = f"chunk-{idx:0{width}d}"
        wb, sheet_map = build_recipe_workbook(group, plain_names=plain_names)
        path = os.path.join(outdir, f"{chunk_id}.xlsx")
        wb.save(path)

        recipe_entries = []
        for rec in group:
            checks = []
            for chk in rec["checks"]:
                stored = (chk["formula"] if plain_names
                          else to_storage_formula_all(chk["formula"]))
                checks.append({
                    "key": chk["key"],
                    "kind": chk["kind"],
                    "variant_index": chk["variant_index"],
                    "check_index": chk["check_index"],
                    "heading": chk["heading"],
                    "label": chk["label"],
                    "sheet": sheet_map[(rec["slug"], chk["key"])],
                    "anchor": chk["anchor"],
                    "check_range": chk["check_range"],
                    "formula_display": chk["formula"],
                    "formula_stored_xlsx": stored,
                    # Whether this exact check's bytes differ from the ones the
                    # LibreOffice reference run executed. Only true where the
                    # formula names a function xlfn_map rewrites.
                    "differs_from_lo_serialization":
                        stored != to_storage_formula_all(chk["formula"]),
                    "serialization": "plain" if plain_names else "xlfn",
                    "expected": chk["expected"],
                    # None = all engines. Recorded so an ingest (and a human
                    # reading the manifest) can see that a check is an
                    # engine-scoped alternative rather than a shared formula.
                    "engines": chk["engines"],
                })
            recipe_entries.append({
                "slug": rec["slug"],
                "title": rec["title"],
                "index": rec["index"],
                "n_checks": len(checks),
                "n_variants": len({c["variant_index"] for c in checks
                                   if c["variant_index"] is not None}),
                "checks": checks,
            })

        manifest_chunks.append({
            "chunk": chunk_id,
            "file": os.path.basename(path),
            "bytes": os.path.getsize(path),
            "sha256": _sha256(path),
            "n_recipes": len(group),
            "n_checks": sum(r["n_checks"] for r in recipe_entries),
            "slugs": [r["slug"] for r in recipe_entries],
            "sheets": sorted(sheet_map.values()),
            "recipes": recipe_entries,
        })

    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "corpus": "recipes",
        "corpus_source": "data/recipes/*.json",
        # The engine this chunk set was built FOR. Checks carrying an
        # "engines" list that excludes it were left out of the workbooks.
        "engine_scope": engine,
        "chunk_size": args.chunk_size,
        "n_chunks": len(chunks),
        "n_recipes": sum(c["n_recipes"] for c in manifest_chunks),
        "n_checks": sum(c["n_checks"] for c in manifest_chunks),
        "subset_only": sorted(set(args.only)) if args.only else None,
        "plain_names": plain_names,
        "serialization": "plain" if plain_names else "xlfn",
        "manifest_note": getattr(args, "manifest_note", None),
        "skipped_multi_sheet": [
            {"slug": r["slug"], "n_checks": r["n_checks"],
             "setup_sheets": r["setup_sheet_names"]}
            for r in skipped
        ],
        "skipped_multi_sheet_reason": (
            "v1 does not build recipes whose checks need extra worksheets. Their "
            "formulas resolve sheet names from cell VALUES (INDIRECT), span "
            "consecutive tabs (3-D refs), or deliberately assert #NAME?/#REF! for "
            "a wrong-syntax or missing tab, and a single recipe can define the "
            "same tab name twice with different contents -- so neither prefix-"
            "renaming nor one-workbook-per-recipe is correct. They are built by "
            "`build-recipes-multisheet` instead, which gives each CHECK its own "
            "workbook with exactly the tabs it declares, in the recipe's own "
            "order, renaming nothing."
        ),
        "skipped_multi_sheet_built_by": "build-recipes-multisheet",
        "canary": {
            "arithmetic_formula": CANARY_ARITH_FORMULA,
            "arithmetic_expected": CANARY_ARITH_EXPECTED,
            "arithmetic_cell_every_sheet": CANARY_ANCHOR,
            "meta_sheet": META_SHEET,
            "meta_volatile_cell": META_VOLATILE_CELL,
            "meta_volatile_formula": "=NOW()",
            "meta_arithmetic_cell": META_ARITH_CELL,
        },
        "chunks": manifest_chunks,
    }
    manifest_path = os.path.join(outdir, MANIFEST_NAME)
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2, default=str)
        f.write("\n")

    mode = ("PLAIN NAMES (formulas exactly as authored -- Sheets maps unprefixed "
            "names)" if plain_names else "xlfn-translated")
    print(f"Built {len(chunks)} chunk(s) from {manifest['n_recipes']} recipe(s), "
          f"{manifest['n_checks']} check(s) -> {outdir}  [{mode}]")
    n_scoped = sum(1 for c in manifest_chunks for r in c["recipes"]
                   for chk in r["checks"] if chk.get("engines"))
    print(f"Engine scope: {engine} -- {n_scoped} of {manifest['n_checks']} check(s) "
          f"are engine-scoped, the rest run in every engine")
    if manifest["manifest_note"]:
        print(f"Manifest note: {manifest['manifest_note']}")
    total = 0
    print(f"  {'file':>16}  {'bytes':>9}  {'recipes':>7}  {'checks':>6}")
    for c in manifest_chunks:
        total += c["bytes"]
        flag = "  <-- OVER SIZE BUDGET" if c["bytes"] > SIZE_WARN_BYTES else ""
        print(f"  {c['file']:>16}  {c['bytes']:>9,}  {c['n_recipes']:>7}  "
              f"{c['n_checks']:>6}{flag}")
    print(f"  {'TOTAL':>16}  {total:>9,} "
          f"(base64 ~{int(total * 4 / 3):,} bytes)")
    if skipped:
        print(f"\nSkipped {len(skipped)} multi-sheet recipe(s) "
              f"({sum(r['n_checks'] for r in skipped)} check(s)) -- these need extra "
              f"worksheets; run build-recipes-multisheet to execute them in Sheets "
              f"(see the manifest's 'skipped_multi_sheet_reason'):")
        for r in skipped:
            print(f"  {r['slug']:45} {r['n_checks']:>3} check(s)  "
                  f"tabs: {', '.join(r['setup_sheet_names'])}")
    print(f"\nWrote {manifest_path}")
    return manifest


# --------------------------------------------------------------------------
# ingest-recipes
# --------------------------------------------------------------------------

def _read_check_value(ws, entry):
    """Read one check's result back exactly the way verify_recipes.run_case
    returns it, so the two engines' `actual` values are directly comparable:

      * check_range  -> row-major list of norm()'d values with the Nones
                        dropped (an unfilled spill cell reads back as None)
      * otherwise    -> norm() of the single anchor cell

    norm() is recipe_corpus's, i.e. verify_recipes' own."""
    if entry["check_range"]:
        grid = cell_addrs_in_range(entry["check_range"])
        vals = [norm(ws[addr].value) for row in grid for addr in row]
        return [v for v in vals if v is not None]
    return norm(ws[entry["anchor"]].value)


def _readback_artifact_notes(expected, raw, actual):
    """Export-format artifacts worth flagging to a human, never silently
    coerced away. Same policy as the function-corpus ingest."""
    import datetime as _dt
    notes = []
    if isinstance(expected, bool) and isinstance(actual, str) \
            and actual.upper() in ("TRUE", "FALSE"):
        notes.append("NOTE: read back as the TEXT 'TRUE'/'FALSE' rather than a "
                     "boolean cell -- likely an .xlsx export serialization "
                     "artifact, verify before reporting it as an engine difference")
    if isinstance(raw, (_dt.datetime, _dt.date, _dt.time)):
        notes.append(f"NOTE: read back as a date/time-formatted cell ({raw!r}); "
                     f"the underlying value is the serial "
                     f"{normalize_readback_value(raw)!r}")
    if actual == "#ERROR!":
        notes.append("NOTE: #ERROR! is Google Sheets' parse-failure error (no Excel "
                     "equivalent); it means Sheets could not parse the formula, "
                     "which is not the same as #NAME?")
    return notes


def ingest_recipe_exports(export_paths, manifest, engine_label):
    """Read exported workbooks -> (recipe_results, canary, trusted, stats)."""
    plain_names = bool(manifest.get("plain_names", True))
    serialization = manifest.get("serialization", "plain" if plain_names else "xlfn")
    manifest_note = manifest.get("manifest_note")

    recipes_out = {}
    per_chunk = []
    sheet_canary_failures = []
    n_sheet_canaries = 0
    volatile_values = {}
    meta_arith = {}
    n_checks = 0

    chunks_by_id = {c["chunk"]: c for c in manifest["chunks"]}

    for path in export_paths:
        chunk_id = _chunk_id_from_path(path, manifest)
        chunk = chunks_by_id[chunk_id]
        wb = openpyxl.load_workbook(path, data_only=True)

        missing = set(chunk["sheets"]) - set(wb.sheetnames)
        if missing:
            raise SystemExit(
                f"{os.path.basename(path)} claims to be {chunk_id} but is missing "
                f"{len(missing)} of its sheets (e.g. {sorted(missing)[:3]}). "
                f"Wrong file for this chunk, or the manifest is stale -- rebuild."
            )

        if META_SHEET in wb.sheetnames:
            meta = wb[META_SHEET]
            volatile_values[chunk_id] = meta[META_VOLATILE_CELL].value
            meta_arith[chunk_id] = meta[META_ARITH_CELL].value
        else:
            volatile_values[chunk_id] = None
            meta_arith[chunk_id] = None

        for rec in chunk["recipes"]:
            main = None
            variants = {}
            extra = []
            recipe_ok = True
            for entry in rec["checks"]:
                ws = wb[entry["sheet"]]
                n_sheet_canaries += 1
                n_checks += 1
                canary_val = ws[CANARY_ANCHOR].value
                canary_ok = canary_val == CANARY_ARITH_EXPECTED
                if not canary_ok:
                    sheet_canary_failures.append(
                        {"chunk": chunk_id, "sheet": entry["sheet"],
                         "slug": rec["slug"], "value": canary_val})

                raw = (None if entry["check_range"]
                       else ws[entry["anchor"]].value)
                # The comparison is inside the try for the same reason it is in
                # recipe_corpus.run_check(): a list `expected` against a
                # non-iterable `actual` raises, and that is recorded as an ERR
                # result rather than crashing the ingest.
                try:
                    actual = _read_check_value(ws, entry)
                    ok = compare_check(entry["expected"], actual)
                except Exception as e:  # noqa: BLE001
                    actual = f"ERR {e}"
                    ok = False
                ok = bool(ok)

                notes = _readback_artifact_notes(entry["expected"], raw, actual)
                if not canary_ok:
                    notes.insert(0, "UNTRUSTED_RECALC: per-sheet canary failed "
                                    "on this sheet")
                if entry.get("differs_from_lo_serialization"):
                    notes.append(
                        "NOTE: written with the PLAIN function name; the "
                        "LibreOffice reference run executed the _xlfn. storage "
                        "form of this formula, so the two runs are not "
                        "byte-identical inputs")
                if not ok:
                    notes.append(f"MISMATCH vs expected: expected "
                                 f"{entry['expected']!r}, got {actual!r}")
                if not ok:
                    recipe_ok = False

                payload = {
                    # Stable key: what the site merges this result by, so a
                    # later-appended engine-scoped check cannot shift this
                    # value onto a neighbouring formula.
                    "key": entry["key"],
                    "label": entry["label"],
                    "formula": entry["formula_display"],
                    "formula_stored_xlsx": entry["formula_stored_xlsx"],
                    "serialization": entry.get("serialization", serialization),
                    "expected": entry["expected"],
                    "actual": actual,
                    "verified": ok,
                    "sheet": entry["sheet"],
                    "canary_ok_this_sheet": canary_ok,
                    "notes": "; ".join(notes) if notes else None,
                }
                if entry.get("engines"):
                    payload["engines"] = entry["engines"]
                if entry["kind"] == "main":
                    main = payload
                elif entry["kind"] == "extra":
                    extra.append(payload)
                else:
                    variants.setdefault(
                        entry["variant_index"],
                        {"heading": entry["heading"], "checks": []},
                    )["checks"].append(payload)

            if main is None:      # cannot happen for a manifest this tool wrote
                raise SystemExit(f"{rec['slug']}: manifest has no 'main' check")

            record = {
                "verified": recipe_ok,
                "engine": "Google Sheets",
                "engine_label": engine_label,
                "serialization": serialization,
                "key": main["key"],
                "formula": main["formula"],
                "formula_stored_xlsx": main["formula_stored_xlsx"],
                "expected": main["expected"],
                "actual": main["actual"],
                "main_verified": main["verified"],
                "sheet": main["sheet"],
                "canary_ok_this_sheet": main["canary_ok_this_sheet"],
                "notes": main["notes"],
            }
            if variants:
                record["variants"] = [variants[i] for i in sorted(variants)]
            if extra:
                record["extra_checks"] = extra
            recipes_out[rec["slug"]] = record

        per_chunk.append({
            "chunk": chunk_id,
            "export_file": os.path.basename(path),
            "n_recipes": chunk["n_recipes"],
            "n_checks": chunk["n_checks"],
        })

    arith_ok = (not sheet_canary_failures
                and n_sheet_canaries > 0
                and all(v == CANARY_ARITH_EXPECTED for v in meta_arith.values()))

    canary = {
        "arithmetic_formula": CANARY_ARITH_FORMULA,
        "arithmetic_expected": CANARY_ARITH_EXPECTED,
        "arithmetic_actual": (CANARY_ARITH_EXPECTED if arith_ok
                              else (sheet_canary_failures[0]["value"]
                                    if sheet_canary_failures else None)),
        "arithmetic_ok": arith_ok,
        "arithmetic_sheets_checked": n_sheet_canaries,
        "arithmetic_sheet_failures": sheet_canary_failures[:20],
        "meta_arithmetic_per_chunk": meta_arith,
        "volatile_formula": "=NOW()",
        "volatile_per_chunk": {k: str(v) for k, v in volatile_values.items()},
        "now_differs_across_runs": None,
        "method": (
            "openpyxl writes formulas with NO cached <v> value, so the uploaded "
            "chunk workbooks contain zero cached results. Google Drive's "
            "auto-conversion to a Google Sheet recalculates every formula with "
            "Google's engine; exporting that Sheet back to .xlsx carries those "
            "computed values out. The deterministic canary =1111+2222 in "
            f"{CANARY_ANCHOR} of EVERY check's sheet reading back exactly "
            f"{CANARY_ARITH_EXPECTED} therefore proves genuine computation -- "
            "without recalculation the cell would read back blank (None), since "
            "nothing was ever cached. The volatile =NOW() canary is recorded per "
            "chunk as corroboration, but a single Drive import yields one "
            "timestamp, so no cross-run volatile comparison is possible: "
            "now_differs_across_runs is null by design."
        ),
        "engine_label": engine_label,
    }

    stats = {"chunks": per_chunk,
             "n_recipes": len(recipes_out),
             "n_checks": n_checks,
             "serialization": serialization,
             "manifest_note": manifest_note}
    return recipes_out, canary, arith_ok, stats


def write_recipe_results(out_path, recipes, canary, trusted, engine_label,
                         allow_label_change=False, engine=ENGINE_ID,
                         recalc_method=RECIPE_RECALC_METHOD,
                         serialization=None, manifest_note=None):
    """Write (or incrementally merge into) results/recipes-verified-sheets.json.

    Merge semantics are the function corpus's, one level down: a recipe that
    was re-executed is replaced WHOLESALE (its variants come as one block),
    every other recipe is preserved byte-for-byte, and `trusted` is the AND of
    the previous file's flag and this run's, so a merge can only ever downgrade
    trust. Merging under a different engine label is refused without
    --allow-label-change, because Sheets is a continuously-updated hosted
    product and two dates' results should not be blended by accident."""
    generated_at = datetime.now(timezone.utc).isoformat()
    output = {
        "generated_at": generated_at,
        "corpus": "recipes",
        "engine": engine,
        "engine_label": engine_label,
        "recalc_method": recalc_method,
        "trusted": trusted,
        "canary": canary,
        "recipes": recipes,
    }

    if os.path.exists(out_path):
        with open(out_path) as f:
            prev = json.load(f)
        prev_label = prev.get("engine_label")
        if prev_label != engine_label and not allow_label_change:
            sys.exit(
                f"Refusing to merge: {out_path} records engine_label "
                f"{prev_label!r} but this ingest is labelled {engine_label!r}. "
                f"Re-run with the same --engine-label, or pass "
                f"--allow-label-change to record both."
            )
        merged = dict(prev)
        merged_recipes = dict(prev.get("recipes") or {})
        merged_recipes.update(recipes)
        merged["recipes"] = merged_recipes
        merged["generated_at"] = generated_at
        merged["corpus"] = "recipes"
        merged["engine"] = engine
        merged["engine_label"] = engine_label
        merged["recalc_method"] = recalc_method
        merged["canary"] = canary
        merged["trusted"] = bool(prev.get("trusted", False)) and trusted
        if prev_label != engine_label:
            merged.setdefault("engine_label_history", [])
            if prev_label not in merged["engine_label_history"]:
                merged["engine_label_history"].append(prev_label)
            merged["engine_label_history"].append(engine_label)
        merged.setdefault("subset_runs", []).append({
            "generated_at": generated_at,
            "recipes": sorted(recipes),
            "trusted_this_run": trusted,
            "engine_label": engine_label,
            "serialization": serialization,
            "manifest_note": manifest_note,
            "superseded_generated_at": prev.get("generated_at"),
        })
        untouched = len(merged_recipes) - len(recipes)
        print(f"Incremental ingest: merged {len(recipes)} recipe(s) into "
              f"{os.path.basename(out_path)}; {untouched} other recipe result(s) "
              f"preserved unchanged.")
        output = merged

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
        f.write("\n")
    return output


def _load_recipe_manifest(args):
    path = args.manifest or os.path.join(os.path.abspath(args.chunkdir), MANIFEST_NAME)
    if not os.path.exists(path):
        sys.exit(f"Manifest not found: {path}. Run `run_sheets.py build-recipes` first.")
    with open(path) as f:
        manifest = json.load(f)
    if manifest.get("corpus") != "recipes":
        sys.exit(f"{path} is not a recipe manifest (corpus="
                 f"{manifest.get('corpus', 'functions')!r}). Point --chunkdir at "
                 f"the directory `build-recipes` wrote.")
    return manifest


def _resolve_recipe_exports(args):
    """The exported workbooks to ingest, from --export files and/or --export-dir.

    A directory is accepted because the MULTI-SHEET build emits one workbook
    per check: a human who has just downloaded 34 converted Sheets wants to
    point at the folder, not type 34 paths. Every file still has to identify
    itself through the manifest (`ms-NNN` / `chunk-NN` in its name, then its
    sheet list), so widening the input never widens what can be mis-ingested.
    """
    paths = list(getattr(args, "export", None) or [])
    for d in (getattr(args, "export_dir", None) or []):
        if not os.path.isdir(d):
            sys.exit(f"--export-dir is not a directory: {d}")
        found = sorted(p for p in glob.glob(os.path.join(d, "*.xlsx"))
                       if not os.path.basename(p).startswith("~$"))
        if not found:
            sys.exit(f"No .xlsx files in {d}")
        paths.extend(found)
    if not paths:
        sys.exit("Nothing to ingest: pass --export FILE ... and/or --export-dir DIR")
    seen, uniq = set(), []
    for p in paths:
        ap = os.path.abspath(p)
        if ap in seen:
            continue
        seen.add(ap)
        uniq.append(p)
    for p in uniq:
        if not os.path.exists(p):
            sys.exit(f"Export not found: {p}")
    return uniq


def cmd_ingest_recipes(args):
    """Ingest either chunk layout. Which one is not a flag the caller has to
    remember: the manifest records `mode`, and `build-recipes-multisheet`
    stamps it, so pointing --chunkdir at a multi-sheet build selects the
    one-workbook-per-check reader automatically."""
    manifest = _load_recipe_manifest(args)
    exports = _resolve_recipe_exports(args)
    multisheet = manifest.get("mode") == MULTISHEET_MODE

    if multisheet:
        recipes, canary, trusted, stats = ingest_multisheet_exports(
            exports, manifest, args.engine_label)
        recalc_method = RECIPE_MS_RECALC_METHOD
    else:
        recipes, canary, trusted, stats = ingest_recipe_exports(
            exports, manifest, args.engine_label)
        recalc_method = RECIPE_RECALC_METHOD

    layout = ("one workbook per check (multi-sheet)" if multisheet
              else "one sheet per check (shared workbook)")
    print(f"Ingested {len(stats['chunks'])} export(s) [{layout}]: "
          f"{stats['n_recipes']} recipe(s), {stats['n_checks']} check(s). "
          f"[serialization={stats['serialization']}]")
    for inc in stats.get("incomplete") or []:
        # Not written, and said out loud: a recipe is merged wholesale and its
        # badge is the AND over its checks, so a partial download must never
        # become a published verdict.
        print(f"  INCOMPLETE  {inc['slug']}: {inc['have']} of {inc['want']} "
              f"workbook(s) ingested, missing {inc['missing']} -- NOT written")
    for bad in stats.get("setup_failures") or []:
        print(f"  SETUP ALTERED  {bad['slug']} [{bad['key']}]: {bad['cells']}")
    for ren in stats.get("importer_renames") or []:
        # Loud, and never silently corrected: the importer changed a tab name,
        # so this check ran against a workbook the reference run never saw.
        pairs = ", ".join(f"{a!r} -> {b!r}" for a, b in ren["renamed"].items())
        print(f"  TAB RENAMED BY IMPORTER  {ren['slug']} [{ren['key']}] "
              f"({ren['chunk']}): {pairs} -- recorded, not comparable to the "
              f"LibreOffice run, excluded from that recipe's verdict")
    if not recipes:
        sys.exit("Nothing complete to write: every recipe in this ingest was "
                 "missing at least one of its workbooks. Download the rest and "
                 "re-run; results/ is unchanged.")

    write_recipe_results(args.out, recipes, canary, trusted, args.engine_label,
                         args.allow_label_change,
                         recalc_method=recalc_method,
                         serialization=stats["serialization"],
                         manifest_note=stats["manifest_note"])

    n_ok = sum(1 for r in recipes.values() if r["verified"])
    if stats["manifest_note"]:
        print(f"Manifest note: {stats['manifest_note']}")
    print(f"Deterministic canary OK on all {canary['arithmetic_sheets_checked']} "
          f"sheet(s): {canary['arithmetic_ok']}")
    if canary["arithmetic_sheet_failures"]:
        print(f"  CANARY FAILURES: {canary['arithmetic_sheet_failures']}")
    print(f"Volatile =NOW() per chunk: {canary['volatile_per_chunk']}")
    print(f"Trusted recalculation: {trusted}")
    print(f"Recipes where every check matched its expected result: "
          f"{n_ok} / {len(recipes)}")
    # Name the CHECK that differed, not just the recipe: a recipe is only
    # "verified" when its main example AND all of its variant checks match, so
    # printing the main example's value for a recipe that failed on a variant
    # would read as "got 60, want 60" and look like a harness bug.
    for slug, r in sorted(recipes.items()):
        if r["verified"]:
            continue
        if not r["main_verified"]:
            print(f"  DIFFERS  {slug} [main] {r['formula']}")
            print(f"           got={r['actual']!r} want={r['expected']!r}")
            if r.get("notes"):
                print(f"           {r['notes']}")
        for vi, var in enumerate(r.get("variants") or []):
            for ci, ch in enumerate(var.get("checks") or []):
                if ch["verified"]:
                    continue
                print(f"  DIFFERS  {slug} [{ch.get('key', f'v{vi}c{ci}')}] "
                      f"{ch['formula']}")
                print(f"           got={ch['actual']!r} want={ch['expected']!r}")
                if ch.get("notes"):
                    print(f"           {ch['notes']}")
        for ch in r.get("extra_checks") or []:
            if ch["verified"]:
                continue
            print(f"  DIFFERS  {slug} [{ch.get('key')}] {ch['formula']}")
            print(f"           got={ch['actual']!r} want={ch['expected']!r}")
            if ch.get("notes"):
                print(f"           {ch['notes']}")
    print(f"Wrote {args.out}")
    return recipes


# --------------------------------------------------------------------------
# selftest-recipes
# --------------------------------------------------------------------------

def cmd_selftest_recipes(args):
    """build-recipes -> LibreOffice standing in for Drive -> ingest-recipes,
    then compare every value against results/recipes-verified.json.

    This proves the CELL MAPPING (right formula on the right sheet, right
    anchor, right check_range read back into the right recipe/variant slot),
    nothing about Google Sheets: the values recovered are LibreOffice's, which
    is exactly why they can be checked against the LibreOffice reference run.
    It writes to a scratch file and never into results/."""
    # Build the LIBREOFFICE check set, not the Google Sheets one: the engine
    # standing in for Drive here IS LibreOffice, and the reference values this
    # command proves the mapping against are LibreOffice's. Building the Sheets
    # set would hand soffice the Sheets-only alternative formulas, which have
    # no LibreOffice reference value to compare with (and several of which
    # LibreOffice cannot evaluate at all).
    args.engine = LIBREOFFICE
    print("=== selftest-recipes: build ===")
    manifest = cmd_build_recipes(args)

    chunkdir = os.path.abspath(args.outdir)
    soffice = os.environ.get("SOFFICE_BIN", "soffice")

    with tempfile.TemporaryDirectory() as tmp:
        export_dir = os.path.join(tmp, "exports")
        os.makedirs(export_dir)
        exports = []
        print("\n=== selftest-recipes: simulated export (soffice --convert-to xlsx) ===")
        for chunk in manifest["chunks"]:
            src = os.path.join(chunkdir, chunk["file"])
            outd = os.path.join(tmp, chunk["chunk"])
            os.makedirs(outd, exist_ok=True)
            proc = subprocess.run(
                [soffice, "--headless", "--convert-to", "xlsx", "--outdir", outd, src],
                capture_output=True, text=True, timeout=900)
            produced = os.path.join(outd, chunk["file"])
            if proc.returncode != 0 or not os.path.exists(produced):
                sys.exit(f"soffice conversion failed for {chunk['file']}: "
                         f"{proc.stderr.strip()}")
            dest = os.path.join(export_dir, f"{chunk['chunk']}-export.xlsx")
            shutil.copyfile(produced, dest)
            exports.append(dest)
            print(f"  {chunk['file']} -> {os.path.basename(dest)} "
                  f"({os.path.getsize(dest):,} bytes)")

        print("\n=== selftest-recipes: ingest ===")
        recipes, canary, trusted, stats = ingest_recipe_exports(
            exports, manifest, args.engine_label)

        out = os.path.abspath(args.out)
        if os.path.commonpath([out, os.path.abspath(RESULTS_DIR)]) == \
                os.path.abspath(RESULTS_DIR):
            sys.exit(f"selftest-recipes refuses to write inside results/ ({out}). "
                     f"These are LibreOffice values produced through the Sheets "
                     f"plumbing, not Google Sheets values.")
        if os.path.exists(out):
            os.remove(out)  # scratch file: always a fresh write, never a merge
        write_recipe_results(out, recipes, canary, trusted, args.engine_label,
                             allow_label_change=True,
                             engine=RECIPE_SELFTEST_ENGINE_ID,
                             recalc_method=(
                                 "selftest plumbing check: soffice --convert-to "
                                 "xlsx standing in for the Drive import -- NOT "
                                 "Google Sheets"),
                             serialization=stats["serialization"],
                             manifest_note=stats["manifest_note"])

    # ---- the actual proof: same values as the LibreOffice reference run ----
    if not os.path.exists(LO_RECIPE_RESULTS_PATH):
        sys.exit(f"{LO_RECIPE_RESULTS_PATH} not found -- run "
                 f"scripts/verify_recipes.py first; the selftest compares "
                 f"against it.")
    with open(LO_RECIPE_RESULTS_PATH) as f:
        reference = json.load(f).get("recipes", {})

    stored_by_key = {}
    for chunk in manifest["chunks"]:
        for rec in chunk["recipes"]:
            for entry in rec["checks"]:
                stored_by_key[(rec["slug"], entry["key"])] = entry

    def _ref_checks(slug):
        """Flatten the reference run into {key: {formula, actual}}."""
        ref = reference.get(slug)
        if ref is None:
            return None
        return {k: {"formula": p.get("formula"), "actual": p.get("actual")}
                for k, p in result_checks_by_key(ref).items()}

    print("\n=== selftest-recipes: values vs results/recipes-verified.json ===")
    n_match = n_diff = n_skipped = 0
    for slug in sorted(recipes):
        ref = _ref_checks(slug)
        if ref is None:
            print(f"  {slug}: NOT in the LibreOffice reference run -- cannot compare")
            n_diff += 1
            continue
        got = {k: p["actual"]
               for k, p in result_checks_by_key(recipes[slug]).items()}
        for key in sorted(got, key=lambda k: (k != "main", k)):
            entry = stored_by_key[(slug, key)]
            mine, theirs = got[key], ref.get(key, {}).get("actual")
            # str() on both sides: the reference file was written with
            # json.dump(default=str), so a value that was a list on the way in
            # comes back as a list of the same reprs -- comparing reprs is the
            # apples-to-apples test.
            same = str(mine) == str(theirs)
            if entry.get("differs_from_lo_serialization"):
                # Not comparable, and honestly so: this check was written with
                # the plain function name while the reference run executed the
                # _xlfn. storage form, i.e. genuinely different input bytes.
                n_skipped += 1
                print(f"  n/c  {slug} {key:8} {entry['formula_display'][:52]:54} "
                      f"plain={mine!r} vs _xlfn-run={theirs!r} "
                      f"(different serialization, not comparable)")
                continue
            if same:
                n_match += 1
            else:
                n_diff += 1
                print(f"  DIFF {slug} {key:8} {entry['formula_display'][:52]:54} "
                      f"got={mine!r} reference={theirs!r}")

    print(f"\nChunks ingested        : {len(stats['chunks'])}")
    print(f"Recipes / checks       : {stats['n_recipes']} / {stats['n_checks']}")
    print(f"Per-sheet canaries     : {canary['arithmetic_sheets_checked']} checked, "
          f"ok={canary['arithmetic_ok']}")
    print(f"Volatile =NOW()        : {canary['volatile_per_chunk']}")
    print(f"trusted                : {trusted}")
    print(f"Values identical to LO : {n_match}")
    print(f"Values differing       : {n_diff}")
    print(f"Not comparable         : {n_skipped} (plain-name serialization)")
    print(f"\nWrote scratch results -> {args.out} "
          f"(LibreOffice values; plumbing proof only)")
    if n_diff:
        sys.exit(f"selftest-recipes FAILED: {n_diff} value(s) do not match the "
                 f"LibreOffice reference run -- the cell mapping is wrong.")
    print("selftest-recipes PASSED: every comparable value round-tripped to the "
          "same result the LibreOffice reference run recorded.")


# --------------------------------------------------------------------------
# build-recipes-multisheet: the six recipes `build-recipes` skips
# --------------------------------------------------------------------------
#
# WHY A SEPARATE COMMAND (AND WHY ONE WORKBOOK PER CHECK)
# -------------------------------------------------------
# `build-recipes` puts one worksheet per check into a shared workbook, which
# is why it cannot carry the six recipes whose checks declare `setup_sheets`:
# their formulas name the data tabs LITERALLY, so two checks that both want a
# tab called `Data` -- with different contents -- cannot coexist, and the
# obvious fix (prefix-rename the tabs, rewrite the references) is exactly what
# these recipes must not have done to them:
#
#   * `=INDIRECT("'"&A1&"'!B2")` resolves the tab name from a CELL VALUE, and
#     one check deliberately points A1 at a tab that does NOT exist to assert
#     #REF!. Renaming tabs changes the answer or destroys the assertion.
#   * `=SUM(Q1:Q3!A1)` is a 3-D reference over a SPAN of consecutive tabs, so
#     it depends on sheet ORDER, not just names.
#   * `=$'Q1 Data'.B2`, `=Data.B2` and `=SUM(Q1.A1:Q3.A1)` exist to assert
#     #NAME? -- a rewriter that "fixed" the separator would delete the test.
#
# One workbook per CHECK removes the whole problem instead of solving it: each
# workbook holds exactly the tabs that one check asks for, named exactly as
# its formula spells them, in the order the recipe JSON lists them. Nothing is
# renamed and nothing is rewritten, so what Google Sheets executes is
# byte-for-byte the formula the recipe teaches. The cost is honest and
# unavoidable: 34 Drive round-trips instead of one.
#
# SHEET ORDER IS LOAD-BEARING
# ---------------------------
# `_meta` first, then the formula sheet, then the data tabs in the recipe's
# own JSON order. The data tabs therefore stay CONSECUTIVE, which is what
# makes `Q1:Q3!A1` mean Q1+Q2+Q3 and not something else; scripts/verify_
# recipes.py's LibreOffice reference run has the same property (it appends the
# extra sheets after the formula sheet, in dict order), so the two engines are
# handed the same workbook shape.
#
# WHERE THE CANARY GOES, AND WHERE IT DELIBERATELY DOES NOT
# ---------------------------------------------------------
# `=1111+2222` goes in Z1 of the FORMULA sheet and in `_meta!A2`, not on the
# data tabs. Several of these checks aggregate WHOLE COLUMNS of a data tab
# (`=SUM('Q1 Data'!B:B)`, `=COUNT(Data!B:B)`) or reach across a 3-D span;
# writing extra formulas into those tabs would put harness cells inside the
# very ranges under test. Z1 is outside every range these recipes touch today,
# but "the canary must never be able to change the answer" is the stronger
# rule, and the formula sheet's canary already proves the workbook was
# recalculated -- it is on the sheet whose result we read.
#
# Instead, the data tabs get a check the shared-workbook path cannot do: their
# setup cells are LITERALS, so ingest reads them back and confirms they still
# hold what the builder wrote (`setup_intact`). That catches a tab that Drive
# renamed, dropped or reformatted -- the failure mode that would otherwise
# turn into a fake "Google Sheets disagrees".

DEFAULT_RECIPE_MS_CHUNK_DIR = os.path.join(
    REPO_ROOT, "harness", "recipe_chunks_multisheet")
DEFAULT_RECIPE_MS_EXPORT_DIR = os.path.join(
    REPO_ROOT, "harness", "recipe_exports_multisheet")
# Scratch output for `selftest-recipes-multisheet`. Outside results/ for the
# same reason every other selftest path is: it holds synthesized values.
RECIPE_MS_SELFTEST_DIR = os.path.join(
    REPO_ROOT, "harness", "recipe_selftest_multisheet")
RECIPE_MS_SELFTEST_CHUNK_DIR = os.path.join(RECIPE_MS_SELFTEST_DIR, "chunks")
RECIPE_MS_SELFTEST_RESULTS_PATH = os.path.join(
    RECIPE_MS_SELFTEST_DIR, "plumbing-check.json")

MULTISHEET_MODE = "multisheet"
RECIPE_MS_RECALC_METHOD = (
    "Drive import + xlsx export readback (how-to recipe corpus, multi-sheet: "
    "one workbook per check)")
MS_ID_RE = re.compile(r"(ms-\d+)")


def multisheet_workbook_id(n):
    """Stable id for the nth workbook. Ingest recovers it from the FILENAME,
    so it has to survive a Drive upload/download round-trip -- `ms-007` does,
    and it is then re-verified against the workbook's sheet list."""
    return f"ms-{n:03d}"


def multisheet_formula_sheet_name(recipe_index, check_key, slug, data_tabs):
    """Name for the sheet the formula itself sits on.

    Same `r<index><key>_<slug>` shape `build-recipes` uses, but the `used` set
    is SEEDED with the data tab names (and `_meta`) so the formula sheet can
    never collide with -- or be mistaken for -- a tab the formula references.
    """
    used = {t.lower() for t in data_tabs}
    used.add(META_SHEET.lower())
    name = recipe_sheet_name(recipe_index, check_key, slug, used)
    if name.lower() in {t.lower() for t in data_tabs}:  # unreachable; guard anyway
        raise ValueError(f"formula sheet name {name!r} collides with a data tab")
    return name


def collect_multisheet_recipe_checks(only=None, engine=GOOGLE_SHEETS):
    """Return (multisheet, single_sheet, missing).

    The inverse split of `collect_recipe_checks()`: what that function calls
    `skipped` is precisely what this command BUILDS, and its `buildable`
    single-sheet recipes are what this command declines to build (they belong
    to `build-recipes`). Enumeration, engine scoping and stable keys are
    unchanged -- it is the same call, read the other way round -- so a check
    lands under the same key here as it would there.
    """
    single_sheet, multisheet, missing = collect_recipe_checks(only, engine=engine)
    return multisheet, single_sheet, missing


def build_multisheet_workbook(rec, chk, plain_names=True):
    """One workbook for ONE check. Returns (wb, formula_sheet, data_tabs).

    Every tab the check's `setup_sheets` asks for is created with its EXACT
    name -- no sanitizing, no truncation, no de-duplication -- because the
    formula spells that name literally. A name that .xlsx or Google Sheets
    could not carry unchanged is a hard error here rather than a silent
    rename that would quietly change what the recipe means.
    """
    setup_sheets = chk["setup_sheets"] or {}
    if not setup_sheets:
        raise ValueError(
            f"{rec['slug']} {chk['key']}: no setup_sheets -- this is a "
            f"single-sheet check and belongs in `build-recipes`")

    seen = set()
    for name in setup_sheets:
        assert_sheets_safe_name(name)
        if len(name) > SHEET_NAME_MAX:
            raise ValueError(
                f"{rec['slug']} {chk['key']}: data tab {name!r} is longer than "
                f"{SHEET_NAME_MAX} chars, which .xlsx cannot carry unchanged -- "
                f"and renaming it would change the formula's meaning")
        if name.lower() in seen:
            raise ValueError(f"{rec['slug']} {chk['key']}: duplicate data tab "
                             f"{name!r}")
        seen.add(name.lower())
    data_tabs = list(setup_sheets)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fs_name = multisheet_formula_sheet_name(
        rec["index"], chk["key"], rec["slug"], data_tabs)
    ws = wb.create_sheet(fs_name)
    for addr, val in (chk["setup_cells"] or {}).items():
        ws[addr] = val
    formula = chk["formula"] if plain_names else to_storage_formula_all(chk["formula"])
    if chk["check_range"]:
        # Same reasoning as corpus.build_workbook(): a spilling result must be
        # written as a real array formula over its full range.
        ws[chk["anchor"]] = ArrayFormula(chk["check_range"], formula)
    else:
        ws[chk["anchor"]] = formula
    ws[CANARY_ANCHOR] = CANARY_ARITH_FORMULA

    # Data tabs AFTER the formula sheet and in the recipe's own order, so a
    # 3-D span (Q1:Q3) stays contiguous and in the order the JSON declares.
    for name, cells in setup_sheets.items():
        tab = wb.create_sheet(title=name)
        for addr, val in (cells or {}).items():
            tab[addr] = val

    meta = wb.create_sheet(META_SHEET, 0)
    meta[META_VOLATILE_CELL] = "=NOW()"
    meta[META_ARITH_CELL] = CANARY_ARITH_FORMULA
    assert_sheets_safe_name(META_SHEET)
    return wb, fs_name, data_tabs


def cmd_build_recipes_multisheet(args):
    plain_names = bool(getattr(args, "plain_names", True))
    engine = getattr(args, "engine", GOOGLE_SHEETS)
    multisheet, single_sheet, missing = collect_multisheet_recipe_checks(
        args.only, engine=engine)
    if missing:
        sys.exit(f"No data/recipes/*.json for: {', '.join(missing)}")
    if not multisheet:
        sys.exit("No multi-sheet recipes matched. This command builds ONLY the "
                 "recipes whose checks declare setup_sheets; the single-sheet "
                 "ones are `build-recipes`' job.")

    outdir = os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    # Stale workbooks from a previous build would still be ingestible against
    # the new manifest, so clear them first (same rule as build-recipes).
    for stale in glob.glob(os.path.join(outdir, "ms-*.xlsx")):
        os.remove(stale)

    manifest_chunks = []
    recipes_index = {}
    n = 0
    for rec in multisheet:
        recipes_index[rec["slug"]] = {
            "index": rec["index"],
            "title": rec["title"],
            "n_checks": len(rec["checks"]),
            # Every key this recipe needs before it may be written to the
            # results file. Ingest refuses to emit a half-covered recipe: a
            # recipe is merged WHOLESALE and its `verified` flag is the AND
            # over all its checks, so writing one from a partial download
            # would publish a verdict nobody executed.
            "keys": [c["key"] for c in rec["checks"]],
        }
        for chk in rec["checks"]:
            n += 1
            wb_id = multisheet_workbook_id(n)
            wb, fs_name, data_tabs = build_multisheet_workbook(
                rec, chk, plain_names=plain_names)
            fname = f"{wb_id}-{rec['slug']}-{chk['key']}.xlsx"
            path = os.path.join(outdir, fname)
            wb.save(path)

            stored = (chk["formula"] if plain_names
                      else to_storage_formula_all(chk["formula"]))
            entry = {
                "key": chk["key"],
                "kind": chk["kind"],
                "variant_index": chk["variant_index"],
                "check_index": chk["check_index"],
                "heading": chk["heading"],
                "label": chk["label"],
                "sheet": fs_name,
                "anchor": chk["anchor"],
                "check_range": chk["check_range"],
                "formula_display": chk["formula"],
                "formula_stored_xlsx": stored,
                "differs_from_lo_serialization":
                    stored != to_storage_formula_all(chk["formula"]),
                "serialization": "plain" if plain_names else "xlfn",
                "expected": chk["expected"],
                "engines": chk["engines"],
                # Multi-sheet specifics: the tabs this check needs, in order,
                # and the literal values ingest re-reads to prove they came
                # back intact.
                "data_sheets": data_tabs,
                "setup_cells": chk["setup_cells"] or {},
                "setup_sheets": {name: dict(cells or {})
                                 for name, cells in (chk["setup_sheets"] or {}).items()},
            }
            manifest_chunks.append({
                "chunk": wb_id,
                "file": fname,
                "bytes": os.path.getsize(path),
                "sha256": _sha256(path),
                "n_recipes": 1,
                "n_checks": 1,
                "slugs": [rec["slug"]],
                # Every sheet the export must still contain, checked by ingest
                # before a single value is read.
                "sheets": sorted([fs_name] + data_tabs + [META_SHEET]),
                "formula_sheet": fs_name,
                "data_sheets": data_tabs,
                "recipes": [{
                    "slug": rec["slug"],
                    "title": rec["title"],
                    "index": rec["index"],
                    "n_checks": 1,
                    "n_variants": (1 if chk["variant_index"] is not None else 0),
                    "checks": [entry],
                }],
            })

    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "corpus": "recipes",
        "mode": MULTISHEET_MODE,
        "corpus_source": "data/recipes/*.json",
        "engine_scope": engine,
        "workbook_layout": (
            "ONE WORKBOOK PER CHECK. Sheet order is _meta, the formula sheet, "
            "then the check's data tabs in the recipe's own JSON order -- the "
            "data tabs stay consecutive so a 3-D reference (Q1:Q3!A1) keeps "
            "meaning what the recipe says it means. Data tab names are the "
            "recipe's LITERAL names: nothing is renamed, prefixed or "
            "truncated, because the formulas spell them out."
        ),
        "n_workbooks": len(manifest_chunks),
        "n_recipes": len(recipes_index),
        "n_checks": sum(c["n_checks"] for c in manifest_chunks),
        "subset_only": sorted(set(args.only)) if args.only else None,
        "plain_names": plain_names,
        "serialization": "plain" if plain_names else "xlfn",
        "manifest_note": getattr(args, "manifest_note", None),
        "recipes_index": recipes_index,
        "not_built_single_sheet": [
            {"slug": r["slug"], "n_checks": r["n_checks"]} for r in single_sheet
        ],
        "canary": {
            "arithmetic_formula": CANARY_ARITH_FORMULA,
            "arithmetic_expected": CANARY_ARITH_EXPECTED,
            "arithmetic_cell_formula_sheet": CANARY_ANCHOR,
            "arithmetic_on_data_tabs": False,
            "arithmetic_on_data_tabs_reason": (
                "Data tabs carry only the recipe's literal setup values. Adding "
                "a formula to them would put a harness cell inside ranges these "
                "checks aggregate whole-column and across 3-D spans. Their "
                "integrity is proven instead by reading the setup literals back "
                "(setup_intact)."
            ),
            "meta_sheet": META_SHEET,
            "meta_volatile_cell": META_VOLATILE_CELL,
            "meta_volatile_formula": "=NOW()",
            "meta_arithmetic_cell": META_ARITH_CELL,
        },
        "chunks": manifest_chunks,
    }
    manifest_path = os.path.join(outdir, MANIFEST_NAME)
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2, default=str)
        f.write("\n")

    mode = ("PLAIN NAMES (formulas exactly as authored)" if plain_names
            else "xlfn-translated")
    print(f"Built {manifest['n_workbooks']} workbook(s) -- one per check -- from "
          f"{manifest['n_recipes']} multi-sheet recipe(s) -> {outdir}  [{mode}]")
    total = 0
    print(f"  {'workbook':<58} {'tabs':>4}  {'bytes':>7}")
    for c in manifest_chunks:
        total += c["bytes"]
        flag = "  <-- OVER SIZE BUDGET" if c["bytes"] > SIZE_WARN_BYTES else ""
        print(f"  {c['file']:<58} {len(c['sheets']):>4}  {c['bytes']:>7,}{flag}")
    print(f"  {'TOTAL':<58} {'':>4}  {total:>7,} "
          f"(base64 ~{int(total * 4 / 3):,} bytes)")
    print(f"\nWrote {manifest_path}")
    return manifest


# --------------------------------------------------------------------------
# ingest-recipes, multi-sheet path
# --------------------------------------------------------------------------

def _multisheet_id_from_path(path, manifest):
    """Which manifest workbook an exported file is, from its FILENAME.

    Same contract as `_chunk_id_from_path`, with `ms-NNN` in place of
    `chunk-NN`; the id is then re-verified against the workbook's sheet list
    by the caller, so a file downloaded under the wrong name cannot be
    ingested against another check's cell map."""
    base = os.path.basename(path)
    ids = {c["chunk"] for c in manifest["chunks"]}
    m = MS_ID_RE.search(base)
    if m and m.group(1) in ids:
        return m.group(1)
    raise SystemExit(
        f"Cannot tell which workbook {base!r} is. Keep the uploaded name "
        f"(Drive's .xlsx export already does), or rename it so it contains "
        f"its manifest id (ms-001 ... {sorted(ids)[-1] if ids else 'ms-001'})."
    )


def _literals_match(ws, cells):
    """Does this worksheet hold exactly the setup literals `cells` describes?

    This is the IDENTITY test for a tab whose name did not survive the import.
    Data tabs contain only literals the builder wrote, so their contents are a
    fingerprint. An empty `cells` fingerprints nothing and therefore proves
    nothing -- it returns False rather than matching anything that happens to
    be lying around."""
    if not cells:
        return False
    for addr, wrote in cells.items():
        got = norm(ws[addr].value)
        want = norm(wrote)
        if got != want and str(got) != str(want):
            return False
    return True


def _resolve_export_sheets(wb, entry):
    """Map every sheet the BUILDER wrote to the name it came BACK under.

    Returns (name_map, renames): `name_map[built_name] -> export_name`, and
    `renames` the subset where those differ.

    WHY THIS EXISTS (observed, not hypothetical)
    --------------------------------------------
    Google Drive's .xlsx importer does not always preserve a worksheet name.
    A tab built as `Jon's Data` came back as `Jons Data` -- the apostrophe
    stripped -- and Sheets rewrote the formula's reference to match, so the
    formula computed a real answer against a tab this harness never built.
    Refusing the whole run for that (which is what a flat "is every sheet
    name still here?" check does) throws away 33 good workbooks and, worse,
    hides the finding. Silently accepting it would be the opposite failure:
    the LibreOffice reference run executed against the ORIGINAL name, so the
    two engines were no longer handed the same workbook and their values are
    not comparable.

    So the rule is identity by CONTENTS, never by guesswork. A data tab that
    is missing by name is matched to a leftover tab only when that tab holds
    exactly the setup literals the builder wrote into it, and only when that
    pairing is unambiguous -- one candidate for one expected tab, matching
    nothing else. Anything less (no candidate, several candidates, or one
    candidate that fits two expected tabs equally) is a genuinely missing tab
    and still hard-refuses the workbook, because at that point we cannot say
    what Sheets executed.

    `_meta` and the formula sheet must come back under their own names: they
    are plain ASCII by construction, they carry formulas rather than
    fingerprintable literals, and the anchor we read lives on the formula
    sheet. Nothing is renamed back -- the export is read exactly as it came.
    """
    present = set(wb.sheetnames)
    name_map = {}
    problems = []

    for structural in (META_SHEET, entry["sheet"]):
        if structural in present:
            name_map[structural] = structural
        else:
            problems.append(f"{structural!r} (harness sheet, must come back "
                            f"under its own name)")

    expected = entry.get("setup_sheets") or {}
    unmatched = []
    for name in expected:
        if name in present:
            name_map[name] = name
        else:
            unmatched.append(name)

    renames = {}
    if unmatched:
        claimed = set(name_map.values())
        spare = [s for s in wb.sheetnames
                 if s not in claimed and s not in expected]
        candidates = {name: [s for s in spare
                             if _literals_match(wb[s], expected[name])]
                      for name in unmatched}
        for name in unmatched:
            cands = candidates[name]
            if not cands:
                problems.append(
                    f"{name!r} (absent, and no leftover tab holds its setup "
                    f"literals {expected[name]!r}; leftover tabs were {spare!r})")
                continue
            if len(cands) > 1:
                problems.append(f"{name!r} (absent, and {len(cands)} leftover "
                                f"tabs match its literals: {cands!r} -- "
                                f"ambiguous, cannot prove which one it is)")
                continue
            rival = [o for o in unmatched if o != name and cands[0] in candidates[o]]
            if rival:
                problems.append(f"{name!r} (absent; leftover tab {cands[0]!r} "
                                f"matches it AND {rival!r} equally -- ambiguous)")
                continue
            name_map[name] = cands[0]
            renames[name] = cands[0]

    if problems:
        raise SystemExit(
            "Export is missing sheet(s) this workbook was built with:\n  "
            + "\n  ".join(problems)
            + "\nA tab renamed by the importer is matched by its setup "
              "literals and ingested with a SETUP_ALTERED note; these could "
              "not be identified that way, so what Google executed is "
              "unknown. Wrong file for this workbook, or a stale manifest -- "
              "rebuild and re-upload.")
    return name_map, renames


def _setup_intact_failures(wb, entry, name_map):
    """Every setup literal that did NOT come back the way the builder wrote it.

    Data tabs hold plain values, never formulas, so a Drive round-trip has to
    return them unchanged. If it does not, the check's inputs were not the
    ones the recipe describes and its result says nothing about the formula --
    which is why this is reported rather than absorbed.

    `name_map` is `_resolve_export_sheets()`'s: a tab the importer renamed is
    read under the name it came back as, and its literals still have to match
    (they are what identified it). A rename is reported separately, as a
    rename -- it is not folded in here as if the data had changed."""
    bad = []
    for name, cells in (entry.get("setup_sheets") or {}).items():
        export_name = name_map.get(name)
        if export_name is None or export_name not in wb.sheetnames:
            bad.append({"sheet": name, "cell": None, "wrote": None,
                        "read": "SHEET MISSING"})
            continue
        tab = wb[export_name]
        for addr, wrote in (cells or {}).items():
            got = norm(tab[addr].value)
            want = norm(wrote)
            if got != want and str(got) != str(want):
                bad.append({"sheet": name, "export_sheet": export_name,
                            "cell": addr, "wrote": want, "read": got})
    return bad


def ingest_multisheet_exports(export_paths, manifest, engine_label):
    """Read one-workbook-per-check exports -> (recipe_results, canary,
    trusted, stats).

    Per workbook the verification is the shared-workbook path's, plus the
    multi-sheet extras: every declared tab must still exist, and every setup
    literal must read back unchanged. Values are then evaluated with exactly
    `_read_check_value()` / `compare_check()` -- the same functions
    scripts/verify_recipes.py judges LibreOffice with -- so a difference
    between the two engines is a difference in the engines.

    COMPLETENESS: a recipe is emitted only when EVERY key the manifest lists
    for it was ingested in this call. `write_recipe_results()` replaces a
    recipe wholesale and the site reads `verified` as the AND over its checks,
    so a recipe assembled from half its workbooks would publish a verdict that
    was never executed. Incomplete recipes are reported and left alone.
    """
    plain_names = bool(manifest.get("plain_names", True))
    serialization = manifest.get("serialization", "plain" if plain_names else "xlfn")
    manifest_note = manifest.get("manifest_note")
    recipes_index = manifest.get("recipes_index") or {}

    chunks_by_id = {c["chunk"]: c for c in manifest["chunks"]}
    collected = {}           # slug -> {key: (entry, payload)}
    per_workbook = []
    sheet_canary_failures = []
    setup_failures = []
    importer_renames = []
    n_sheet_canaries = 0
    volatile_values = {}
    meta_arith = {}
    n_checks = 0

    for path in sorted(export_paths):
        wb_id = _multisheet_id_from_path(path, manifest)
        chunk = chunks_by_id[wb_id]
        wb = openpyxl.load_workbook(path, data_only=True)

        rec = chunk["recipes"][0]
        entry = rec["checks"][0]
        try:
            name_map, renames = _resolve_export_sheets(wb, entry)
        except SystemExit as e:
            raise SystemExit(f"{os.path.basename(path)} (claims to be {wb_id}): {e}")
        if renames:
            importer_renames.append({"chunk": wb_id, "slug": rec["slug"],
                                     "key": entry["key"], "renamed": dict(renames)})

        if META_SHEET in wb.sheetnames:
            meta = wb[META_SHEET]
            volatile_values[wb_id] = meta[META_VOLATILE_CELL].value
            meta_arith[wb_id] = meta[META_ARITH_CELL].value
        else:                                   # cannot happen: checked above
            volatile_values[wb_id] = None
            meta_arith[wb_id] = None

        ws = wb[entry["sheet"]]
        n_sheet_canaries += 1
        n_checks += 1

        canary_val = ws[CANARY_ANCHOR].value
        canary_ok = canary_val == CANARY_ARITH_EXPECTED
        if not canary_ok:
            sheet_canary_failures.append(
                {"chunk": wb_id, "sheet": entry["sheet"], "slug": rec["slug"],
                 "key": entry["key"], "value": canary_val})

        bad_setup = _setup_intact_failures(wb, entry, name_map)
        if bad_setup:
            setup_failures.append({"chunk": wb_id, "slug": rec["slug"],
                                   "key": entry["key"], "cells": bad_setup[:10]})

        raw = None if entry["check_range"] else ws[entry["anchor"]].value
        try:
            actual = _read_check_value(ws, entry)
            ok = compare_check(entry["expected"], actual)
        except Exception as e:  # noqa: BLE001 -- report, never abort
            actual = f"ERR {e}"
            ok = False
        ok = bool(ok)

        # A renamed tab means the workbook Google executed is not the workbook
        # the LibreOffice reference run executed, so the two values are not
        # comparable -- whatever came back is recorded verbatim, and the check
        # is excluded from the recipe's verdict rather than counted as a
        # formula divergence.
        comparable = not renames
        not_comparable_reason = None
        lead = []
        if renames:
            not_comparable_reason = "importer_renamed_tab"
            pairs = "; ".join(f"Google Drive import renamed tab {built!r} -> "
                              f"{got_name!r}"
                              for built, got_name in renames.items())
            lead.append(
                f"SETUP_ALTERED: {pairs} (identity proven by reading that tab's "
                f"setup literals back unchanged, not by guessing at the name); "
                f"formula {entry['formula_display']} result recorded against the "
                f"renamed reality. NOT comparable to the LibreOffice reference "
                f"run, which executed against the name as built, and NOT counted "
                f"toward this recipe's verdict. Nothing was renamed back.")
        if bad_setup:
            lead.append(f"SETUP_ALTERED: {len(bad_setup)} setup literal(s) did "
                        f"not survive the round-trip (e.g. {bad_setup[0]}) -- "
                        f"this check's inputs were not the recipe's, so its "
                        f"result says nothing about the formula")
        if not canary_ok:
            lead.insert(0, "UNTRUSTED_RECALC: per-sheet canary failed on this "
                           "workbook's formula sheet")
        notes = lead + _readback_artifact_notes(entry["expected"], raw, actual)
        if entry.get("differs_from_lo_serialization"):
            notes.append(
                "NOTE: written with the PLAIN function name; the LibreOffice "
                "reference run executed the _xlfn. storage form of this formula, "
                "so the two runs are not byte-identical inputs")
        if not ok:
            notes.append(
                (f"DIFFERS from the LibreOffice-authored expectation: expected "
                 f"{entry['expected']!r}, got {actual!r} -- but see SETUP_ALTERED "
                 f"above: this is not a formula divergence, the tabs were not the "
                 f"ones built")
                if not comparable else
                f"MISMATCH vs expected: expected {entry['expected']!r}, "
                f"got {actual!r}")

        payload = {
            "key": entry["key"],
            "label": entry["label"],
            "formula": entry["formula_display"],
            "formula_stored_xlsx": entry["formula_stored_xlsx"],
            "serialization": entry.get("serialization", serialization),
            "expected": entry["expected"],
            "actual": actual,
            "verified": ok,
            "sheet": entry["sheet"],
            "workbook": chunk["file"],
            "data_sheets": entry.get("data_sheets") or [],
            "canary_ok_this_sheet": canary_ok,
            # setup_intact: the literals came back unchanged (true even for a
            # renamed tab -- that is how it was identified). sheet_names_intact:
            # the tabs also came back under the names the builder wrote.
            "setup_intact": not bad_setup,
            "sheet_names_intact": not renames,
            "importer_renamed_tabs": dict(renames),
            "comparable": comparable,
            "not_comparable_reason": not_comparable_reason,
            "notes": "; ".join(notes) if notes else None,
        }
        if entry.get("engines"):
            payload["engines"] = entry["engines"]

        collected.setdefault(rec["slug"], {})[entry["key"]] = (entry, payload)
        per_workbook.append({
            "chunk": wb_id,
            "export_file": os.path.basename(path),
            "slug": rec["slug"],
            "key": entry["key"],
            "n_recipes": 1,
            "n_checks": 1,
        })

    # ---- assemble complete recipes only --------------------------------
    recipes_out = {}
    incomplete = []
    for slug, got in sorted(collected.items()):
        want = list((recipes_index.get(slug) or {}).get("keys") or sorted(got))
        missing_keys = [k for k in want if k not in got]
        if missing_keys:
            incomplete.append({"slug": slug, "have": len(got), "want": len(want),
                               "missing": missing_keys})
            continue
        main = None
        variants = {}
        extra = []
        recipe_ok = True
        not_comparable = []
        renamed_here = {}
        for key in want:
            entry, payload = got[key]
            if payload.get("comparable", True):
                # The verdict is the AND over the checks that ARE comparable.
                # A check whose tabs the importer renamed executed against a
                # workbook neither engine agreed on, so counting it either way
                # would be a claim about a formula nobody tested.
                if not payload["verified"]:
                    recipe_ok = False
            else:
                not_comparable.append(key)
                renamed_here.update(payload.get("importer_renamed_tabs") or {})
            if entry["kind"] == "main":
                main = payload
            elif entry["kind"] == "extra":
                extra.append(payload)
            else:
                variants.setdefault(
                    entry["variant_index"],
                    {"heading": entry["heading"], "checks": []},
                )["checks"].append(payload)
        if main is None:      # cannot happen for a manifest this tool wrote
            raise SystemExit(f"{slug}: manifest has no 'main' check")
        record = {
            "verified": recipe_ok,
            # Which checks the verdict above is (and is not) an AND over.
            "verdict_over_comparable_checks_only": bool(not_comparable),
            "n_checks": len(want),
            "n_not_comparable": len(not_comparable),
            "not_comparable_keys": not_comparable,
            "importer_renamed_tabs": renamed_here,
            "engine": "Google Sheets",
            "engine_label": engine_label,
            "serialization": serialization,
            "multisheet": True,
            "key": main["key"],
            "formula": main["formula"],
            "formula_stored_xlsx": main["formula_stored_xlsx"],
            "expected": main["expected"],
            "actual": main["actual"],
            "main_verified": main["verified"],
            "sheet": main["sheet"],
            "workbook": main["workbook"],
            "data_sheets": main["data_sheets"],
            "canary_ok_this_sheet": main["canary_ok_this_sheet"],
            "setup_intact": main["setup_intact"],
            "sheet_names_intact": main["sheet_names_intact"],
            "comparable": main["comparable"],
            "notes": main["notes"],
        }
        if variants:
            record["variants"] = [variants[i] for i in sorted(variants)]
        if extra:
            record["extra_checks"] = extra
        recipes_out[slug] = record

    arith_ok = (not sheet_canary_failures
                and n_sheet_canaries > 0
                and all(v == CANARY_ARITH_EXPECTED for v in meta_arith.values()))

    canary = {
        "arithmetic_formula": CANARY_ARITH_FORMULA,
        "arithmetic_expected": CANARY_ARITH_EXPECTED,
        "arithmetic_actual": (CANARY_ARITH_EXPECTED if arith_ok
                              else (sheet_canary_failures[0]["value"]
                                    if sheet_canary_failures else None)),
        "arithmetic_ok": arith_ok,
        "arithmetic_sheets_checked": n_sheet_canaries,
        "arithmetic_sheet_failures": sheet_canary_failures[:20],
        "meta_arithmetic_per_chunk": meta_arith,
        "setup_intact_failures": setup_failures[:20],
        "importer_renamed_tabs": importer_renames[:40],
        "volatile_formula": "=NOW()",
        "volatile_per_chunk": {k: str(v) for k, v in volatile_values.items()},
        "now_differs_across_runs": None,
        "method": (
            "One workbook per check. openpyxl writes formulas with NO cached <v> "
            "value, so each uploaded workbook contains zero cached results. "
            "Google Drive's auto-conversion to a Google Sheet recalculates every "
            "formula with Google's engine; exporting back to .xlsx carries those "
            f"values out. The deterministic canary =1111+2222 in {CANARY_ANCHOR} "
            f"of each workbook's FORMULA sheet reading back exactly "
            f"{CANARY_ARITH_EXPECTED} proves genuine computation -- without "
            "recalculation the cell would read back blank (None). The canary is "
            "deliberately NOT written to the data tabs, whose whole columns and "
            "3-D spans these checks aggregate; those tabs are proven instead by "
            "reading their setup literals back unchanged (setup_intact). The "
            "volatile =NOW() canary is recorded per workbook as corroboration, "
            "but a single Drive import yields one timestamp per workbook, so no "
            "cross-run volatile comparison is possible: now_differs_across_runs "
            "is null by design."
        ),
        "engine_label": engine_label,
    }

    stats = {"chunks": per_workbook,
             "n_recipes": len(recipes_out),
             "n_checks": n_checks,
             "incomplete": incomplete,
             "setup_failures": setup_failures,
             "importer_renames": importer_renames,
             "serialization": serialization,
             "manifest_note": manifest_note}
    return recipes_out, canary, arith_ok, stats


# --------------------------------------------------------------------------
# selftest-recipes-multisheet
# --------------------------------------------------------------------------
#
# WHY THIS ONE DOES NOT DRIVE LIBREOFFICE
# ---------------------------------------
# `selftest-recipes` converts its chunks with soffice and then proves the
# recovered values equal results/recipes-verified.json. That works there
# because every check in its slice is one LibreOffice ran. It would NOT work
# here: three of these checks exist to assert LibreOffice-syntax failures
# (#NAME?) and one asserts a #REF! that LibreOffice's IFERROR did not trap --
# behaviour that is the ENGINE's, not the harness's, and re-deriving it would
# prove nothing about the plumbing while making the test depend on which
# LibreOffice build happens to be installed.
#
# So this selftest synthesizes the export instead: it writes each check's
# expected value into the anchor of the workbook the builder just produced,
# stamps the canaries, and pushes the result through the REAL ingest path.
# What that proves is exactly what is in doubt -- that the right formula is on
# the right sheet at the right anchor with the right tabs beside it, that
# ingest reads it back into the right recipe/variant/extra slot, that the
# merge preserves every recipe it did not touch, and that a broken canary is
# caught rather than shrugged off. It proves nothing about any engine, and
# claims nothing about one.

def _simulate_multisheet_export(src_path, entry, dest_path, value,
                                canary_value=CANARY_ARITH_EXPECTED,
                                meta_arith_value=CANARY_ARITH_EXPECTED):
    """Write the workbook Drive WOULD hand back if Sheets computed `value`.

    Formulas become their results, the canaries become their computed
    constants, `_meta!A1` becomes a real timestamp, and the data tabs are left
    exactly as the builder wrote them (literals survive a round-trip, which is
    the property `setup_intact` checks). Nothing here touches Drive or
    LibreOffice."""
    wb = openpyxl.load_workbook(src_path)
    ws = wb[entry["sheet"]]
    if entry["check_range"]:
        grid = cell_addrs_in_range(entry["check_range"])
        flat = [addr for row in grid for addr in row]
        vals = list(value) if isinstance(value, (list, tuple)) else [value]
        for addr in flat:
            ws[addr] = None
        for addr, v in zip(flat, vals):
            ws[addr] = v
    else:
        ws[entry["anchor"]] = value
    ws[CANARY_ANCHOR] = canary_value
    meta = wb[META_SHEET]
    meta[META_VOLATILE_CELL] = datetime.now(timezone.utc).replace(tzinfo=None)
    meta[META_ARITH_CELL] = meta_arith_value
    wb.save(dest_path)
    return dest_path


def cmd_selftest_recipes_multisheet(args):
    """build-recipes-multisheet -> synthesized export -> ingest-recipes,
    with four assertions that each cover one way this could go wrong.

    Never writes inside results/ and never uploads anything."""
    failures = []

    def _check(label, cond, detail=""):
        print(f"  {'PASS' if cond else 'FAIL'}  {label}"
              + (f"   {detail}" if detail else ""))
        if not cond:
            failures.append(f"{label}{(' -- ' + detail) if detail else ''}")

    args.engine = getattr(args, "engine", GOOGLE_SHEETS)
    print("=== selftest-recipes-multisheet: build ===")
    manifest = cmd_build_recipes_multisheet(args)
    chunkdir = os.path.abspath(args.outdir)

    out = os.path.abspath(args.out)
    if os.path.commonpath([out, os.path.abspath(RESULTS_DIR)]) == \
            os.path.abspath(RESULTS_DIR):
        sys.exit(f"selftest-recipes-multisheet refuses to write inside results/ "
                 f"({out}). These are synthesized values, not engine output.")

    label = args.engine_label
    with tempfile.TemporaryDirectory() as tmp:
        export_dir = os.path.join(tmp, "exports")
        os.makedirs(export_dir)

        # ---- 1. round-trip every workbook through the real ingest path ----
        print("\n=== selftest-recipes-multisheet: synthesized export ===")
        exports = []
        entries = {}
        for chunk in manifest["chunks"]:
            entry = chunk["recipes"][0]["checks"][0]
            entries[chunk["chunk"]] = (chunk, entry)
            dest = os.path.join(export_dir, f"{chunk['chunk']}-export.xlsx")
            _simulate_multisheet_export(
                os.path.join(chunkdir, chunk["file"]), entry, dest,
                entry["expected"])
            exports.append(dest)
        print(f"  synthesized {len(exports)} export workbook(s) "
              f"(expected values written into each anchor)")

        print("\n=== selftest-recipes-multisheet: ingest ===")
        recipes, canary, trusted, stats = ingest_multisheet_exports(
            exports, manifest, label)

        n_expected_recipes = manifest["n_recipes"]
        n_expected_checks = manifest["n_checks"]
        _check("every workbook ingested",
               stats["n_checks"] == n_expected_checks,
               f"{stats['n_checks']} / {n_expected_checks} check(s)")
        _check("every multi-sheet recipe assembled",
               len(recipes) == n_expected_recipes,
               f"{len(recipes)} / {n_expected_recipes} recipe(s): "
               f"{sorted(recipes)}")
        _check("no recipe left incomplete", not stats["incomplete"],
               str(stats["incomplete"]))
        _check("every data tab's setup literals survived",
               not stats["setup_failures"], str(stats["setup_failures"][:2]))
        _check("deterministic canary OK on every workbook", canary["arithmetic_ok"],
               f"{canary['arithmetic_sheets_checked']} formula sheet(s) checked")
        _check("run marked trusted", trusted is True)
        bad = [(s, k) for s, r in recipes.items()
               for k, p in result_checks_by_key(r).items() if not p.get("verified")]
        _check("every check evaluated to its expected value", not bad, str(bad[:5]))
        # The mapping proof: each stored payload must carry back the formula the
        # manifest says lives at that key. A transposed anchor or a check written
        # to the wrong sheet shows up here and nowhere else.
        wrong = []
        for chunk, entry in entries.values():
            slug = chunk["recipes"][0]["slug"]
            payload = result_checks_by_key(recipes[slug]).get(entry["key"])
            if payload is None or payload["formula"] != entry["formula_display"]:
                wrong.append((slug, entry["key"]))
        _check("every check landed under its own key with its own formula",
               not wrong, str(wrong[:5]))
        _check("volatile =NOW() recorded per workbook",
               len(canary["volatile_per_chunk"]) == n_expected_checks)

        # ---- 2. the merge preserves every recipe it did not touch ----------
        print("\n=== selftest-recipes-multisheet: merge preservation ===")
        merge_target = os.path.join(tmp, "merge-into.json")
        if os.path.exists(DEFAULT_RECIPE_RESULTS_PATH):
            shutil.copyfile(DEFAULT_RECIPE_RESULTS_PATH, merge_target)
            with open(merge_target) as f:
                before = json.load(f)
            prev_recipes = before.get("recipes") or {}
            prev_label = before.get("engine_label")
            prev_json = {k: json.dumps(v, sort_keys=True, default=str)
                         for k, v in prev_recipes.items()}

            # (a) the engine-date label guard fires without --allow-label-change
            try:
                write_recipe_results(
                    merge_target, recipes, canary, trusted,
                    "Google Sheets (Drive import, 1999-01-01)",
                    allow_label_change=False,
                    engine=RECIPE_SELFTEST_ENGINE_ID,
                    recalc_method=RECIPE_MS_RECALC_METHOD)
                guarded = False
            except SystemExit:
                guarded = True
            _check("engine-label change refused without --allow-label-change",
                   guarded)
            with open(merge_target) as f:
                after_guard = json.load(f)
            _check("refused merge left the target file untouched",
                   json.dumps(after_guard, sort_keys=True, default=str)
                   == json.dumps(before, sort_keys=True, default=str))

            # (b) the real merge, under the file's own label
            write_recipe_results(
                merge_target, recipes, canary, trusted, prev_label,
                allow_label_change=False,
                engine=RECIPE_SELFTEST_ENGINE_ID,
                recalc_method=RECIPE_MS_RECALC_METHOD,
                serialization=stats["serialization"],
                manifest_note=stats["manifest_note"])
            with open(merge_target) as f:
                after = json.load(f)
            after_recipes = after.get("recipes") or {}
            # The invariant is about the recipes this ingest did NOT touch.
            # Once a real Sheets run has landed, the target file already
            # contains these six, so "prev + new" is the wrong arithmetic and
            # "every prior entry unchanged" is the wrong claim -- a re-ingest
            # is SUPPOSED to replace them. What must never move is everything
            # else, which is what the merge exists to protect.
            untouched = {k: blob for k, blob in prev_json.items()
                         if k not in recipes}
            changed = [k for k, blob in untouched.items()
                       if json.dumps(after_recipes.get(k), sort_keys=True,
                                     default=str) != blob]
            _check(f"all {len(untouched)} untouched recipe(s) preserved "
                   f"byte-identical", not changed, str(changed[:5]))
            _check("merged file holds every prior recipe plus the ingested ones",
                   set(after_recipes) == set(prev_json) | set(recipes)
                   and len(after_recipes) == len(set(prev_json) | set(recipes)),
                   f"{len(after_recipes)} vs "
                   f"{len(set(prev_json) | set(recipes))}")
            _check("re-ingested recipes were actually replaced",
                   all(json.dumps(after_recipes[k], sort_keys=True, default=str)
                       == json.dumps(recipes[k], sort_keys=True, default=str)
                       for k in recipes))
            _check("trust is the AND of both runs",
                   after["trusted"] == (bool(before.get("trusted", False))
                                        and trusted))
        else:
            _check(f"{DEFAULT_RECIPE_RESULTS_PATH} exists to merge into", False,
                   "cannot run the merge-preservation assertions")

        # ---- 3. a canary failure is caught --------------------------------
        print("\n=== selftest-recipes-multisheet: canary failure detection ===")
        bad_dir = os.path.join(tmp, "bad-exports")
        os.makedirs(bad_dir)
        bad_exports = []
        first_id = manifest["chunks"][0]["chunk"]
        for chunk in manifest["chunks"]:
            entry = chunk["recipes"][0]["checks"][0]
            dest = os.path.join(bad_dir, f"{chunk['chunk']}-export.xlsx")
            # Exactly what an un-recalculated import looks like: openpyxl never
            # wrote a cached value, so the canary comes back blank.
            _simulate_multisheet_export(
                os.path.join(chunkdir, chunk["file"]), entry, dest,
                entry["expected"],
                canary_value=(None if chunk["chunk"] == first_id
                              else CANARY_ARITH_EXPECTED))
            bad_exports.append(dest)
        bad_recipes, bad_canary, bad_trusted, bad_stats = \
            ingest_multisheet_exports(bad_exports, manifest, label)
        _check("blank canary marks the run untrusted", bad_trusted is False)
        _check("the failing workbook is named",
               len(bad_canary["arithmetic_sheet_failures"]) == 1
               and bad_canary["arithmetic_sheet_failures"][0]["chunk"] == first_id,
               str(bad_canary["arithmetic_sheet_failures"][:1]))
        failed_entry = manifest["chunks"][0]["recipes"][0]["checks"][0]
        failed_slug = manifest["chunks"][0]["recipes"][0]["slug"]
        note = (result_checks_by_key(bad_recipes[failed_slug])
                .get(failed_entry["key"], {}).get("notes") or "")
        _check("the affected check carries UNTRUSTED_RECALC",
               "UNTRUSTED_RECALC" in note, note[:90])

        # ---- 4. a half-downloaded recipe is refused, not half-written -----
        print("\n=== selftest-recipes-multisheet: partial-download guard ===")
        multi = max(manifest["recipes_index"].items(),
                    key=lambda kv: kv[1]["n_checks"])
        multi_slug = multi[0]
        drop = [c["chunk"] for c in manifest["chunks"]
                if c["recipes"][0]["slug"] == multi_slug][:1]
        partial = [p for p in exports
                   if not any(d in os.path.basename(p) for d in drop)]
        part_recipes, _pc, _pt, part_stats = ingest_multisheet_exports(
            partial, manifest, label)
        _check(f"{multi_slug} (missing 1 of {multi[1]['n_checks']} workbooks) "
               f"is NOT written", multi_slug not in part_recipes)
        _check("the incomplete recipe is reported",
               any(i["slug"] == multi_slug for i in part_stats["incomplete"]),
               str(part_stats["incomplete"][:1]))
        _check("the complete recipes still ingest",
               len(part_recipes) == n_expected_recipes - 1,
               f"{len(part_recipes)} / {n_expected_recipes - 1}")

        # ---- 5. an importer-RENAMED tab is recorded, not refused ----------
        # Observed for real: Drive's importer stripped the apostrophe from a tab
        # built as `Jon's Data`. The workbook still computes -- against a tab
        # this harness never built -- so it must ingest, be flagged, and be kept
        # out of the recipe's verdict. None of those three is optional.
        print("\n=== selftest-recipes-multisheet: importer-renamed tab ===")

        def _fixture(chunk, entry, dest, value, rename=None, drop=None,
                     corrupt=None):
            """A synthesized export, then mangled the way an importer might."""
            _simulate_multisheet_export(
                os.path.join(chunkdir, chunk["file"]), entry, dest, value)
            fx = openpyxl.load_workbook(dest)
            if rename:
                fx[rename[0]].title = rename[1]
            if drop:
                fx.remove(fx[drop])
            if corrupt:
                fx[corrupt[0]][corrupt[1]] = corrupt[2]
            fx.save(dest)
            return dest

        # Prefer the real-world case (a tab name with an apostrophe); fall back
        # to any single-data-tab check with a fingerprintable literal.
        def _rename_candidate():
            single = [(c, c["recipes"][0]["checks"][0]) for c in manifest["chunks"]
                      if len(c["recipes"][0]["checks"][0].get("setup_sheets") or {}) == 1
                      and all((cells or {}) for cells in
                              (c["recipes"][0]["checks"][0]["setup_sheets"] or {}).values())]
            apos = [t for t in single if "'" in list(t[1]["setup_sheets"])[0]]
            return (apos or single)[0]

        ren_chunk, ren_entry = _rename_candidate()
        ren_slug = ren_chunk["recipes"][0]["slug"]
        ren_key = ren_entry["key"]
        built_tab = list(ren_entry["setup_sheets"])[0]
        export_tab = built_tab.replace("'", "") or (built_tab + " 1")
        if export_tab == built_tab:
            export_tab = built_tab + " 1"
        fx_dir = os.path.join(tmp, "fixtures")
        os.makedirs(fx_dir)

        def _exports_with(replacement_path):
            return [replacement_path if os.path.basename(p).startswith(ren_chunk["chunk"])
                    else p for p in exports]

        ren_path = _fixture(
            ren_chunk, ren_entry,
            os.path.join(fx_dir, f"{ren_chunk['chunk']}-export.xlsx"),
            ren_entry["expected"], rename=(built_tab, export_tab))
        ren_recipes, _rc, _rt, ren_stats = ingest_multisheet_exports(
            _exports_with(ren_path), manifest, label)
        _check("a renamed tab does NOT block the run",
               len(ren_recipes) == n_expected_recipes
               and ren_stats["n_checks"] == n_expected_checks,
               f"{len(ren_recipes)} recipe(s), {ren_stats['n_checks']} check(s)")
        _check("the rename is reported by name",
               [r["renamed"] for r in ren_stats["importer_renames"]]
               == [{built_tab: export_tab}],
               str(ren_stats["importer_renames"]))
        ren_payload = result_checks_by_key(ren_recipes[ren_slug]).get(ren_key) or {}
        _check("the affected check is flagged not-comparable",
               ren_payload.get("comparable") is False
               and ren_payload.get("not_comparable_reason") == "importer_renamed_tab",
               f"{ren_slug} [{ren_key}]")
        ren_note = ren_payload.get("notes") or ""
        _check("the note records BOTH names verbatim",
               "SETUP_ALTERED" in ren_note and repr(built_tab) in ren_note
               and repr(export_tab) in ren_note, ren_note[:140])
        _check("nothing was renamed back; literals still intact",
               ren_payload.get("setup_intact") is True
               and ren_payload.get("sheet_names_intact") is False
               and ren_payload.get("importer_renamed_tabs") == {built_tab: export_tab})
        _check("the value it actually returned is recorded as-is",
               str(ren_payload.get("actual")) == str(ren_entry["expected"]),
               f"actual={ren_payload.get('actual')!r}")
        others_comparable = sum(
            1 for r in ren_recipes.values()
            for k, pl in result_checks_by_key(r).items()
            if pl.get("comparable", True))
        _check("every other check stays comparable",
               others_comparable == n_expected_checks - 1,
               f"{others_comparable} / {n_expected_checks - 1}")

        # The load-bearing half: a renamed tab whose formula then FAILED must
        # not drag the recipe's badge down, because no engine was tested.
        ren_bad_path = _fixture(
            ren_chunk, ren_entry,
            os.path.join(fx_dir, f"{ren_chunk['chunk']}-bad-export.xlsx"),
            "#REF!", rename=(built_tab, export_tab))
        os.replace(ren_bad_path,
                   os.path.join(fx_dir, f"{ren_chunk['chunk']}-export.xlsx"))
        ren_bad_path = os.path.join(fx_dir, f"{ren_chunk['chunk']}-export.xlsx")
        bad2_recipes, _bc, _bt, _bs = ingest_multisheet_exports(
            _exports_with(ren_bad_path), manifest, label)
        bad2_payload = result_checks_by_key(bad2_recipes[ren_slug]).get(ren_key) or {}
        _check("a renamed tab's failure is NOT counted as a formula divergence",
               bad2_recipes[ren_slug]["verified"] is True
               and bad2_payload.get("verified") is False
               and bad2_payload.get("comparable") is False,
               f"recipe verified={bad2_recipes[ren_slug]['verified']}, "
               f"check verified={bad2_payload.get('verified')}")
        _check("the recipe records which checks its verdict excluded",
               bad2_recipes[ren_slug]["n_not_comparable"] == 1
               and ren_key in bad2_recipes[ren_slug]["not_comparable_keys"]
               and bad2_recipes[ren_slug]["verdict_over_comparable_checks_only"] is True)

        # ---- 6. a genuinely missing tab still refuses ---------------------
        print("\n=== selftest-recipes-multisheet: missing-tab refusal ===")
        gone = _fixture(ren_chunk, ren_entry,
                        os.path.join(fx_dir, "gone.xlsx"),
                        ren_entry["expected"], drop=built_tab)
        try:
            ingest_multisheet_exports([gone], manifest, label)
            refused_missing = False
        except SystemExit:
            refused_missing = True
        _check("a data tab that is simply GONE still hard-refuses",
               refused_missing)

        # Renamed AND altered: the literals no longer fingerprint the tab, so
        # its identity cannot be proven and the workbook must be refused rather
        # than matched on a resemblance.
        addr = list(ren_entry["setup_sheets"][built_tab])[0]
        unprovable = _fixture(
            ren_chunk, ren_entry, os.path.join(fx_dir, "unprovable.xlsx"),
            ren_entry["expected"], rename=(built_tab, export_tab),
            corrupt=(export_tab, addr, "NOT THE VALUE WE WROTE"))
        try:
            ingest_multisheet_exports([unprovable], manifest, label)
            refused_unprovable = False
        except SystemExit:
            refused_unprovable = True
        _check("a renamed tab whose literals do NOT match is refused",
               refused_unprovable)

        # ---- scratch results file, for a human to read -------------------
        if os.path.exists(out):
            os.remove(out)          # scratch: always a fresh write, never a merge
        write_recipe_results(out, recipes, canary, trusted, args.engine_label,
                             allow_label_change=True,
                             engine=RECIPE_SELFTEST_ENGINE_ID,
                             recalc_method=(
                                 "selftest plumbing check: expected values "
                                 "synthesized into the builder's own workbooks "
                                 "-- NOT Google Sheets, NOT LibreOffice"),
                             serialization=stats["serialization"],
                             manifest_note=stats["manifest_note"])

    print(f"\nWorkbooks / checks     : {manifest['n_workbooks']} / {n_expected_checks}")
    print(f"Recipes assembled      : {len(recipes)} ({', '.join(sorted(recipes))})")
    print(f"Wrote scratch results -> {out} (synthesized values; plumbing proof only)")
    if failures:
        print(f"\nselftest-recipes-multisheet FAILED: {len(failures)} assertion(s)")
        for f in failures:
            print(f"  - {f}")
        sys.exit(1)
    print("\nselftest-recipes-multisheet PASSED: every check round-tripped to its "
          "own key, the merge preserved every untouched recipe, the engine-label "
          "guard held, a broken canary was caught, a half-downloaded recipe was "
          "refused, an importer-renamed tab was ingested and flagged rather than "
          "blocking the run or being counted as a formula divergence, and a "
          "genuinely missing (or unidentifiable) tab still refused.")


# --------------------------------------------------------------------------
# selftest-excel-web: prove the excel_web ingest identity + merge discipline
# --------------------------------------------------------------------------

def _simulate_chunk_export(src_path, chunk, dest_path,
                           value_for=None,
                           canary_value=CANARY_ARITH_EXPECTED,
                           meta_arith_value=CANARY_ARITH_EXPECTED,
                           meta_volatile=None):
    """Write the workbook an engine WOULD hand back for a `build` chunk.

    Every case's formula cell becomes its expected value, the per-sheet canary
    becomes its computed constant, and `_meta` gets a volatile timestamp. This
    executes NOTHING -- no LibreOffice, no OneDrive, no Drive -- so it makes no
    claim about any engine's behaviour; it exists to exercise the cell map, the
    canary logic and the merge, which are this harness's code and not an
    engine's.

    `value_for(case) -> value` overrides the value written for a given case
    (used to inject an engine-specific token); default is case["expected"].
    """
    wb = openpyxl.load_workbook(src_path)
    for case in chunk["cases"]:
        ws = wb[case["sheet"]]
        value = case["expected"] if value_for is None else value_for(case)
        if case["check_range"]:
            grid = cell_addrs_in_range(case["check_range"])
            flat = [addr for row in grid for addr in row]
            vals = list(value) if isinstance(value, (list, tuple)) else [value]
            for addr in flat:
                ws[addr] = None
            for addr, v in zip(flat, vals):
                ws[addr] = v
        else:
            ws[case["anchor"]] = value
        ws[CANARY_ANCHOR] = canary_value
    meta = wb[META_SHEET]
    # Excel for the web returns =NOW() as a bare float serial, so the
    # simulation writes a serial too -- otherwise the decode path this engine
    # actually uses would never be exercised.
    if meta_volatile is None:
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        meta_volatile = (now - datetime(1899, 12, 30)).total_seconds() / 86400.0
    meta[META_VOLATILE_CELL] = meta_volatile
    meta[META_ARITH_CELL] = meta_arith_value
    wb.save(dest_path)
    return dest_path


def cmd_selftest_excel_web(args):
    """Fixture-based proof of the excel_web ingest path. Executes no engine.

    What this DOES prove: the engine registry writes an Excel-for-the-web
    identity (never Google's) into its own results file, the Sheets-only
    #ERROR! token is not applied to this engine, a synthesized export is
    correctly flagged as NOT written by Excel, the canary marks a bad run
    untrusted, and -- the main event -- the keyed incremental merge preserves
    every untouched function byte-identically, refuses a date-label change
    without --allow-label-change, and refuses an engine-IDENTITY change
    outright.

    What this does NOT prove, and must never be read as proving: anything
    whatsoever about how Excel for the web computes. The values here are the
    corpus's own documented expecteds written back into the builder's
    workbook. Real Excel-web behaviour lives in results/excel-web.json, and
    only a real OneDrive round-trip can put it there.
    """
    failures = []

    def _check(label, cond, detail=""):
        print(f"  {'PASS' if cond else 'FAIL'}  {label}"
              + (f"   {detail}" if detail else ""))
        if not cond:
            failures.append(f"{label}{(' -- ' + detail) if detail else ''}")

    # Build the fixture from data/tests/ rather than depending on a committed
    # workbook, exactly like every other chunk dir in this harness: the .xlsx
    # is reproducible at any time and only ingested RESULTS are committed.
    chunkdir = os.path.abspath(args.chunkdir)
    print(f"=== selftest-excel-web: build fixture ({len(args.only)} function(s)) ===")
    manifest = cmd_build(argparse.Namespace(
        chunk_size=args.chunk_size, only=list(args.only), outdir=chunkdir,
        plain_names=False, manifest_note=args.manifest_note))
    chunk = manifest["chunks"][0]
    src = os.path.join(chunkdir, chunk.get("file") or f"{chunk['chunk']}.xlsx")
    if not os.path.exists(src):
        sys.exit(f"Chunk workbook not found: {src}")

    out = os.path.abspath(args.out)
    if os.path.commonpath([out, os.path.abspath(RESULTS_DIR)]) == \
            os.path.abspath(RESULTS_DIR):
        sys.exit(f"selftest-excel-web refuses to write inside results/ ({out}). "
                 f"These are synthesized values, not engine output.")

    label = args.engine_label
    spec = _engine_spec(EXCEL_WEB_ENGINE_ID)
    print(f"=== selftest-excel-web: synthesized export from {chunk['chunk']} "
          f"({chunk['n_functions']} fn / {chunk['n_cases']} cases) ===")
    print("    executes NO engine -- values are the corpus's own expecteds")

    with tempfile.TemporaryDirectory() as tmp:
        export = _simulate_chunk_export(
            src, chunk, os.path.join(tmp, f"{chunk['chunk']}-export.xlsx"))

        # ---- 1. identity: this must never come back as Google Sheets ------
        print("\n=== selftest-excel-web: engine identity ===")
        fr, canary, trusted, stats = ingest_exports(
            [export], manifest, label, engine=EXCEL_WEB_ENGINE_ID)
        _check("every function ingested",
               len(fr) == chunk["n_functions"],
               f"{len(fr)} / {chunk['n_functions']}")
        _check("every case ingested",
               stats["n_cases"] == chunk["n_cases"],
               f"{stats['n_cases']} / {chunk['n_cases']}")
        _check("stats report the excel_web engine",
               stats["engine"] == EXCEL_WEB_ENGINE_ID, stats["engine"])
        _check("canary OK on every sheet", canary["arithmetic_ok"],
               f"{canary['arithmetic_sheets_checked']} sheet(s)")
        _check("run marked trusted", trusted is True)
        _check("canary prose describes Excel for the web, not Drive",
               "Excel for the web" in canary["method"]
               and "Google Drive" not in canary["method"])
        # The mapping proof: each case must carry back the formula the
        # manifest says lives at that cell.
        wrong = [c["test_id"] for c in chunk["cases"]
                 if fr.get(c["function"], {}).get(c["test_id"], {})
                 .get("formula_display") != c["formula_display"]]
        _check("every case landed under its own id with its own formula",
               not wrong, str(wrong[:5]))
        _check("volatile =NOW() decoded from a float serial",
               bool(canary.get("volatile_decoded_per_chunk")),
               str(canary.get("volatile_decoded_per_chunk")))

        # ---- 2. a synthesized export is NOT passed off as Excel's ---------
        print("\n=== selftest-excel-web: package provenance ===")
        _check("openpyxl-written export is flagged as not-Excel-written",
               stats["app_provenance_ok"] is False,
               str(stats["app_provenance"]))
        _check("the writing application is recorded verbatim anyway",
               "chunk-01" in (canary.get("app_provenance_per_chunk") or {}))

        # ---- 3. the Sheets-only #ERROR! token is not applied here ---------
        print("\n=== selftest-excel-web: error vocabulary ===")
        target = chunk["cases"][0]
        err_export = _simulate_chunk_export(
            src, chunk, os.path.join(tmp, "chunk-01-export-err.xlsx"),
            value_for=lambda c: ("#ERROR!" if c["test_id"] == target["test_id"]
                                 else c["expected"]))
        xw_fr, _c, _t, _s = ingest_exports(
            [err_export], manifest, label, engine=EXCEL_WEB_ENGINE_ID)
        gs_fr, _c, _t, _s = ingest_exports(
            [err_export], manifest, label, engine=ENGINE_ID)
        xw_case = xw_fr[target["function"]][target["test_id"]]
        gs_case = gs_fr[target["function"]][target["test_id"]]
        _check("#ERROR! is NOT an error token for excel_web",
               xw_case["error"] is None, repr(xw_case["error"]))
        _check("#ERROR! IS still an error token for google_sheets",
               gs_case["error"] == "#ERROR!", repr(gs_case["error"]))
        _check("the Google-specific #ERROR! note is not attached to excel_web",
               "parse-failure" not in (xw_case["notes"] or ""))

        # ---- 4. the merge preserves every function it did not touch -------
        print("\n=== selftest-excel-web: merge preservation ===")
        # A subset ingest: only these functions are re-executed; every other
        # function in the target file must survive byte-identical, including
        # the executed_at date of the run that produced it.
        subset_fns = sorted({c["function"] for c in chunk["cases"]})[:2]
        sub_manifest = json.loads(json.dumps(manifest))
        sub_chunk = sub_manifest["chunks"][0]
        sub_chunk["cases"] = [c for c in sub_chunk["cases"]
                              if c["function"] in subset_fns]
        sub_fr, sub_canary, sub_trusted, sub_stats = ingest_exports(
            [export], sub_manifest, label, engine=EXCEL_WEB_ENGINE_ID)
        _check(f"subset ingest carries only {subset_fns}",
               sorted(sub_fr) == subset_fns, str(sorted(sub_fr)))

        merge_target = os.path.join(tmp, "merge-into.json")
        # Seed the target with the FULL run, stamped with an older date, so
        # "preserved unchanged" is a claim with teeth.
        write_results(merge_target, json.loads(json.dumps(fr)), canary, True,
                      label, engine=EXCEL_WEB_ENGINE_ID,
                      recalc_method=spec["recalc_method"],
                      readback_artifacts=spec["readback_artifacts"])
        with open(merge_target) as f:
            before = json.load(f)
        prev_json = {k: json.dumps(v, sort_keys=True, default=str)
                     for k, v in before["function_results"].items()}

        # (a) a DATE-label change is refused without --allow-label-change
        try:
            write_results(merge_target, sub_fr, sub_canary, sub_trusted,
                          "Excel for the web (recalc, 1999-01-01)",
                          allow_label_change=False,
                          engine=EXCEL_WEB_ENGINE_ID,
                          recalc_method=spec["recalc_method"])
            guarded = False
        except SystemExit:
            guarded = True
        _check("date-label change refused without --allow-label-change", guarded)
        with open(merge_target) as f:
            after_guard = json.load(f)
        _check("refused merge left the target file untouched",
               json.dumps(after_guard, sort_keys=True, default=str)
               == json.dumps(before, sort_keys=True, default=str))

        # (b) an engine-IDENTITY change is refused EVEN WITH the flag. This is
        #     the guard that keeps Excel-web values out of Google's file (and
        #     vice versa) -- a label flag must never be able to authorise it.
        try:
            write_results(merge_target, sub_fr, sub_canary, sub_trusted,
                          "Google Sheets (Drive import, 2026-09-01)",
                          allow_label_change=True,
                          engine=ENGINE_ID, recalc_method=RECALC_METHOD)
            id_guarded = False
        except SystemExit:
            id_guarded = True
        _check("engine-IDENTITY change refused even with --allow-label-change",
               id_guarded)
        with open(merge_target) as f:
            after_id = json.load(f)
        _check("refused identity merge left the target file untouched",
               json.dumps(after_id, sort_keys=True, default=str)
               == json.dumps(before, sort_keys=True, default=str))

        # (c) the real subset merge, under the file's own label
        write_results(merge_target, sub_fr, sub_canary, sub_trusted, label,
                      allow_label_change=False, engine=EXCEL_WEB_ENGINE_ID,
                      recalc_method=spec["recalc_method"],
                      serialization=sub_stats["serialization"],
                      manifest_note=sub_stats["manifest_note"],
                      readback_artifacts=spec["readback_artifacts"])
        with open(merge_target) as f:
            after = json.load(f)
        after_fr = after["function_results"]
        untouched = {k: b for k, b in prev_json.items() if k not in sub_fr}
        changed = [k for k, b in untouched.items()
                   if json.dumps(after_fr.get(k), sort_keys=True,
                                 default=str) != b]
        _check(f"all {len(untouched)} untouched function(s) preserved "
               f"byte-identical", not changed, str(changed[:5]))
        _check("merged file holds every prior function plus the ingested ones",
               set(after_fr) == set(prev_json) | set(sub_fr))
        _check("re-ingested functions were actually replaced",
               all(json.dumps(after_fr[k], sort_keys=True, default=str)
                   == json.dumps(sub_fr[k], sort_keys=True, default=str)
                   for k in sub_fr))
        _check("every untouched function kept its own executed_at",
               all(after_fr[k].get("executed_at")
                   == before["function_results"][k].get("executed_at")
                   for k in untouched))
        _check("the merged file still records the excel_web engine",
               after["engine"] == EXCEL_WEB_ENGINE_ID, after["engine"])
        _check("the merged file still records the Excel-web recalc method",
               after["recalc_method"] == spec["recalc_method"])
        _check("readback_artifacts survive the merge",
               "readback_artifacts" in after)
        _check("the subset run is recorded for audit",
               after["subset_runs"][-1]["functions"] == subset_fns,
               str(after.get("subset_runs", [])[-1:]))

        # (d) trust can only ever go down
        bad_export = _simulate_chunk_export(
            src, chunk, os.path.join(tmp, "chunk-01-export-bad.xlsx"),
            canary_value=None)
        bad_fr, bad_canary, bad_trusted, _bs = ingest_exports(
            [bad_export], manifest, label, engine=EXCEL_WEB_ENGINE_ID)
        print("\n=== selftest-excel-web: canary failure detection ===")
        _check("blank canary marks the run untrusted", bad_trusted is False)
        _check("the failing sheets are named",
               len(bad_canary["arithmetic_sheet_failures"]) > 0)
        note = (bad_fr[target["function"]][target["test_id"]]["notes"] or "")
        _check("the affected case carries UNTRUSTED_RECALC",
               "UNTRUSTED_RECALC" in note, note[:80])
        write_results(merge_target, bad_fr, bad_canary, bad_trusted, label,
                      allow_label_change=False, engine=EXCEL_WEB_ENGINE_ID,
                      recalc_method=spec["recalc_method"])
        with open(merge_target) as f:
            after_bad = json.load(f)
        _check("an untrusted merge downgrades the file's trust",
               after_bad["trusted"] is False)

        # ---- 5. the site must not file this under desktop Excel -----------
        print("\n=== selftest-excel-web: site engine classification ===")
        try:
            site_src = os.path.join(REPO_ROOT, "site", "build_site.py")
            ns = {}
            with open(site_src) as f:
                text = f.read()
            start = text.index("def engine_key_from_engine_name")
            exec(text[start:text.index("def iso_date")], ns)  # noqa: S102
            keyfn = ns["engine_key_from_engine_name"]
            _check("engine id 'excel_web' is NOT classified as desktop 'excel'",
                   keyfn(EXCEL_WEB_ENGINE_ID) != "excel",
                   f"got {keyfn(EXCEL_WEB_ENGINE_ID)!r}")
            _check("the dated Excel-web label is NOT classified as 'excel'",
                   keyfn(spec["label_template"].format(date="2026-09-01"))
                   != "excel")
            _check("desktop Excel still classifies as 'excel'",
                   keyfn("excel") == "excel")
            _check("Google Sheets still classifies as 'google_sheets'",
                   keyfn("google_sheets") == "google_sheets")
        except (OSError, ValueError, KeyError) as exc:
            _check("site engine classification checked", False, repr(exc))

        # Write the scratch results file so the run leaves an inspectable
        # artifact, exactly like the other selftests.
        os.makedirs(os.path.dirname(out), exist_ok=True)
        if os.path.exists(out):
            os.remove(out)
        write_results(out, fr, canary, trusted, label,
                      engine=EXCEL_WEB_SELFTEST_ENGINE_ID,
                      recalc_method=EXCEL_WEB_SELFTEST_RECALC_METHOD,
                      readback_artifacts=spec["readback_artifacts"])

    print(f"\nWrote scratch results -> {out} "
          f"(synthesized values; plumbing proof only)")
    if failures:
        print(f"\nselftest-excel-web: {len(failures)} FAILURE(S)")
        for f_ in failures:
            print(f"  - {f_}")
        sys.exit(1)
    print("\nselftest-excel-web: all checks passed. Proven: the excel_web "
          "ingest writes an Excel-for-the-web identity into its own results "
          "file and never Google's, the Sheets-only #ERROR! token is not "
          "applied to it, a synthesized export is flagged as not written by "
          "Excel, a blank canary marks the run untrusted and downgrades the "
          "merged file, the keyed merge preserves every untouched function "
          "byte-identical with its own executed_at, a date-label change is "
          "refused without --allow-label-change, an engine-identity change is "
          "refused even with it, and the site generator does not file this "
          "engine under desktop Excel. NOTHING here claims anything about how "
          "Excel for the web computes.")


# --------------------------------------------------------------------------

def main():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    default_label = f"Google Sheets (Drive import, {today})"

    ap = argparse.ArgumentParser(
        prog="run_sheets.py",
        description="Google Sheets engine runner: build chunked .xlsx workbooks "
                    "for Drive import, and ingest the exported .xlsx readback.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__.split("USAGE\n-----\n", 1)[1].split("\n\nINCREMENTAL")[0],
    )
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="emit chunk-NN.xlsx workbooks + manifest.json")
    b.add_argument("--chunk-size", type=int, default=40,
                   help="max FUNCTIONS per chunk workbook (default 40)")
    b.add_argument("--only", nargs="+", metavar="FN",
                   help="build only these functions (default: all of data/tests)")
    b.add_argument("--outdir", default=DEFAULT_CHUNK_DIR,
                   help=f"output directory (default {DEFAULT_CHUNK_DIR})")
    b.add_argument("--plain-names", action="store_true",
                   help="write formulas EXACTLY as authored in data/tests -- no "
                        "_xlfn./_xlfn._xlws. storage-form translation at all. Use "
                        "for functions Google Sheets recognizes by their bare name "
                        "but whose xlfn/xlws-prefixed storage form its xlsx "
                        "importer does NOT map (verified: FILTER, SORT, and the "
                        "LAMBDA-family functions), which otherwise show a false "
                        "#NAME?/#ERROR! that looks like a real support gap. See "
                        "README.md 'Phase 2: Google Sheets runner'.")
    b.add_argument("--manifest-note", default=None, metavar="TEXT",
                   help="free-text note stored in manifest.json's top-level "
                        "'manifest_note' field and carried into ingest's results "
                        "provenance (subset_runs entries)")
    b.set_defaults(func=cmd_build)

    i = sub.add_parser("ingest", help="read exported .xlsx back into results JSON")
    i.add_argument("--export", nargs="+", required=True, metavar="XLSX",
                   help="exported workbook(s), named <chunk-id>-export.xlsx")
    i.add_argument("--engine", default=ENGINE_ID, choices=sorted(INGEST_ENGINES),
                   help="which engine executed these exports. The chunk "
                        "workbooks are engine-neutral (formula-only, zero "
                        "cached values), so the engine is declared at INGEST "
                        "time and it selects the results identity, the error "
                        "vocabulary, the canary prose and the default "
                        "--out/--chunkdir/--engine-label (default %(default)s)")
    i.add_argument("--chunkdir", default=None,
                   help="directory holding manifest.json (default: the "
                        "selected engine's chunk dir)")
    i.add_argument("--manifest", help="explicit path to manifest.json")
    i.add_argument("--out", default=None,
                   help=f"results file (default: the selected engine's, e.g. "
                        f"{DEFAULT_RESULTS_PATH} or {EXCEL_WEB_RESULTS_PATH})")
    i.add_argument("--engine-label", default=None,
                   help="honest engine_version label. Neither hosted engine "
                        "exposes a version, so this records the execution "
                        "DATE (default: the selected engine's dated template, "
                        "e.g. %r or %r)"
                        % (INGEST_ENGINES[ENGINE_ID]["label_template"]
                           .format(date=today),
                           INGEST_ENGINES[EXCEL_WEB_ENGINE_ID]["label_template"]
                           .format(date=today)))
    i.add_argument("--allow-label-change", action="store_true",
                   help="permit merging into a file recorded under a different "
                        "engine label (records both in engine_version_history). "
                        "Covers a DATE change on one engine only -- an engine "
                        "IDENTITY change is refused regardless")
    i.set_defaults(func=cmd_ingest)

    s = sub.add_parser("selftest",
                       help="dry run: build + LibreOffice-simulated export + "
                            "ingest, to a scratch results file")
    s.add_argument("--chunk-size", type=int, default=40)
    s.add_argument("--only", nargs="+", metavar="FN",
                   default=["COUNT", "SUM", "MROUND", "XLOOKUP", "DATEDIF", "ISNUMBER"])
    s.add_argument("--outdir", default=DEFAULT_CHUNK_DIR)
    s.add_argument("--out", default=SELFTEST_RESULTS_PATH,
                   help="scratch results path (never results/google-sheets.json -- "
                        "these are LibreOffice values, so they must not land in "
                        "results/ alongside real engine results)")
    s.add_argument("--engine-label",
                   default="SELFTEST (LibreOffice values via simulated export "
                           "- NOT Google Sheets)")
    s.add_argument("--allow-label-change", action="store_true", default=True,
                   help=argparse.SUPPRESS)
    s.add_argument("--plain-names", action="store_true",
                   help="dry-run the --plain-names build path too (see `build --help`)")
    s.add_argument("--manifest-note", default=None, metavar="TEXT", help=argparse.SUPPRESS)
    s.set_defaults(func=cmd_selftest)

    # ---- recipe corpus -----------------------------------------------------
    br = sub.add_parser(
        "build-recipes",
        help="emit chunk-NN.xlsx workbooks + manifest.json for the how-to "
             "RECIPE corpus (data/recipes/*.json)",
        description="One worksheet per recipe check (the main worked example, "
                    "plus every variant check), each carrying that check's "
                    "setup_cells, its formula at the same anchor "
                    "scripts/verify_recipes.py uses, and the deterministic "
                    "=1111+2222 canary in Z1. Recipes needing extra worksheets "
                    "(setup_sheets) are SKIPPED and listed -- see the manifest's "
                    "'skipped_multi_sheet_reason'.")
    br.add_argument("--chunk-size", type=int, default=60,
                    help="max RECIPES per chunk workbook (default %(default)s)")
    br.add_argument("--only", nargs="+", metavar="SLUG",
                    help="build only these recipe slugs (default: all of "
                         "data/recipes)")
    br.add_argument("--outdir", default=DEFAULT_RECIPE_CHUNK_DIR,
                    help="output directory (default %(default)s)")
    br.add_argument("--plain-names", dest="plain_names", action="store_true",
                    default=True,
                    help="write every formula EXACTLY as authored, with no "
                         "_xlfn./_xlfn._xlws. storage-form translation. THIS IS "
                         "THE DEFAULT for recipes: Google Sheets maps bare "
                         "modern names on xlsx import but does NOT map "
                         "_xlfn._xlws.FILTER/SORT or the LAMBDA family, and a "
                         "recipe is by definition the formula a user types.")
    br.add_argument("--xlfn-names", dest="plain_names", action="store_false",
                    help="opt out of --plain-names and write the OOXML storage "
                         "form instead (what the LibreOffice reference run "
                         "executes)")
    br.add_argument("--manifest-note", default=None, metavar="TEXT",
                    help="free-text note stored in the manifest and carried into "
                         "ingest provenance")
    br.add_argument("--engine", default=GOOGLE_SHEETS,
                    choices=[GOOGLE_SHEETS, LIBREOFFICE],
                    help="engine the chunks are built FOR; checks carrying an "
                         "\"engines\" list that excludes it are left out "
                         "(default %(default)s)")
    br.set_defaults(func=cmd_build_recipes)

    ir = sub.add_parser("ingest-recipes",
                        help="read exported recipe .xlsx back into "
                             "results/recipes-verified-sheets.json")
    ir.add_argument("--export", nargs="+", metavar="XLSX", default=None,
                    help="exported workbook(s). Shared-workbook builds are named "
                         "<chunk-id>-export.xlsx; multi-sheet builds keep the "
                         "uploaded name, which already carries its ms-NNN id")
    ir.add_argument("--export-dir", nargs="+", metavar="DIR", default=None,
                    help="directory/-ies of exported .xlsx files to ingest (every "
                         "*.xlsx in them). This is the usual way to ingest a "
                         "MULTI-SHEET build, which produces one export per check")
    ir.add_argument("--chunkdir", default=DEFAULT_RECIPE_CHUNK_DIR,
                    help="directory holding the recipe manifest.json (default "
                         "%(default)s; point it at recipe_chunks_multisheet/ to "
                         "ingest a multi-sheet build -- the manifest's \"mode\" "
                         "selects the reader, no flag needed)")
    ir.add_argument("--manifest", help="explicit path to manifest.json")
    ir.add_argument("--out", default=DEFAULT_RECIPE_RESULTS_PATH,
                    help="results file (default %(default)s)")
    ir.add_argument("--engine-label", default=default_label,
                    help="honest engine label; Sheets has no version, so this "
                         "records the import DATE (default %(default)r)")
    ir.add_argument("--allow-label-change", action="store_true",
                    help="permit merging into a file recorded under a different "
                         "engine label (records both in engine_label_history)")
    ir.set_defaults(func=cmd_ingest_recipes)

    bm = sub.add_parser(
        "build-recipes-multisheet",
        help="emit ONE workbook per check for the six MULTI-SHEET recipes that "
             "`build-recipes` skips",
        description="The six recipes whose checks declare setup_sheets cannot "
                    "share a workbook: their formulas name the data tabs "
                    "literally (and one deliberately names a tab that must NOT "
                    "exist), two tabs called Data with different contents "
                    "collide, and a 3-D reference depends on tab ORDER. This "
                    "command gives each CHECK its own workbook holding exactly "
                    "the tabs it asks for, named and ordered exactly as the "
                    "recipe declares them -- nothing renamed, nothing rewritten. "
                    "Cost: one Drive round-trip per check.")
    bm.add_argument("--only", nargs="+", metavar="SLUG",
                    help="build only these recipe slugs (default: all six "
                         "multi-sheet recipes)")
    bm.add_argument("--outdir", default=DEFAULT_RECIPE_MS_CHUNK_DIR,
                    help="output directory (default %(default)s)")
    bm.add_argument("--plain-names", dest="plain_names", action="store_true",
                    default=True,
                    help="write every formula EXACTLY as authored, with no "
                         "_xlfn. storage-form translation. THE DEFAULT, for the "
                         "same reason as build-recipes: Sheets maps bare modern "
                         "names on xlsx import, and a recipe is by definition "
                         "the formula a user types")
    bm.add_argument("--xlfn-names", dest="plain_names", action="store_false",
                    help="opt out of --plain-names and write the OOXML storage "
                         "form instead")
    bm.add_argument("--manifest-note", default=None, metavar="TEXT",
                    help="free-text note stored in the manifest and carried into "
                         "ingest provenance")
    bm.add_argument("--engine", default=GOOGLE_SHEETS,
                    choices=[GOOGLE_SHEETS, LIBREOFFICE],
                    help="engine the workbooks are built FOR; checks carrying an "
                         "\"engines\" list that excludes it are left out "
                         "(default %(default)s)")
    bm.set_defaults(func=cmd_build_recipes_multisheet)

    sm = sub.add_parser(
        "selftest-recipes-multisheet",
        help="dry run of the multi-sheet pipeline: build + SYNTHESIZED export + "
             "ingest, with merge/canary/partial-download assertions",
        description="Builds the per-check workbooks, writes each check's "
                    "expected value into the workbook the builder just made "
                    "(that is the simulated Drive export), and pushes the result "
                    "through the real ingest path. Proves the cell mapping, the "
                    "keyed merge (every pre-existing recipe byte-identical), the "
                    "engine-label guard, canary-failure detection, the "
                    "partial-download refusal, and the importer-renamed-tab "
                    "path (ingested and flagged, not blocked and not counted as "
                    "a divergence) against a still-hard refusal for a tab that "
                    "is genuinely gone. Executes NO engine: it neither "
                    "drives LibreOffice nor touches Drive, and claims nothing "
                    "about either engine's behaviour.")
    sm.add_argument("--only", nargs="+", metavar="SLUG", default=None)
    sm.add_argument("--outdir", default=RECIPE_MS_SELFTEST_CHUNK_DIR)
    sm.add_argument("--out", default=RECIPE_MS_SELFTEST_RESULTS_PATH,
                    help="scratch results path (never inside results/ -- these "
                         "are synthesized values, not engine output)")
    sm.add_argument("--engine-label",
                    default="SELFTEST (synthesized values via simulated export "
                            "- NOT Google Sheets)")
    sm.add_argument("--plain-names", dest="plain_names", action="store_true",
                    default=True, help=argparse.SUPPRESS)
    sm.add_argument("--xlfn-names", dest="plain_names", action="store_false",
                    help="dry-run the --xlfn-names build path instead")
    sm.add_argument("--manifest-note", default=None, metavar="TEXT",
                    help=argparse.SUPPRESS)
    sm.set_defaults(func=cmd_selftest_recipes_multisheet, engine=GOOGLE_SHEETS)

    sxw = sub.add_parser(
        "selftest-excel-web",
        help="dry run of the excel_web ingest identity + merge discipline "
             "(fixture-based; executes no engine)",
        description="Synthesizes an export by writing the corpus's own "
                    "expected values into the builder's chunk workbook, then "
                    "pushes it through the real ingest path under --engine "
                    "excel_web. Proves the Excel-for-the-web results identity "
                    "(never Google's), that the Sheets-only #ERROR! token is "
                    "not applied to this engine, that an openpyxl-written "
                    "package is flagged as not written by Excel, canary "
                    "failure detection, and the keyed incremental merge: "
                    "untouched functions byte-identical with their own "
                    "executed_at, a date-label change refused without "
                    "--allow-label-change, and an engine-IDENTITY change "
                    "refused even with it. Executes NO engine -- no "
                    "LibreOffice, no OneDrive, no Drive -- and therefore "
                    "claims nothing about how Excel for the web computes.")
    sxw.add_argument("--chunkdir", default=EXCEL_WEB_SELFTEST_CHUNK_DIR,
                     help="scratch dir the fixture chunk is built into "
                          "(default %(default)s)")
    sxw.add_argument("--only", nargs="+", metavar="FN",
                     default=list(EXCEL_WEB_SELFTEST_FUNCTIONS),
                     help="functions to build the fixture from (default: the "
                          "same six the Excel-web probe covered)")
    sxw.add_argument("--chunk-size", type=int, default=40)
    sxw.add_argument("--manifest-note", default=None, metavar="TEXT",
                     help=argparse.SUPPRESS)
    sxw.add_argument("--out", default=EXCEL_WEB_SELFTEST_RESULTS_PATH,
                     help="scratch results path (never inside results/ -- these "
                          "are synthesized values, not engine output)")
    sxw.add_argument("--engine-label",
                     default="SELFTEST (synthesized values via simulated export "
                             "- NOT Excel for the web)")
    sxw.set_defaults(func=cmd_selftest_excel_web)

    sr = sub.add_parser("selftest-recipes",
                        help="dry run of the recipe pipeline: build + "
                             "LibreOffice-simulated export + ingest, then "
                             "compare every value to results/recipes-verified.json")
    sr.add_argument("--chunk-size", type=int, default=60)
    sr.add_argument("--only", nargs="+", metavar="SLUG",
                    default=list(RECIPE_SELFTEST_SLUGS))
    sr.add_argument("--outdir", default=RECIPE_SELFTEST_CHUNK_DIR)
    sr.add_argument("--out", default=RECIPE_SELFTEST_RESULTS_PATH,
                    help="scratch results path (never inside results/ -- these "
                         "are LibreOffice values)")
    sr.add_argument("--engine-label",
                    default="SELFTEST (LibreOffice values via simulated export "
                            "- NOT Google Sheets)")
    sr.add_argument("--plain-names", dest="plain_names", action="store_true",
                    default=True, help=argparse.SUPPRESS)
    sr.add_argument("--xlfn-names", dest="plain_names", action="store_false",
                    help="dry-run the --xlfn-names build path instead")
    sr.add_argument("--allow-label-change", action="store_true", default=True,
                    help=argparse.SUPPRESS)
    sr.add_argument("--manifest-note", default=None, metavar="TEXT",
                    help=argparse.SUPPRESS)
    # cmd_selftest_recipes overrides this anyway; the default documents intent.
    sr.set_defaults(func=cmd_selftest_recipes, engine=LIBREOFFICE)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
